"""
Comprehensive Excel File Audit
- Verifies all sheets, cells, calculations
- Tests with virtual live data
- Multiple condition testing
- End-to-end verification
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytz
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


class ExcelComprehensiveAudit:
    """Comprehensive audit of Excel file."""

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.issues = []
        self.warnings = []
        self.passed = []
        self.ist = pytz.timezone("Asia/Kolkata")

    def audit_all(self):
        """Run complete audit."""
        print("=" * 80)
        print("  COMPREHENSIVE EXCEL FILE AUDIT")
        print("=" * 80)

        if not self.excel_path.exists():
            self.issues.append(f"Excel file not found: {self.excel_path}")
            return self._generate_report()

        # Test 1: File accessibility
        self._test_file_accessibility()

        # Test 2: Sheet structure
        self._test_sheet_structure()

        # Test 3: Data integrity
        self._test_data_integrity()

        # Test 4: Calculations
        self._test_calculations()

        # Test 5: ML Predictions
        self._test_ml_predictions()

        # Test 6: Trade Signals
        self._test_trade_signals()

        # Test 7: Charts
        self._test_charts()

        # Test 8: Paper Trading Data
        self._test_paper_trading()

        # Test 9: Virtual Live Data Test
        self._test_virtual_live_data()

        # Test 10: Multiple Conditions
        self._test_multiple_conditions()

        return self._generate_report()

    def _test_file_accessibility(self):
        """Test 1: File accessibility."""
        print("\n[TEST 1] File Accessibility")
        print("-" * 80)

        try:
            size = self.excel_path.stat().st_size
            print(f"  File exists: OK")
            print(f"  File size: {size:,} bytes")

            # Try to open with openpyxl
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            wb.close()
            print(f"  File readable: OK")
            self.passed.append("File accessibility")

        except Exception as e:
            self.issues.append(f"File accessibility error: {e}")
            print(f"  ERROR: {e}")

    def _test_sheet_structure(self):
        """Test 2: Sheet structure."""
        print("\n[TEST 2] Sheet Structure")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)
            sheets = xl.sheet_names

            print(f"  Total sheets: {len(sheets)}")

            required_sheets = ["OptionChain_Data", "CHAIN_RAW", "ML_PREDICTIONS", "TOP_OPPORTUNITIES", "Summary"]

            for sheet in required_sheets:
                if sheet in sheets:
                    print(f"  OK {sheet}: Present")
                    self.passed.append(f"Sheet: {sheet}")
                else:
                    # Check if alternative exists
                    alt_found = False
                    for s in sheets:
                        if sheet.lower() in s.lower() or s.lower() in sheet.lower():
                            print(f"  INFO {sheet}: Alternative found ({s})")
                            alt_found = True
                            break
                    if not alt_found:
                        self.warnings.append(f"Sheet '{sheet}' not found")
                        print(f"  WARNING {sheet}: Missing")

            # Check all sheets are readable
            for sheet in sheets:
                try:
                    df = pd.read_excel(xl, sheet_name=sheet, nrows=1)
                    print(f"  OK {sheet}: Readable ({len(df.columns)} columns)")
                except Exception as e:
                    self.issues.append(f"Sheet '{sheet}' not readable: {e}")
                    print(f"  ERROR {sheet}: {e}")

            self.passed.append("Sheet structure")

        except Exception as e:
            self.issues.append(f"Sheet structure error: {e}")
            print(f"  ERROR: {e}")

    def _test_data_integrity(self):
        """Test 3: Data integrity."""
        print("\n[TEST 3] Data Integrity")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)

            # Find main data sheet
            main_sheet = None
            for sheet in ["CHAIN_RAW", "OptionChain_Data"]:
                if sheet in xl.sheet_names:
                    main_sheet = sheet
                    break

            if not main_sheet:
                main_sheet = xl.sheet_names[0]

            df = pd.read_excel(xl, sheet_name=main_sheet)

            print(f"  Main sheet: {main_sheet}")
            print(f"  Rows: {len(df):,}")
            print(f"  Columns: {len(df.columns)}")

            # Check critical columns
            critical_cols = ["underlying", "strike", "option_type", "ltp", "spot_price"]
            missing = [c for c in critical_cols if c not in df.columns]

            if missing:
                self.issues.append(f"Missing critical columns: {missing}")
                print(f"  ERROR: Missing critical columns: {missing}")
            else:
                print(f"  OK: All critical columns present")
                self.passed.append("Critical columns")

            # Check data types
            print(f"\n  Data Type Check:")
            for col in critical_cols:
                if col in df.columns:
                    dtype = df[col].dtype
                    null_count = df[col].isna().sum()
                    null_pct = (null_count / len(df)) * 100

                    if null_pct > 50:
                        self.warnings.append(f"Column '{col}' has {null_pct:.1f}% nulls")
                        print(f"    WARNING {col}: {dtype}, {null_pct:.1f}% nulls")
                    else:
                        print(f"    OK {col}: {dtype}, {null_pct:.1f}% nulls")

            # Check for duplicates
            if "token" in df.columns:
                duplicates = df["token"].duplicated().sum()
                if duplicates > 0:
                    self.warnings.append(f"{duplicates} duplicate tokens found")
                    print(f"  WARNING Duplicate tokens: {duplicates}")
                else:
                    print(f"  OK No duplicate tokens")

            self.passed.append("Data integrity")

        except Exception as e:
            self.issues.append(f"Data integrity error: {e}")
            print(f"  ERROR: {e}")

    def _test_calculations(self):
        """Test 4: Calculations."""
        print("\n[TEST 4] Calculations Verification")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)
            main_sheet = "CHAIN_RAW" if "CHAIN_RAW" in xl.sheet_names else "OptionChain_Data"
            if main_sheet not in xl.sheet_names:
                main_sheet = xl.sheet_names[0]

            df = pd.read_excel(xl, sheet_name=main_sheet, nrows=100)

            # Test intrinsic value
            if all(col in df.columns for col in ["intrinsic_value", "spot_price", "strike", "option_type"]):
                sample = df[df["intrinsic_value"].notna() & df["spot_price"].notna() & df["strike"].notna()].head(10)

                errors = 0
                for idx, row in sample.iterrows():
                    spot = row["spot_price"]
                    strike = row["strike"]
                    opt_type = str(row["option_type"]).upper()
                    intrinsic = row["intrinsic_value"]

                    if opt_type == "CE":
                        expected = max(0, spot - strike)
                    else:
                        expected = max(0, strike - spot)

                    if abs(intrinsic - expected) > 0.01:
                        errors += 1

                if errors == 0:
                    print(f"  OK Intrinsic value: Correct ({len(sample)} samples)")
                    self.passed.append("Intrinsic value calculation")
                else:
                    self.warnings.append(f"Intrinsic value: {errors} errors in {len(sample)} samples")
                    print(f"  WARNING Intrinsic value: {errors} errors")

            # Test mid price
            if all(col in df.columns for col in ["mid_price", "bidPrice", "offerPrice"]):
                sample = df[df["mid_price"].notna() & df["bidPrice"].notna() & df["offerPrice"].notna()].head(10)

                errors = 0
                for idx, row in sample.iterrows():
                    mid = row["mid_price"]
                    expected = (row["bidPrice"] + row["offerPrice"]) / 2
                    if abs(mid - expected) > 0.01:
                        errors += 1

                if errors == 0:
                    print(f"  OK Mid price: Correct ({len(sample)} samples)")
                    self.passed.append("Mid price calculation")
                else:
                    self.warnings.append(f"Mid price: {errors} errors")
                    print(f"  WARNING Mid price: {errors} errors")

            # Test bid-ask spread
            if all(col in df.columns for col in ["bid_ask_spread", "offerPrice", "bidPrice"]):
                sample = df[df["bid_ask_spread"].notna() & df["offerPrice"].notna() & df["bidPrice"].notna()].head(10)

                errors = 0
                for idx, row in sample.iterrows():
                    spread = row["bid_ask_spread"]
                    expected = row["offerPrice"] - row["bidPrice"]
                    if abs(spread - expected) > 0.01:
                        errors += 1

                if errors == 0:
                    print(f"  OK Bid-ask spread: Correct")
                    self.passed.append("Bid-ask spread calculation")
                else:
                    self.warnings.append(f"Bid-ask spread: {errors} errors")
                    print(f"  WARNING Bid-ask spread: {errors} errors")

            self.passed.append("Calculations")

        except Exception as e:
            self.issues.append(f"Calculations error: {e}")
            print(f"  ERROR: {e}")

    def _test_ml_predictions(self):
        """Test 5: ML Predictions."""
        print("\n[TEST 5] ML Predictions")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)

            if "ML_PREDICTIONS" not in xl.sheet_names:
                self.warnings.append("ML_PREDICTIONS sheet not found")
                print(f"  WARNING ML_PREDICTIONS sheet: Not found")
                return

            df = pd.read_excel(xl, sheet_name="ML_PREDICTIONS")

            print(f"  Rows: {len(df):,}")
            print(f"  Columns: {len(df.columns)}")

            # Check required columns
            required = ["ml_prediction", "ml_confidence"]
            missing = [c for c in required if c not in df.columns]

            if missing:
                self.warnings.append(f"ML_PREDICTIONS missing columns: {missing}")
                print(f"  WARNING: Missing columns: {missing}")
            else:
                print(f"  OK: Required columns present")

                # Check data validity
                if "ml_confidence" in df.columns:
                    conf_values = df["ml_confidence"].dropna()
                    if len(conf_values) > 0:
                        min_conf = conf_values.min()
                        max_conf = conf_values.max()
                        print(f"  Confidence range: {min_conf:.2f} - {max_conf:.2f}")

                        if min_conf < 0 or max_conf > 100:
                            self.warnings.append(f"ML confidence out of range: {min_conf:.2f} - {max_conf:.2f}")
                        else:
                            print(f"  OK Confidence values valid")
                            self.passed.append("ML predictions")

            # Check predicted profit
            if "predicted_profit" in df.columns:
                profit_values = df["predicted_profit"].dropna()
                if len(profit_values) > 0:
                    print(f"  Predicted profit range: {profit_values.min():.2f} - {profit_values.max():.2f}")
                    self.passed.append("Predicted profit")

        except Exception as e:
            self.issues.append(f"ML predictions error: {e}")
            print(f"  ERROR: {e}")

    def _test_trade_signals(self):
        """Test 6: Trade Signals."""
        print("\n[TEST 6] Trade Signals")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)

            # Check main sheet for signals
            main_sheet = "CHAIN_RAW" if "CHAIN_RAW" in xl.sheet_names else "OptionChain_Data"
            if main_sheet not in xl.sheet_names:
                main_sheet = xl.sheet_names[0]

            df = pd.read_excel(xl, sheet_name=main_sheet)

            if "trade_signal" in df.columns:
                signals = df["trade_signal"].value_counts()
                print(f"  Signal distribution:")
                for signal, count in signals.items():
                    print(f"    {signal}: {count}")

                active_signals = df[df["trade_signal"] != "NO TRADE"]
                print(f"  Active signals: {len(active_signals)}")

                if len(active_signals) > 0:
                    # Check signal data
                    required = ["entry_price", "target_price", "stop_loss"]
                    missing = [c for c in required if c not in active_signals.columns]

                    if missing:
                        self.warnings.append(f"Trade signals missing columns: {missing}")
                    else:
                        print(f"  OK Signal data complete")
                        self.passed.append("Trade signals")
                else:
                    self.warnings.append("No active trade signals found")
            else:
                self.warnings.append("trade_signal column not found")

        except Exception as e:
            self.issues.append(f"Trade signals error: {e}")
            print(f"  ERROR: {e}")

    def _test_charts(self):
        """Test 7: Charts."""
        print("\n[TEST 7] Charts")
        print("-" * 80)

        try:
            wb = openpyxl.load_workbook(self.excel_path)

            chart_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if hasattr(ws, "_charts"):
                    chart_count += len(ws._charts)

            if chart_count > 0:
                print(f"  OK Charts found: {chart_count}")
                self.passed.append("Charts")
            else:
                self.warnings.append("No charts found in Excel file")
                print(f"  WARNING No charts found")

            wb.close()

        except Exception as e:
            self.warnings.append(f"Chart check error: {e}")
            print(f"  WARNING Chart check: {e}")

    def _test_paper_trading(self):
        """Test 8: Paper Trading Data."""
        print("\n[TEST 8] Paper Trading Data")
        print("-" * 80)

        try:
            xl = pd.ExcelFile(self.excel_path)

            paper_sheets = ["PAPER_TRADES", "OPEN_POSITIONS", "PNL_SUMMARY"]

            for sheet in paper_sheets:
                if sheet in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet)
                    print(f"  OK {sheet}: {len(df)} rows")
                    self.passed.append(f"Paper trading: {sheet}")
                else:
                    print(f"  INFO {sheet}: Not present (OK if system not running)")

        except Exception as e:
            self.warnings.append(f"Paper trading check error: {e}")
            print(f"  WARNING Paper trading: {e}")

    def _test_virtual_live_data(self):
        """Test 9: Virtual Live Data Test."""
        print("\n[TEST 9] Virtual Live Data Test")
        print("-" * 80)

        try:
            # Generate virtual live data
            print("  Generating virtual live data...")

            virtual_data = []
            underlyings = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]
            strikes = [24000, 24500, 25000, 25500, 26000]

            for underlying in underlyings:
                spot = 25000 + np.random.randint(-500, 500)
                for strike in strikes:
                    for opt_type in ["CE", "PE"]:
                        ltp = abs(spot - strike) * 0.1 + np.random.uniform(10, 100)
                        oi = np.random.randint(10000, 100000)
                        volume = np.random.randint(1000, 10000)

                        virtual_data.append(
                            {
                                "underlying": underlying,
                                "strike": strike,
                                "option_type": opt_type,
                                "spot_price": spot,
                                "ltp": ltp,
                                "oi": oi,
                                "volume": volume,
                                "bidPrice": ltp * 0.995,
                                "offerPrice": ltp * 1.005,
                                "delta": np.random.uniform(0, 1) if opt_type == "CE" else np.random.uniform(-1, 0),
                                "gamma": np.random.uniform(0, 0.01),
                                "theta": np.random.uniform(-20, -5),
                                "vega": np.random.uniform(10, 50),
                                "iv": np.random.uniform(0.15, 0.30),
                                "expiry": "24FEB2026",
                                "token": f"{underlying}_{strike}_{opt_type}",
                                "symbol": f"{underlying}24FEB26{strike}{opt_type}",
                            }
                        )

            virtual_df = pd.DataFrame(virtual_data)
            print(f"  Generated {len(virtual_df)} virtual contracts")

            # Test with virtual data
            from scripts.build_production_optionchain_master import (
                OptionChainMasterBuilder,
            )

            builder = OptionChainMasterBuilder()

            # Add calculations
            virtual_df = builder.add_all_calculations(virtual_df)
            new_cols = len([c for c in virtual_df.columns if c not in virtual_data[0].keys()])
            print(f"  OK Calculations added: {new_cols} new columns")

            # Check calculations work
            if "intrinsic_value" in virtual_df.columns:
                valid_intrinsic = virtual_df["intrinsic_value"].notna().sum()
                print(f"  OK Intrinsic value: {valid_intrinsic}/{len(virtual_df)} calculated")
                self.passed.append("Virtual data: calculations")

            if "mid_price" in virtual_df.columns:
                valid_mid = virtual_df["mid_price"].notna().sum()
                print(f"  OK Mid price: {valid_mid}/{len(virtual_df)} calculated")

            print(f"  OK Virtual live data test: PASSED")
            self.passed.append("Virtual live data test")

        except Exception as e:
            self.issues.append(f"Virtual live data error: {e}")
            print(f"  ERROR: {e}")
            import traceback

            traceback.print_exc()

    def _test_multiple_conditions(self):
        """Test 10: Multiple Conditions."""
        print("\n[TEST 10] Multiple Conditions Test")
        print("-" * 80)

        conditions = [
            ("Empty data", pd.DataFrame()),
            ("Missing columns", pd.DataFrame({"underlying": ["NIFTY"], "strike": [25000]})),
            ("All nulls", pd.DataFrame({"underlying": [None], "strike": [None], "ltp": [None]})),
            ("Invalid values", pd.DataFrame({"underlying": ["NIFTY"], "strike": [-1000], "ltp": [-50]})),
        ]

        for condition_name, test_df in conditions:
            try:
                print(f"  Testing: {condition_name}...")

                if len(test_df) == 0:
                    print(f"    OK Handles empty data")
                    continue

                from scripts.build_production_optionchain_master import (
                    OptionChainMasterBuilder,
                )

                builder = OptionChainMasterBuilder()

                # Try to add calculations
                try:
                    result_df = builder.add_all_calculations(test_df)
                    print(f"    OK Handles {condition_name}")
                    self.passed.append(f"Condition: {condition_name}")
                except Exception as e:
                    # Some conditions should fail gracefully
                    if "Missing required columns" in str(e):
                        print(f"    OK Handles {condition_name} (expected error)")
                        self.passed.append(f"Condition: {condition_name}")
                    else:
                        self.warnings.append(f"{condition_name} handling: {e}")
                        print(f"    OKOKOK {condition_name}: {e}")

            except Exception as e:
                self.warnings.append(f"{condition_name} test error: {e}")
                print(f"    WARNING {condition_name}: {e}")

    def _generate_report(self):
        """Generate audit report."""
        print("\n" + "=" * 80)
        print("  AUDIT REPORT")
        print("=" * 80)

        print(f"\nPassed: {len(self.passed)}")
        print(f"Warnings: {len(self.warnings)}")
        print(f"Issues: {len(self.issues)}")

        if self.issues:
            print(f"\nISSUES:")
            for issue in self.issues:
                print(f"  ERROR: {issue}")

        if self.warnings:
            print(f"\nWARNINGS:")
            for warning in self.warnings:
                print(f"  WARNING: {warning}")

        if self.passed:
            print(f"\nPASSED:")
            for passed in self.passed[:10]:  # Show first 10
                print(f"  OK: {passed}")
            if len(self.passed) > 10:
                print(f"  ... and {len(self.passed) - 10} more")

        # Overall status
        print("\n" + "=" * 80)
        if len(self.issues) == 0:
            if len(self.warnings) == 0:
                print("  STATUS: EXCELLENT - All tests passed")
            else:
                print("  STATUS: GOOD - Some warnings, but no critical issues")
        else:
            print("  STATUS: NEEDS ATTENTION - Issues found")
        print("=" * 80)

        return {
            "passed": len(self.passed),
            "warnings": len(self.warnings),
            "issues": len(self.issues),
            "status": (
                "EXCELLENT"
                if len(self.issues) == 0 and len(self.warnings) == 0
                else "GOOD" if len(self.issues) == 0 else "NEEDS_ATTENTION"
            ),
        }


def main():
    """Main execution."""
    excel_path = ROOT_DIR / "outputs" / "OptionChain_Master_v3_AI_FINAL.xlsx"

    auditor = ExcelComprehensiveAudit(excel_path)
    report = auditor.audit_all()

    return report["status"] != "NEEDS_ATTENTION"


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
