#!/usr/bin/env python3
"""SYSTEM3 GCP Micro MRI — one-shot full account clone (names/metadata only, no secret values).

Run:
  python scripts/system3_gcp_micro_mri_full.py

Outputs:
  System3_GCP_MICRO_MRI_FULL.xlsx          — agent-readable workbook (share this)
  reports/latest/gcp_micro_mri/MANIFEST.json
  reports/latest/gcp_micro_mri/latest/*.json — raw dumps
  reports/coordination/System3_GCP_MICRO_MRI_FULL.xlsx — copy for coordination

Never exports: secret values, API keys, tokens, private key material.
Always exports: env var NAMES, secret IDs, IAM bindings, URLs, resource metadata.
"""
from __future__ import annotations

import json
import re
import subprocess
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
PROJ = "system3-openalgo-safe"
REG = "asia-south1"
BASE = "https://genesis-system3-web-doq2wplepa-el.a.run.app"
XLSX = ROOT / "System3_GCP_MICRO_MRI_FULL.xlsx"
STAMP = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
OUT = ROOT / "reports" / "latest" / "gcp_micro_mri" / STAMP
OUT_LATEST = ROOT / "reports" / "latest" / "gcp_micro_mri" / "latest"
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

HF = PatternFill("solid", fgColor="1F4E79")
HFont = Font(color="FFFFFF", bold=True)
WARN = PatternFill("solid", fgColor="FFEB9C")
OKF = PatternFill("solid", fgColor="C6EFCE")
FAILF = PatternFill("solid", fgColor="FFC7CE")

SENSITIVE_NAME = re.compile(r"(token|secret|password|key|credential|pin|totp|api_key)", re.I)
SAFE_LITERAL_SUFFIX = re.compile(
    r"(SECRET_ID|TOKEN_SOURCE|TOKEN_CACHE|TOKEN_ROTATION|TOKEN_REFRESH|PERSIST_TOKEN|_SCHEDULE|_JOB|_TOPIC|_WAIT|_COOLDOWN|_PUBLISH|REQUIRE_API_KEY|SELF_HEAL)$",
    re.I,
)


def should_redact_literal(env_name: str, value: str) -> bool:
    """Redact only opaque secret values — never secret *IDs* or short config flags."""
    if not SENSITIVE_NAME.search(env_name):
        return False
    if SAFE_LITERAL_SUFFIX.search(env_name):
        return False
    naked = {"DHAN_PIN", "DHAN_TOTP", "DHAN_TOTP_SECRET", "DHAN_ACCESS_TOKEN", "API_KEY", "PASSWORD"}
    if env_name.upper() in naked:
        return True
    if len(value) > 48 or value.startswith("eyJ"):
        return True
    return False


def gcloud_bin() -> str:
    for c in (
        r"C:\Program Files (x86)\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        r"C:\Program Files\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        "gcloud",
    ):
        if c == "gcloud" or Path(c).exists():
            return c
    return "gcloud"


def run_gcloud(args: list[str], timeout: int = 90) -> tuple[Any, str | None]:
    exe = gcloud_bin()
    try:
        p = subprocess.run([exe, *args], capture_output=True, shell=False, timeout=timeout)
    except subprocess.TimeoutExpired:
        return None, f"timeout_after_{timeout}s"
    raw = p.stdout or b""
    err = (p.stderr or b"").decode("utf-8", "replace").strip() or None
    if p.returncode != 0 and not raw:
        return None, err or f"rc={p.returncode}"
    for enc in ("utf-8", "utf-16", "utf-16-le"):
        try:
            text = raw.decode(enc).lstrip("\ufeff").strip()
            break
        except UnicodeDecodeError:
            text = ""
    else:
        text = raw.decode("utf-8", "replace").strip()
    if not text:
        return ([] if p.returncode == 0 else None), err
    try:
        return json.loads(text), err
    except json.JSONDecodeError:
        return text, err


def save(name: str, data: Any) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    OUT_LATEST.mkdir(parents=True, exist_ok=True)
    text = json.dumps(data, indent=2) if not isinstance(data, str) else data
    (OUT / name).write_text(text, encoding="utf-8")
    (OUT_LATEST / name).write_text(text, encoding="utf-8")
    return OUT / name


class Collector:
    def __init__(self) -> None:
        self.coverage: list[dict[str, Any]] = []
        self.data: dict[str, Any] = {}

    def fetch(self, key: str, args: list[str], *, required: bool = False, timeout: int = 90) -> Any:
        print(f"  FETCH {key} ...", flush=True)
        data, err = run_gcloud(args, timeout=timeout)
        # Some list commands return empty list with warnings on stderr — still OK
        ok = data is not None
        if isinstance(data, dict) and data.get("error") and len(data) == 1:
            ok = False
        count = len(data) if isinstance(data, list) else (len(data.get("bindings", [])) if isinstance(data, dict) and "bindings" in data else (1 if data else 0))
        self.coverage.append({
            "key": key,
            "command": "gcloud " + " ".join(args[:8]) + ("..." if len(args) > 8 else ""),
            "ok": ok,
            "count": count,
            "error": (err or "")[:300],
            "required": required,
        })
        self.data[key] = data if data is not None else {"error": err}
        save(f"{key}.json", self.data[key])
        print(f"    -> ok={ok} count={count}", flush=True)
        return self.data[key]

    def http_live(self, key: str, path: str) -> Any:
        try:
            req = urllib.request.Request(f"{BASE}/{path}", headers={"User-Agent": "system3-gcp-micro-mri/1.0"})
            with urllib.request.urlopen(req, timeout=45) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            self.coverage.append({"key": key, "command": f"GET {BASE}/{path}", "ok": True, "count": 1, "error": "", "required": True})
            self.data[key] = data
            save(f"live_{key}.json", data)
            return data
        except Exception as exc:
            self.coverage.append({"key": key, "command": f"GET {BASE}/{path}", "ok": False, "count": 0, "error": str(exc), "required": True})
            self.data[key] = {"error": str(exc)}
            save(f"live_{key}.json", self.data[key])
            return self.data[key]


def dig(d: Any, *keys: str, default: Any = "") -> Any:
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur if cur is not None else default


def parse_env_rows(resource_type: str, resource_name: str, container_env: list[dict] | None) -> list[dict]:
    rows = []
    for e in container_env or []:
        name = e.get("name", "")
        if "valueFrom" in e:
            sk = dig(e, "valueFrom", "secretKeyRef", "name", default="")
            sk_key = dig(e, "valueFrom", "secretKeyRef", "key", default="latest")
            rows.append({
                "resource_type": resource_type,
                "resource_name": resource_name,
                "env_name": name,
                "source_type": "secretKeyRef",
                "secret_id": sk,
                "secret_version_key": sk_key,
                "literal_value": "***SECRET_REF***",
            })
        else:
            val = str(e.get("value", ""))
            if should_redact_literal(name, val):
                val = "***REDACTED***"
            rows.append({
                "resource_type": resource_type,
                "resource_name": resource_name,
                "env_name": name,
                "source_type": "literal",
                "secret_id": "",
                "secret_version_key": "",
                "literal_value": val,
            })
    return rows


def style_header(ws, n: int) -> None:
    for c in range(1, n + 1):
        cell = ws.cell(1, c)
        cell.fill = HF
        cell.font = HFont
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], widths: list[int]):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws, len(headers))
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 70)
    return ws


def load_repo_safe_env() -> list[tuple[str, str]]:
    impl = ROOT / "scripts" / "gcp_cloud_run_auto_deploy_impl.py"
    if not impl.exists():
        return []
    text = impl.read_text(encoding="utf-8")
    m = re.search(r"SAFE_ENV\s*=\s*\((.*?)\)\n\n", text, re.S)
    if not m:
        return []
    block = m.group(1)
    return re.findall(r'\("([^"]+)",\s*"([^"]*)"\)', block)


def collect_all() -> Collector:
    c = Collector()
    c.fetch("project", ["projects", "describe", PROJ, "--format=json"], required=True)
    c.fetch("billing", ["billing", "projects", "describe", PROJ, "--format=json"])
    c.fetch("iam_policy", ["projects", "get-iam-policy", PROJ, "--format=json"], required=True)
    c.fetch("iam_custom_roles", ["iam", "roles", "list", f"--project={PROJ}", "--format=json"])
    c.fetch("org_policies", ["resource-manager", "org-policies", "list", f"--project={PROJ}", "--format=json"])
    c.fetch("apis_enabled", ["services", "list", "--enabled", f"--project={PROJ}", "--format=json"])
    # Production region only (asia-south1). Multi-region scan is optional via env.
    import os
    regions = [REG]
    if os.environ.get("SYSTEM3_MRI_ALL_REGIONS", "").strip() in {"1", "true", "YES"}:
        rr = c.fetch("run_regions", ["run", "regions", "list", "--format=json"], timeout=60)
        if isinstance(rr, list):
            regions = sorted({dig(x, "locationId", default=REG) for x in rr if dig(x, "locationId")} | {REG})

    all_services: list[dict] = []
    all_jobs: list[dict] = []
    all_revisions: list[dict] = []
    for region in regions:
        print(f"REGION {region}", flush=True)
        svcs = c.fetch(f"run_services_{region}", ["run", "services", "list", f"--project={PROJ}", f"--region={region}", "--format=json"])
        if isinstance(svcs, list):
            for s in svcs:
                s["_region"] = region
            all_services.extend(svcs)
        jobs = c.fetch(f"run_jobs_{region}", ["run", "jobs", "list", f"--project={PROJ}", f"--region={region}", "--format=json"])
        if isinstance(jobs, list):
            for j in jobs:
                j["_region"] = region
            all_jobs.extend(jobs)
        for svc in (svcs or []) if isinstance(svcs, list) else []:
            sn = dig(svc, "metadata", "name")
            if sn:
                c.fetch(f"run_service_{sn}", ["run", "services", "describe", sn, f"--project={PROJ}", f"--region={region}", "--format=json"])
                revs = c.fetch(f"run_revisions_{sn}", ["run", "revisions", "list", f"--service={sn}", f"--project={PROJ}", f"--region={region}", "--limit=15", "--format=json"])
                if isinstance(revs, list):
                    for r in revs:
                        r["_service"] = sn
                        r["_region"] = region
                    all_revisions.extend(revs)
        for job in (jobs or []) if isinstance(jobs, list) else []:
            jn = dig(job, "metadata", "name")
            if jn:
                c.fetch(f"run_job_{jn}", ["run", "jobs", "describe", jn, f"--project={PROJ}", f"--region={region}", "--format=json"])
                c.fetch(f"run_job_exec_{jn}", ["run", "jobs", "executions", "list", f"--job={jn}", f"--project={PROJ}", f"--region={region}", "--limit=10", "--format=json"])

    c.data["run_services_all"] = all_services
    c.data["run_jobs_all"] = all_jobs
    c.data["run_revisions_all"] = all_revisions
    save("run_services_all.json", all_services)
    save("run_jobs_all.json", all_jobs)

    c.fetch("scheduler_jobs", ["scheduler", "jobs", "list", f"--project={PROJ}", f"--location={REG}", "--format=json"])
    for sj in (c.data.get("scheduler_jobs") or []) if isinstance(c.data.get("scheduler_jobs"), list) else []:
        name = str(sj.get("name", "")).split("/")[-1]
        if name:
            c.fetch(f"scheduler_{name}", ["scheduler", "jobs", "describe", name, f"--project={PROJ}", f"--location={REG}", "--format=json"])

    secrets = c.fetch("secrets", ["secrets", "list", f"--project={PROJ}", "--format=json"], required=True)
    secret_details: list[dict] = []
    if isinstance(secrets, list):
        for s in secrets:
            sid = str(dig(s, "name", default="")).split("/")[-1]
            if not sid:
                continue
            vers = c.fetch(f"secret_versions_{sid}", ["secrets", "versions", "list", sid, f"--project={PROJ}", "--limit=20", "--format=json"])
            iam = c.fetch(f"secret_iam_{sid}", ["secrets", "get-iam-policy", sid, f"--project={PROJ}", "--format=json"])
            secret_details.append({"secret_id": sid, "meta": s, "versions": vers, "iam": iam})
    c.data["secrets_micro"] = secret_details
    save("secrets_micro.json", secret_details)

    sas = c.fetch("service_accounts", ["iam", "service-accounts", "list", f"--project={PROJ}", "--format=json"], required=True)
    sa_micro: list[dict] = []
    if isinstance(sas, list):
        for sa in sas:
            email = sa.get("email", "")
            if not email:
                continue
            keys = c.fetch(f"sa_keys_{email.split('@')[0]}", ["iam", "service-accounts", "keys", "list", f"--iam-account={email}", "--format=json"])
            sa_iam = c.fetch(f"sa_iam_{email.split('@')[0]}", ["iam", "service-accounts", "get-iam-policy", email, "--format=json"])
            sa_micro.append({"email": email, "meta": sa, "keys": keys, "iam": sa_iam})
    c.data["service_accounts_micro"] = sa_micro
    save("service_accounts_micro.json", sa_micro)

    c.fetch("wif_pools", ["iam", "workload-identity-pools", "list", "--location=global", f"--project={PROJ}", "--format=json"])
    wif_micro: list[dict] = []
    for pool in (c.data.get("wif_pools") or []) if isinstance(c.data.get("wif_pools"), list) else []:
        pname = str(pool.get("name", "")).split("/")[-1]
        if pname:
            # Full pool resource path required for providers.list
            pool_path = pool.get("name") or f"projects/{PROJ}/locations/global/workloadIdentityPools/{pname}"
            prov = c.fetch(
                f"wif_providers_{pname}",
                ["iam", "workload-identity-pools", "providers", "list", f"--workload-identity-pool={pname}", "--location=global", f"--project={PROJ}", "--format=json"],
                timeout=60,
            )
            wif_micro.append({"pool": pool, "providers": prov, "pool_path": pool_path})
    c.data["wif_micro"] = wif_micro
    save("wif_micro.json", wif_micro)

    c.fetch("pubsub_topics", ["pubsub", "topics", "list", f"--project={PROJ}", "--format=json"])
    subs: list[dict] = []
    for topic in (c.data.get("pubsub_topics") or []) if isinstance(c.data.get("pubsub_topics"), list) else []:
        tname = topic if isinstance(topic, str) else topic.get("name", "")
        short = str(tname).split("/")[-1]
        if short:
            s = c.fetch(f"pubsub_subs_{short}", ["pubsub", "subscriptions", "list", f"--project={PROJ}", f"--filter=topic:{tname}", "--format=json"])
            if isinstance(s, list):
                subs.extend(s)
    c.data["pubsub_subscriptions"] = subs
    save("pubsub_subscriptions.json", subs)

    c.fetch("buckets", ["storage", "buckets", "list", f"--project={PROJ}", "--format=json"])
    bucket_micro: list[dict] = []
    for b in (c.data.get("buckets") or []) if isinstance(c.data.get("buckets"), list) else []:
        bname = b.get("name") if isinstance(b, dict) else str(b)
        if bname:
            detail = c.fetch(f"bucket_{bname.replace('.', '_')}", ["storage", "buckets", "describe", f"gs://{bname}", "--format=json"])
            bucket_micro.append({"name": bname, "detail": detail})
    c.data["buckets_micro"] = bucket_micro

    c.fetch("artifact_repos", ["artifacts", "repositories", "list", f"--project={PROJ}", f"--location={REG}", "--format=json"])
    for repo in (c.data.get("artifact_repos") or []) if isinstance(c.data.get("artifact_repos"), list) else []:
        rname = str(dig(repo, "name", default="")).split("/")[-1]
        if rname:
            c.fetch(f"artifact_images_{rname}", ["artifacts", "docker", "images", "list", f"{REG}-docker.pkg.dev/{PROJ}/{rname}", "--include-tags", "--format=json"], timeout=120)

    c.fetch("build_triggers", ["builds", "triggers", "list", f"--project={PROJ}", "--format=json"])
    c.fetch("builds_recent", ["builds", "list", f"--project={PROJ}", "--limit=20", "--format=json"])

    c.fetch("firestore_databases", ["firestore", "databases", "list", f"--project={PROJ}", "--format=json"])
    c.fetch("logging_sinks", ["logging", "sinks", "list", f"--project={PROJ}", "--format=json"])

    # Optional / often-empty surfaces — short timeout, never block MRI
    for key, args in [
        ("monitoring_policies", ["monitoring", "policies", "list", f"--project={PROJ}", "--format=json"]),
        ("uptime_configs", ["monitoring", "uptime", "list-configs", f"--project={PROJ}", "--format=json"]),
        ("compute_instances", ["compute", "instances", "list", f"--project={PROJ}", "--format=json"]),
        ("sql_instances", ["sql", "instances", "list", f"--project={PROJ}", "--format=json"]),
        ("functions", ["functions", "list", f"--project={PROJ}", f"--region={REG}", "--format=json"]),
        ("eventarc_triggers", ["eventarc", "triggers", "list", f"--project={PROJ}", f"--location={REG}", "--format=json"]),
        ("dns_zones", ["dns", "managed-zones", "list", f"--project={PROJ}", "--format=json"]),
        ("bq_datasets", ["bq", "ls", "--project_id", PROJ, "--format=json"]),
    ]:
        c.fetch(key, args, timeout=45)

    for path in ("api/deploy_info", "api/broker/status", "api/health", "api/auto_gates", "api/scheduler/health", "api/system_health"):
        key = "live_" + path.replace("/", "_")
        c.http_live(key, path)

    try:
        main_sha = subprocess.check_output(["git", "rev-parse", "origin/main"], cwd=str(ROOT), text=True).strip()
    except Exception:
        main_sha = "UNKNOWN"
    c.data["github_origin_main"] = main_sha
    save("github_origin_main.json", {"sha": main_sha})

    return c


def build_workbook(c: Collector) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # 00 Agent README
    readme = [
        ["SYSTEM3 GCP MICRO MRI — agent guide", ""],
        ["Generated UTC", NOW],
        ["Project", PROJ],
        ["Primary region", REG],
        ["Workbook", str(XLSX)],
        ["Raw JSON dir", str(OUT_LATEST)],
        ["", ""],
        ["How to inspect", "Start sheet 01_Executive → 02_Coverage → 04_Env_Vars_Micro → 05_Secrets_Micro"],
        ["Env vars", "Sheet 04 lists every Cloud Run env NAME + secret_id mapping (values redacted)"],
        ["Secrets", "Sheet 05 lists every Secret Manager secret + versions + IAM (no values)"],
        ["Gaps", "Sheet 03 lists what cannot be exported (secret values, full billing line items)"],
        ["Live truth", "Sheet 25_Live_APIs — same-session production JSON"],
        ["Repo authority", "Sheet 26_Repo_SAFE_ENV vs sheet 27 live diff"],
        ["Share with agent", "Send this xlsx + MANIFEST.json; never send secret value files"],
    ]
    ws = wb.create_sheet("00_Agent_README")
    for row in readme:
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    # Coverage
    cov_rows = [[x["key"], x["ok"], x["count"], x["required"], x["error"][:200], x["command"][:120]] for x in c.coverage]
    fails = sum(1 for x in c.coverage if not x["ok"])
    write_sheet(wb, "02_Coverage_Verification", ["key", "ok", "count", "required", "error", "command"], cov_rows, [35, 8, 8, 10, 50, 80])

    # Gaps honest list
    gaps = [
        ["NOT exported (by design)", "Secret Manager secret VALUES / token contents", "Use secret_id + version metadata only"],
        ["NOT exported (by design)", "Service account private key PEM content", "Keys list shows key_id + disabled only"],
        ["NOT exported (by design)", "User passwords / OAuth refresh tokens", ""],
        ["Partial", "Full billing invoice line-item CSV", "Console billing estimate only unless Billing Export enabled"],
        ["Partial", "Historical Cloud Logging full stream", "Use Logging console / query; sinks listed"],
        ["Partial", "Browser F12 HAR unless captured separately", "Run with browser open for network HAR"],
        ["Partial", "Org-level resources outside project", "This scan is project-scoped"],
        ["Verification", f"Coverage commands: {len(c.coverage)}", f"Failed: {fails}"],
    ]
    write_sheet(wb, "03_Gaps_And_Limits", ["category", "item", "note"], gaps, [18, 55, 60])

    # Env vars micro — from all described services/jobs
    env_rows: list[list[Any]] = []
    for key, val in c.data.items():
        if not key.startswith("run_service_") and not key.startswith("run_job_"):
            continue
        if not isinstance(val, dict) or val.get("error"):
            continue
        rtype = "service" if key.startswith("run_service_") else "job"
        rname = dig(val, "metadata", "name")
        if rtype == "service":
            containers = dig(val, "spec", "template", "spec", "containers", default=[]) or []
            envs = (containers[0] if containers else {}).get("env", [])
        else:
            # Cloud Run Jobs: spec.template.spec.template.spec.containers[].env
            containers = dig(val, "spec", "template", "spec", "template", "spec", "containers", default=[]) or []
            if not containers:
                containers = dig(val, "spec", "template", "template", "spec", "containers", default=[]) or []
            envs = (containers[0] if containers else {}).get("env", [])
        for row in parse_env_rows(rtype, rname, envs):
            env_rows.append([row["resource_type"], row["resource_name"], row["env_name"], row["source_type"], row["secret_id"], row["secret_version_key"], row["literal_value"]])
    write_sheet(wb, "04_Env_Vars_Micro", ["resource_type", "resource_name", "env_name", "source_type", "secret_id", "secret_version_key", "literal_or_redacted"], env_rows, [12, 40, 40, 14, 40, 14, 50])

    # Secret to env map
    sem_rows = [[r[2], r[4], r[1], r[0]] for r in env_rows if r[3] == "secretKeyRef" and r[4]]
    write_sheet(wb, "05_Secret_To_Env_Map", ["env_name", "secret_id", "resource_name", "resource_type"], sem_rows, [40, 40, 40, 12])

    # Secrets micro
    sec_rows: list[list[Any]] = []
    for sd in c.data.get("secrets_micro") or []:
        sid = sd.get("secret_id", "")
        meta = sd.get("meta") or {}
        for v in (sd.get("versions") or []) if isinstance(sd.get("versions"), list) else []:
            sec_rows.append([sid, str(dig(v, "name", default="")).split("/")[-1], v.get("state"), v.get("createTime"), v.get("destroyTime"), json.dumps(meta.get("labels") or {})])
        for b in dig(sd.get("iam"), "bindings", default=[]) or []:
            for m in b.get("members") or []:
                sec_rows.append([sid, "IAM", b.get("role"), m, "", ""])
    write_sheet(wb, "06_Secrets_Micro", ["secret_id", "version_or_iam", "state_or_role", "create_or_member", "destroy", "labels"], sec_rows, [40, 20, 40, 50, 22, 30])

    # IAM bindings
    iam_rows = []
    for b in dig(c.data.get("iam_policy"), "bindings", default=[]) or []:
        for m in b.get("members") or []:
            iam_rows.append([b.get("role"), m, m.split(":")[0] if ":" in m else ""])
    write_sheet(wb, "14_IAM_Project_Bindings", ["role", "member", "member_type"], iam_rows, [55, 70, 16])

    # SA micro
    sa_rows = []
    key_rows = []
    for sa in c.data.get("service_accounts_micro") or []:
        email = sa.get("email", "")
        meta = sa.get("meta") or {}
        sa_rows.append([email, meta.get("displayName"), meta.get("uniqueId"), meta.get("disabled"), meta.get("oauth2ClientId")])
        for k in (sa.get("keys") or []) if isinstance(sa.get("keys"), list) else []:
            key_rows.append([email, k.get("name", "").split("/")[-1], k.get("keyType"), k.get("validAfterTime"), k.get("validBeforeTime"), k.get("disabled")])
        for b in dig(sa.get("iam"), "bindings", default=[]) or []:
            for m in b.get("members") or []:
                sa_rows.append([email, f"IAM:{b.get('role')}", m, "", ""])
    write_sheet(wb, "11_Service_Accounts_Micro", ["email", "display_or_iam", "unique_or_member", "disabled", "oauth2"], sa_rows, [55, 40, 40, 10, 20])
    write_sheet(wb, "12_SA_Keys_Metadata", ["email", "key_id", "keyType", "validAfter", "validBefore", "disabled"], key_rows, [55, 20, 14, 22, 22, 10])

    # Cloud Run services all
    svc_rows = []
    for s in c.data.get("run_services_all") or []:
        svc_rows.append([
            s.get("_region"), dig(s, "metadata", "name"), dig(s, "status", "url"),
            dig(s, "status", "latestReadyRevisionName"), dig(s, "spec", "template", "spec", "serviceAccountName"),
            dig(s, "spec", "template", "spec", "containers", default=[{}])[0].get("image", "") if dig(s, "spec", "template", "spec", "containers") else "",
        ])
    write_sheet(wb, "07_CloudRun_Services", ["region", "name", "url", "latestReadyRevision", "serviceAccount", "image"], svc_rows, [14, 30, 55, 40, 45, 70])

    # Jobs
    job_rows = []
    exec_rows = []
    for j in c.data.get("run_jobs_all") or []:
        jn = dig(j, "metadata", "name")
        job_rows.append([j.get("_region"), jn, dig(j, "status", "executionCount"), dig(j, "spec", "template", "template", "spec", "serviceAccountName")])
        ex = c.data.get(f"run_job_exec_{jn}")
        if isinstance(ex, list):
            for e in ex:
                exec_rows.append([jn, dig(e, "metadata", "name"), dig(e, "status", "completionTime"), dig(e, "status", "conditions", default=[{}])[0].get("status") if dig(e, "status", "conditions") else ""])
    write_sheet(wb, "08_CloudRun_Jobs", ["region", "job", "executionCount", "serviceAccount"], job_rows, [14, 40, 12, 45])
    write_sheet(wb, "09_Job_Executions", ["job", "execution", "completionTime", "status"], exec_rows, [40, 45, 24, 12])

    # Scheduler full
    sch_rows = []
    for sj in (c.data.get("scheduler_jobs") or []) if isinstance(c.data.get("scheduler_jobs"), list) else []:
        name = str(sj.get("name", "")).split("/")[-1]
        full = c.data.get(f"scheduler_{name}") or sj
        http = full.get("httpTarget") or {}
        sch_rows.append([
            name, full.get("schedule"), full.get("timeZone"), full.get("state"),
            http.get("uri") or dig(full, "pubsubTarget", "topicName"),
            dig(http, "oauthToken", "serviceAccountEmail"),
            dig(http, "httpMethod"), full.get("lastAttemptTime"),
        ])
    write_sheet(wb, "10_Scheduler_Full", ["name", "schedule", "timezone", "state", "target", "oauth_sa", "method", "lastAttempt"], sch_rows, [40, 18, 16, 12, 70, 45, 10, 24])

    # WIF
    wif_rows = []
    for wm in c.data.get("wif_micro") or []:
        pool = wm.get("pool") or {}
        pname = str(pool.get("name", "")).split("/")[-1]
        wif_rows.append(["pool", pname, pool.get("state"), pool.get("description"), ""])
        for p in (wm.get("providers") or []) if isinstance(wm.get("providers"), list) else []:
            wif_rows.append(["provider", str(p.get("name", "")).split("/")[-1], p.get("state"), p.get("displayName"), json.dumps(p.get("oidc") or p.get("aws") or {})[:200]])
    write_sheet(wb, "15_WIF_Pools_Providers", ["kind", "name", "state", "description", "oidc_aws"], wif_rows, [12, 40, 12, 40, 60])

    # PubSub, buckets, APIs, etc.
    ps_rows = [[t if isinstance(t, str) else t.get("name")] for t in (c.data.get("pubsub_topics") or [])]
    write_sheet(wb, "16_PubSub_Topics", ["topic"], ps_rows, [80])
    sub_rows = [[s.get("name"), s.get("topic")] for s in (c.data.get("pubsub_subscriptions") or []) if isinstance(s, dict)]
    write_sheet(wb, "17_PubSub_Subscriptions", ["subscription", "topic"], sub_rows, [60, 60])

    bkt_rows = []
    for bm in c.data.get("buckets_micro") or []:
        d = bm.get("detail") or {}
        bkt_rows.append([bm.get("name"), d.get("location"), d.get("storageClass"), d.get("versioning", {}).get("enabled"), d.get("lifecycle", {}).get("rule")])
    write_sheet(wb, "18_Storage_Buckets", ["bucket", "location", "class", "versioning", "lifecycle_rules"], bkt_rows, [40, 14, 14, 10, 40])

    api_rows = [[a.get("config", {}).get("name") or a.get("name"), a.get("state")] for a in (c.data.get("apis_enabled") or []) if isinstance(a, dict)]
    write_sheet(wb, "23_APIs_Enabled", ["api", "state"], api_rows, [60, 12])

    # Repo SAFE_ENV vs live
    repo_env = load_repo_safe_env()
    live_env = {r[2]: r[6] for r in env_rows if r[0] == "service" and r[1] == "genesis-system3-web" and r[3] == "literal"}
    repo_rows = [[k, v, live_env.get(k, "MISSING_ON_LIVE"), "MATCH" if live_env.get(k) == v else "DIFF"] for k, v in repo_env]
    write_sheet(wb, "26_Repo_SAFE_ENV", ["env_name", "repo_default", "live_literal", "match"], repo_rows, [40, 40, 40, 10])

    diff_rows = []
    live_names = {r[2] for r in env_rows if r[0] == "service" and r[1] == "genesis-system3-web"}
    repo_names = {k for k, _ in repo_env}
    for n in sorted(live_names - repo_names):
        diff_rows.append(["live_only", n])
    for n in sorted(repo_names - live_names):
        diff_rows.append(["repo_only", n])
    write_sheet(wb, "27_Live_vs_Repo_Env_Diff", ["kind", "env_name"], diff_rows, [12, 40])

    # Live APIs flattened
    live_rows = []
    for key, payload in c.data.items():
        if not key.startswith("live_api_"):
            continue
        if isinstance(payload, dict):
            for k, v in payload.items():
                if isinstance(v, (dict, list)):
                    val = json.dumps(v)[:1500]
                else:
                    val = "***REDACTED***" if SENSITIVE_NAME.search(str(k)) else str(v)
                live_rows.append([key, k, val])
    write_sheet(wb, "25_Live_APIs", ["surface", "key", "value"], live_rows, [25, 35, 100])

    # URL catalog
    urls = [
        ("P0", "GCP", "Home", f"https://console.cloud.google.com/home/dashboard?project={PROJ}"),
        ("P0", "GCP", "IAM", f"https://console.cloud.google.com/iam-admin/iam?project={PROJ}"),
        ("P0", "GCP", "Service accounts", f"https://console.cloud.google.com/iam-admin/serviceaccounts?project={PROJ}"),
        ("P0", "GCP", "Secret Manager", f"https://console.cloud.google.com/security/secret-manager?project={PROJ}"),
        ("P0", "GCP", "Cloud Run web", f"https://console.cloud.google.com/run/detail/{REG}/genesis-system3-web/metrics?project={PROJ}"),
        ("P0", "GCP", "Rotate job", f"https://console.cloud.google.com/run/jobs/details/{REG}/genesis-system3-dhan-token-rotate?project={PROJ}"),
        ("P0", "GCP", "Scheduler", f"https://console.cloud.google.com/cloudscheduler?project={PROJ}"),
        ("P0", "GCP", "Logging", f"https://console.cloud.google.com/logs/query?project={PROJ}"),
        ("P0", "GCP", "Error Reporting", f"https://console.cloud.google.com/errors?project={PROJ}"),
        ("P0", "GCP", "Billing", f"https://console.cloud.google.com/billing?project={PROJ}"),
        ("P0", "GCP", "WIF", f"https://console.cloud.google.com/iam-admin/workload-identity-pools?project={PROJ}"),
        ("P0", "Live", "UI", f"{BASE}/ui/"),
        ("P0", "Live", "deploy_info", f"{BASE}/api/deploy_info"),
        ("P0", "GitHub", "repo", "https://github.com/psw2025-cmd/Genesis_System3"),
        ("P0", "GitHub", "#188", "https://github.com/psw2025-cmd/Genesis_System3/issues/188"),
    ]
    write_sheet(wb, "28_URL_Catalog", ["priority", "surface", "name", "url"], [[*u, NOW] for u in urls], [8, 10, 25, 90])

    # Executive
    dep = c.data.get("live_api_deploy_info") or {}
    br = c.data.get("live_api_broker_status") or {}
    hl = c.data.get("live_api_health") or {}
    gt = c.data.get("live_api_auto_gates") or {}
    exec_rows = [
        ["extracted_utc", NOW],
        ["project", PROJ],
        ["region", REG],
        ["github_main", c.data.get("github_origin_main")],
        ["serving_sha", dep.get("git_sha") if isinstance(dep, dict) else ""],
        ["main_eq_serving", str((dep.get("git_sha") if isinstance(dep, dict) else None) == c.data.get("github_origin_main"))],
        ["live_trading", str(dep.get("live_trading_enabled") if isinstance(dep, dict) else "")],
        ["broker_connected", str(br.get("connected") if isinstance(br, dict) else "")],
        ["broker_auth", str(br.get("auth_classification") if isinstance(br, dict) else "")],
        ["health_status", str(hl.get("status") if isinstance(hl, dict) else "")],
        ["health_qc", str(hl.get("qc_status") if isinstance(hl, dict) else "")],
        ["gates", f"{gt.get('gates_passing')}/{gt.get('gates_total')}" if isinstance(gt, dict) else ""],
        ["env_vars_count", str(len(env_rows))],
        ["secrets_count", str(len(c.data.get("secrets_micro") or []))],
        ["service_accounts_count", str(len(c.data.get("service_accounts_micro") or []))],
        ["iam_bindings", str(len(iam_rows))],
        ["coverage_commands", str(len(c.coverage))],
        ["coverage_failed", str(fails)],
        ["workbook", str(XLSX)],
        ["raw_json", str(OUT_LATEST)],
    ]
    write_sheet(wb, "01_Executive", ["field", "value"], exec_rows, [28, 90])

    # Raw index
    idx = [[p.name, p.stat().st_size] for p in sorted(OUT_LATEST.glob("*")) if p.is_file()]
    write_sheet(wb, "29_Raw_Artifact_Index", ["file", "bytes"], idx, [50, 12])

    return wb


def main() -> int:
    import argparse
    import sys

    ap = argparse.ArgumentParser(description="SYSTEM3 GCP Micro MRI")
    ap.add_argument("--rebuild-from-latest", action="store_true", help="Rebuild xlsx from reports/latest/gcp_micro_mri/latest JSON only")
    args = ap.parse_args()

    if args.rebuild_from_latest:
        print("SYSTEM3 GCP Micro MRI — rebuild from latest JSON...", flush=True)
        c = Collector()
        for p in sorted(OUT_LATEST.glob("*.json")):
            try:
                c.data[p.stem] = json.loads(p.read_text(encoding="utf-8"))
            except Exception as exc:
                c.coverage.append({"key": p.stem, "command": "load", "ok": False, "count": 0, "error": str(exc), "required": False})
                continue
            c.coverage.append({"key": p.stem, "command": "load_cached", "ok": True, "count": 1, "error": "", "required": False})
        # reconstruct aggregates
        c.data["run_services_all"] = c.data.get("run_services_all") or c.data.get(f"run_services_{REG}") or []
        c.data["run_jobs_all"] = c.data.get("run_jobs_all") or c.data.get(f"run_jobs_{REG}") or []
        if isinstance(c.data.get("run_services_all"), list):
            for s in c.data["run_services_all"]:
                if isinstance(s, dict) and "_region" not in s:
                    s["_region"] = REG
        # secrets_micro / sa micro may already exist
        if "secrets_micro" not in c.data and isinstance(c.data.get("secrets"), list):
            c.data["secrets_micro"] = []
            for s in c.data["secrets"]:
                sid = str(dig(s, "name", default="")).split("/")[-1]
                c.data["secrets_micro"].append({
                    "secret_id": sid,
                    "meta": s,
                    "versions": c.data.get(f"secret_versions_{sid}"),
                    "iam": c.data.get(f"secret_iam_{sid}"),
                })
        if "service_accounts_micro" not in c.data and isinstance(c.data.get("service_accounts"), list):
            c.data["service_accounts_micro"] = []
            for sa in c.data["service_accounts"]:
                email = sa.get("email", "")
                short = email.split("@")[0]
                c.data["service_accounts_micro"].append({
                    "email": email,
                    "meta": sa,
                    "keys": c.data.get(f"sa_keys_{short}"),
                    "iam": c.data.get(f"sa_iam_{short}"),
                })
        if "wif_micro" not in c.data:
            c.data["wif_micro"] = []
            for pool in (c.data.get("wif_pools") or []) if isinstance(c.data.get("wif_pools"), list) else []:
                pname = str(pool.get("name", "")).split("/")[-1]
                c.data["wif_micro"].append({"pool": pool, "providers": c.data.get(f"wif_providers_{pname}")})
        if "buckets_micro" not in c.data:
            c.data["buckets_micro"] = []
            for b in (c.data.get("buckets") or []) if isinstance(c.data.get("buckets"), list) else []:
                bname = b.get("name") if isinstance(b, dict) else str(b)
                c.data["buckets_micro"].append({"name": bname, "detail": c.data.get(f"bucket_{str(bname).replace('.', '_')}")})
        if "github_origin_main" not in c.data:
            try:
                c.data["github_origin_main"] = subprocess.check_output(["git", "rev-parse", "origin/main"], cwd=str(ROOT), text=True).strip()
            except Exception:
                c.data["github_origin_main"] = "UNKNOWN"
        elif isinstance(c.data["github_origin_main"], dict):
            c.data["github_origin_main"] = c.data["github_origin_main"].get("sha", "UNKNOWN")
    else:
        print("SYSTEM3 GCP Micro MRI — collecting...", flush=True)
        c = collect_all()

    wb = build_workbook(c)
    try:
        wb.save(XLSX)
        out_path = XLSX
    except PermissionError:
        alt = ROOT / f"System3_GCP_MICRO_MRI_FULL_{STAMP}.xlsx"
        wb.save(alt)
        out_path = alt
        print(f"WARN: primary xlsx locked; wrote {alt}", flush=True)
    copy = ROOT / "reports" / "coordination" / "System3_GCP_MICRO_MRI_FULL.xlsx"
    copy.parent.mkdir(parents=True, exist_ok=True)
    try:
        copy.write_bytes(out_path.read_bytes())
    except Exception as exc:
        print(f"WARN copy: {exc}", flush=True)

    manifest = {
        "schema": "system3_gcp_micro_mri_v1",
        "generated_utc": NOW,
        "project": PROJ,
        "workbook": str(out_path),
        "raw_dir": str(OUT_LATEST),
        "coverage_total": len(c.coverage),
        "coverage_failed": sum(1 for x in c.coverage if not x["ok"]),
        "sheets": wb.sheetnames,
        "mode": "rebuild_from_latest" if args.rebuild_from_latest else "full_collect",
        "counts": {
            "secrets": len(c.data.get("secrets_micro") or []),
            "service_accounts": len(c.data.get("service_accounts_micro") or []),
            "run_services": len(c.data.get("run_services_all") or []),
            "run_jobs": len(c.data.get("run_jobs_all") or []),
        },
    }
    save_path = ROOT / "reports" / "latest" / "gcp_micro_mri" / "MANIFEST.json"
    save_path.parent.mkdir(parents=True, exist_ok=True)
    save_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
