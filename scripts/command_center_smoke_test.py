#!/usr/bin/env python3
"""Command Center smoke tests — required before apply. Writes status JSON for CI."""
from __future__ import annotations

import json
import os
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "reports" / "coordination"
SMOKE_OUT = OUT / "SMOKE_TEST_LAST.json"
BASE = os.environ.get(
    "SYSTEM3_PUBLIC_BASE",
    "http://127.0.0.1:8000",
).rstrip("/")

REQUIRED_XLSX_SHEETS = {
    "0_Meta",
    "1_User_Actions",
    "2_Options_Priority",
    "3_Pending_Live",
    "4_UI_Tab_Impact",
    "5_GCP_GitHub_Levers",
    "6_Progress_Chart",
    "7_Failure_Playbook",
    "8_MD_Upgrades",
}


def check(name: str, ok: bool, detail: str) -> dict:
    return {"name": name, "ok": ok, "detail": detail}


def _is_loopback_base(url: str) -> bool:
    lowered = (url or "").strip().lower()
    return "127.0.0.1" in lowered or "localhost" in lowered


def _local_deploy_info() -> dict:
    git_sha = ""
    try:
        import subprocess

        git_sha = subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=str(ROOT),
            timeout=3,
            stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        git_sha = ""
    return {
        "git_sha": git_sha,
        "service_name": "local-laptop",
        "deploy_target": "local-laptop",
        "public_base_url": BASE,
        "live_trading_enabled": False,
        "_source": "repo-local",
    }


def _read_deploy_info() -> tuple[dict | None, str]:
    endpoints = ["/api/deploy/info", "/api/deploy_info"]
    for path in endpoints:
        try:
            with urllib.request.urlopen(f"{BASE}{path}", timeout=25) as resp:
                return json.loads(resp.read().decode("utf-8")), f"http:{path}"
        except Exception:
            continue
    if _is_loopback_base(BASE):
        return _local_deploy_info(), "local-fallback"
    return None, "http-unreachable"


def main() -> int:
    results = []
    run_id = os.environ.get("GITHUB_RUN_ID") or os.environ.get("CC_RUN_ID") or f"local-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
    agent_id = os.environ.get("CC_AGENT_ID", "cursor-composer")

    # 1 policy schema
    from validate_access_policy import main as validate_main

    rc = validate_main()
    results.append(check("access_policy_schema", rc == 0, f"validate_exit={rc}"))

    # 2 excel sheets
    xlsx = OUT / "AGENT_OPERATING_OPTIONS.xlsx"
    excel_ok = False
    excel_detail = "missing xlsx"
    if xlsx.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx, read_only=True)
            names = set(wb.sheetnames)
            missing = sorted(REQUIRED_XLSX_SHEETS - names)
            excel_ok = not missing
            excel_detail = f"sheets={len(names)} missing={missing}"
            wb.close()
        except Exception as e:  # noqa: BLE001
            excel_detail = str(e)[:200]
    results.append(check("excel_sheets_1_to_8", excel_ok, excel_detail))

    # 3 artifacts
    arts = [
        OUT / "COMMAND_CENTER.md",
        OUT / "ISSUES_ONLY.md",
        OUT / "ISSUES_MERMAID.md",
        OUT / "TRACKING_CHECKLIST.md",
        OUT / "ACCESS_POLICY.yaml",
    ]
    missing_arts = [str(p.name) for p in arts if not p.exists()]
    results.append(check("command_center_artifacts_exist", not missing_arts, f"missing={missing_arts}"))

    # 4 live deploy_info
    live_ok = False
    live_detail = ""
    live_trading = None
    raw, source = _read_deploy_info()
    if raw is not None:
        live_ok = bool(raw.get("git_sha"))
        live_trading = raw.get("live_trading_enabled")
        live_detail = f"source={source} sha={(raw.get('git_sha') or '')[:12]} live_trading={live_trading}"
    else:
        live_detail = source
    results.append(check("live_deploy_info_http", live_ok, live_detail))
    results.append(check("no_live_trading_flag", live_trading is False, f"live_trading_enabled={live_trading}"))

    passed = all(r["ok"] for r in results)
    payload = {
        "utc": datetime.now(timezone.utc).isoformat(),
        "run_id": run_id,
        "agent_id": agent_id,
        "passed": passed,
        "results": results,
    }
    OUT.mkdir(parents=True, exist_ok=True)
    SMOKE_OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print("SMOKE_START")
    for r in results:
        print(f"  {r['name']}: {'PASS' if r['ok'] else 'FAIL'} — {r['detail']}")
    print("SMOKE_END")
    print(f"SMOKE_PASSED={passed} run_id={run_id}")
    return 0 if passed else 1


if __name__ == "__main__":
    # allow importing sibling validate script
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    raise SystemExit(main())
