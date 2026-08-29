#!/usr/bin/env python3
"""Rank System3 coordination CSV/XLSX authorities across laptop copies + live SHAs.

Writes:
  - reports/coordination/AUTHORITY_RANK_ANALYSIS.csv
  - sheet 10_Authority_Rank on reports/coordination/AGENT_OPERATING_OPTIONS.xlsx
"""

from __future__ import annotations

import csv
import json
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _http_json(url: str, timeout: float = 25.0) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "system3-authority-rank/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def live_snapshot() -> dict:
    base = "https://genesis-system3-web-doq2wplepa-el.a.run.app"
    out = {
        "serving_sha": "UNKNOWN",
        "broker": "UNKNOWN",
        "health": "UNKNOWN",
        "gates": "UNKNOWN",
    }
    try:
        d = _http_json(f"{base}/api/deploy_info")
        out["serving_sha"] = d.get("git_sha") or "UNKNOWN"
    except Exception as exc:  # noqa: BLE001
        out["serving_sha"] = f"ERR:{exc}"
    try:
        b = _http_json(f"{base}/api/broker/status")
        out["broker"] = (
            f"connected={b.get('connected')} auth={b.get('auth_classification')} "
            f"live={b.get('live_trading_enabled')}"
        )
    except Exception as exc:  # noqa: BLE001
        out["broker"] = f"ERR:{exc}"
    try:
        h = _http_json(f"{base}/api/health")
        out["health"] = f"status={h.get('status')} qc={h.get('qc_status')}"
    except Exception as exc:  # noqa: BLE001
        out["health"] = f"ERR:{exc}"
    try:
        g = _http_json(f"{base}/api/auto_gates")
        out["gates"] = (
            f"{g.get('gates_passing')}/{g.get('gates_total')} "
            f"trade_ready={g.get('trade_ready')}"
        )
    except Exception as exc:  # noqa: BLE001
        out["gates"] = f"ERR:{exc}"
    return out


def add(candidates: list[dict], path: Path, source_class: str, notes: str = "", github: str = "UNKNOWN") -> None:
    if not path.exists():
        return
    st = path.stat()
    rows = None
    sheets = None
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
                rows = max(sum(1 for _ in f) - 1, 0)
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = ",".join(wb.sheetnames)
            wb.close()
    except Exception as exc:  # noqa: BLE001
        notes = f"{notes} read_err={exc}".strip()
    candidates.append(
        {
            "path": str(path),
            "name": path.name,
            "source_class": source_class,
            "bytes": st.st_size,
            "mtime_utc": datetime.fromtimestamp(st.st_mtime, timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "rows_or_na": rows if rows is not None else "n/a",
            "sheets_or_na": sheets if sheets else "n/a",
            "github_presence": github,
            "notes": notes,
        }
    )


def score_row(c: dict) -> dict:
    score = 0
    reasons: list[str] = []
    if c["source_class"] == "PRIMARY_CLONE":
        score += 40
        reasons.append("+40 primary clone")
    elif c["source_class"] == "BANNED_C_SYSTEM3":
        score -= 100
        reasons.append("-100 banned path")
    elif c["source_class"] == "OVERLAY_C_GENESIS":
        score -= 50
        reasons.append("-50 overlay not clone")
    elif c["source_class"] == "WORKTREE_COPY":
        score += 5
        reasons.append("+5 worktree secondary")

    name = c["name"].lower()
    if name == "github_action_map_status.csv":
        score += 35
        reasons.append("+35 runbook single live action board")
        if c.get("github_presence") == "NOT_ON_MAIN":
            score -= 8
            reasons.append("-8 not on GitHub main yet (laptop+Excel twin)")
    if name == "agent_operating_options.xlsx":
        score += 32
        reasons.append("+32 multi-sheet operator workbook on GitHub")
    if name == "session_issues_master.csv":
        score += 25
        reasons.append("+25 session issues master")
    if name == "pending_issues_master.csv":
        score += 10
        reasons.append("+10 pending master (verify SHAs)")
    if "20260825" in name:
        score -= 5
        reasons.append("-5 dated snapshot")
    if name.startswith("system3_master_mri"):
        score += 8
        reasons.append("+8 MRI workbook not daily P0 board")
    if "ruhi" in name:
        score += 12
        reasons.append("+12 RUHI ledger")
    if c.get("github_presence") == "ON_MAIN":
        score += 8
        reasons.append("+8 present on GitHub main")
    elif c.get("github_presence") == "NOT_ON_MAIN":
        score -= 3
        reasons.append("-3 missing from GitHub main")

    try:
        mt = datetime.strptime(c["mtime_utc"], "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
        age_h = (datetime.now(timezone.utc) - mt).total_seconds() / 3600
        if age_h < 24:
            score += 10
            reasons.append("+10 mtime <24h")
        elif age_h < 72:
            score += 5
            reasons.append("+5 mtime <72h")
        else:
            score -= 5
            reasons.append("-5 older than 72h")
    except Exception:
        pass

    if score >= 70:
        follow = "FOLLOW_FIRST"
    elif score >= 50:
        follow = "FOLLOW_SECOND"
    elif score >= 30:
        follow = "CONTEXT_ONLY"
    elif score < 0:
        follow = "BAN_IGNORE"
    else:
        follow = "CONTEXT_ONLY"

    recs = {
        "GITHUB_ACTION_MAP_STATUS.csv": (
            "BEST daily done/pending action map (Cloud/GitHub/Gmail). "
            "Excel twin = sheet 9_GitHub_Action_Map. Commit to main so cloud agents can see it."
        ),
        "AGENT_OPERATING_OPTIONS.xlsx": (
            "BEST Excel multi-sheet operator file (on GitHub). Open this; use new sheet 10_Authority_Rank."
        ),
        "session_issues_master.csv": (
            "BEST issue-status twin for current serving SHA; overwrite each command-center refresh."
        ),
        "pending_issues_master.csv": "Backlog catalog only; SHAs may be stale — re-verify live.",
        "PENDING_ISSUES_MASTER_20260825.csv": "Historical dated dump — do not drive work from this alone.",
        "System3_Master_MRI_Control.xlsx": "MRI deep-scan workbook — not the daily P0 action board.",
        "ruhi_task_ledger.csv": "RUHI companion; subordinate to GITHUB_ACTION_MAP_STATUS.csv.",
        "SYSTEM3_LIVE_UNRESOLVED_ISSUES.csv": "Audit trail; verify against live APIs before acting.",
    }
    return {
        **c,
        "score": score,
        "follow_rank": follow,
        "score_reasons": "; ".join(reasons),
        "recommendation": recs.get(
            c["name"],
            "Secondary/competing copy — prefer primary reports/coordination/ paths.",
        ),
    }


def main() -> int:
    live = live_snapshot()
    serving = live["serving_sha"]
    main_sha = serving  # same-session fetch below if git available
    try:
        import subprocess

        main_sha = (
            subprocess.check_output(
                ["git", "rev-parse", "origin/main"],
                cwd=str(ROOT),
                text=True,
            ).strip()
            or serving
        )
    except Exception:
        main_sha = serving

    candidates: list[dict] = []
    add(
        candidates,
        ROOT / "reports/coordination/GITHUB_ACTION_MAP_STATUS.csv",
        "PRIMARY_CLONE",
        "Runbook-named live action board",
        github="NOT_ON_MAIN",
    )
    add(
        candidates,
        ROOT / "reports/coordination/session_issues_master.csv",
        "PRIMARY_CLONE",
        "Session overwrite twin",
        github="ON_MAIN",
    )
    add(
        candidates,
        ROOT / "reports/coordination/pending_issues_master.csv",
        "PRIMARY_CLONE",
        "Legacy pending master",
        github="UNKNOWN",
    )
    add(
        candidates,
        ROOT / "reports/latest/pending_issues/PENDING_ISSUES_MASTER_20260825.csv",
        "PRIMARY_CLONE",
        "Dated snapshot",
        github="UNKNOWN",
    )
    add(
        candidates,
        ROOT / "reports/coordination/AGENT_OPERATING_OPTIONS.xlsx",
        "PRIMARY_CLONE",
        "Existing multi-sheet operator workbook",
        github="ON_MAIN",
    )
    add(
        candidates,
        ROOT / "reports/coordination/ruhi_task_ledger.csv",
        "PRIMARY_CLONE",
        "RUHI task ledger",
        github="ON_MAIN",
    )
    add(
        candidates,
        ROOT / "System3_Master_MRI_Control.xlsx",
        "PRIMARY_CLONE",
        "MRI scan workbook",
        github="UNKNOWN",
    )
    add(
        candidates,
        ROOT / "audit/live_agent_issue_ledger/SYSTEM3_LIVE_UNRESOLVED_ISSUES.csv",
        "PRIMARY_CLONE",
        "Audit ledger",
        github="UNKNOWN",
    )
    add(
        candidates,
        Path(r"C:\System3\Genesis_System3\reports\coordination\GITHUB_ACTION_MAP_STATUS.csv"),
        "BANNED_C_SYSTEM3",
        "Hard-ban path",
    )
    add(
        candidates,
        Path(r"C:\System3\Genesis_System3\System3_Master_MRI_Control.xlsx"),
        "BANNED_C_SYSTEM3",
        "Hard-ban path",
    )
    add(
        candidates,
        Path(r"C:\Genesis_System3\worktrees\agent-evidence-catalog-20260825\System3_Master_MRI_Control.xlsx"),
        "OVERLAY_C_GENESIS",
        "Overlay/worktree copy",
    )
    add(
        candidates,
        ROOT / ".worktrees/master-production-closure/System3_Master_MRI_Control.xlsx",
        "WORKTREE_COPY",
        "Linked worktree MRI copy",
    )
    add(
        candidates,
        ROOT / ".worktrees/master-production-closure/reports/coordination/ruhi_task_ledger.csv",
        "WORKTREE_COPY",
        "Worktree copy may lag",
    )

    rank_rows = [score_row(c) for c in candidates]
    rank_rows.sort(key=lambda r: (-r["score"], r["path"]))

    out_csv = ROOT / "reports/coordination/AUTHORITY_RANK_ANALYSIS.csv"
    fields = [
        "rank",
        "follow_rank",
        "score",
        "name",
        "source_class",
        "path",
        "bytes",
        "mtime_utc",
        "rows_or_na",
        "sheets_or_na",
        "github_presence",
        "serving_sha_now",
        "github_main_sha_now",
        "live_broker",
        "live_health",
        "live_gates",
        "score_reasons",
        "recommendation",
        "notes",
        "analyzed_utc",
    ]
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for i, r in enumerate(rank_rows, 1):
            w.writerow(
                {
                    "rank": i,
                    "follow_rank": r["follow_rank"],
                    "score": r["score"],
                    "name": r["name"],
                    "source_class": r["source_class"],
                    "path": r["path"],
                    "bytes": r["bytes"],
                    "mtime_utc": r["mtime_utc"],
                    "rows_or_na": r["rows_or_na"],
                    "sheets_or_na": r["sheets_or_na"],
                    "github_presence": r["github_presence"],
                    "serving_sha_now": serving,
                    "github_main_sha_now": main_sha,
                    "live_broker": live["broker"],
                    "live_health": live["health"],
                    "live_gates": live["gates"],
                    "score_reasons": r["score_reasons"],
                    "recommendation": r["recommendation"],
                    "notes": r["notes"],
                    "analyzed_utc": NOW,
                }
            )

    xlsx = ROOT / "reports/coordination/AGENT_OPERATING_OPTIONS.xlsx"
    wb = load_workbook(xlsx)
    sheet_name = "10_Authority_Rank"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    best_fill = PatternFill("solid", fgColor="C6EFCE")
    ban_fill = PatternFill("solid", fgColor="FFC7CE")
    second_fill = PatternFill("solid", fgColor="FFEB9C")

    ws.append(fields)
    for col, _h in enumerate(fields, 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, r in enumerate(rank_rows, 1):
        ws.append(
            [
                i,
                r["follow_rank"],
                r["score"],
                r["name"],
                r["source_class"],
                r["path"],
                r["bytes"],
                r["mtime_utc"],
                r["rows_or_na"],
                r["sheets_or_na"],
                r["github_presence"],
                serving,
                main_sha,
                live["broker"],
                live["health"],
                live["gates"],
                r["score_reasons"],
                r["recommendation"],
                r["notes"],
                NOW,
            ]
        )
        fr = r["follow_rank"]
        fill = None
        if fr == "FOLLOW_FIRST":
            fill = best_fill
        elif fr == "FOLLOW_SECOND":
            fill = second_fill
        elif fr == "BAN_IGNORE":
            fill = ban_fill
        if fill:
            for col in range(1, len(fields) + 1):
                ws.cell(i + 1, col).fill = fill

    start = len(rank_rows) + 3
    ws.cell(start, 1, "VERDICT").font = Font(bold=True)
    ws.cell(
        start,
        2,
        "FOLLOW_FIRST: AGENT_OPERATING_OPTIONS.xlsx (Excel) + GITHUB_ACTION_MAP_STATUS.csv (daily map) "
        "+ session_issues_master.csv (issue statuses). Ignore C:\\System3 / overlay copies. "
        "Live PASS only from GitHub main + /api/deploy_info.",
    )
    ws.cell(start + 1, 1, "GAP").font = Font(bold=True)
    ws.cell(
        start + 1,
        2,
        "GITHUB_ACTION_MAP_STATUS.csv is NOT on GitHub main (404) — cloud agents cannot see it until committed/pushed. "
        "Prefer Excel sheet 9_GitHub_Action_Map / this workbook until CSV is on main.",
    )
    ws.cell(start + 2, 1, "CLOUD_NOW").font = Font(bold=True)
    ws.cell(
        start + 2,
        2,
        f"serving={serving} main={main_sha} broker={live['broker']} health={live['health']} gates={live['gates']} LIVE=false",
    )

    widths = [6, 14, 8, 36, 18, 70, 10, 22, 10, 40, 14, 42, 42, 36, 28, 22, 55, 70, 40, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(width, 55)

    wb.save(xlsx)
    print(json.dumps({"csv": str(out_csv), "xlsx": str(xlsx), "sheet": sheet_name, "top": [
        {"follow": r["follow_rank"], "score": r["score"], "name": r["name"]} for r in rank_rows[:5]
    ], "live": live, "main": main_sha}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
