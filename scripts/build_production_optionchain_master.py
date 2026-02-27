"""
Build Production-Grade OptionChain Master Excel File
Analyzes existing file, fills missing data, adds all calculations
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings("ignore")

ROOT_DIR = Path(__file__).parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

# Import existing calculation functions
from core.utils.option_chain_calculations import add_calculated_columns
from src.metrics.greeks import calculate_greeks, calculate_greeks_from_market_price
from src.metrics.oi_buildup import compute_deltas, classify_oi_buildup
from src.selector.top_symbol_selector import TopSymbolSelector
from src.metrics.iv_solver import solve_implied_volatility, black_scholes_price


class OptionChainMasterBuilder:
    """
    Builds production-grade OptionChain Master Excel file with all data and calculations.
    """

    def __init__(self):
        self.output_dir = ROOT_DIR / "outputs"
        self.excel_path = self.output_dir / "OptionChain_Master_v3_AI_FINAL.xlsx"
        self.csv_path = self.output_dir / "chain_raw_live.csv"
        self.risk_free_rate = 0.06  # 6% annual

    def load_existing_data(self):
        """Load data from CSV and existing Excel."""
        print("\n[STEP 1] Loading Existing Data")
        print("-" * 80)

        # Load from CSV (most complete)
        if self.csv_path.exists():
            print(f"Loading from CSV: {self.csv_path.name}")
            df_csv = pd.read_csv(self.csv_path, on_bad_lines="skip", engine="python")
            print(f"  CSV Rows: {len(df_csv):,}")
            print(f"  CSV Columns: {len(df_csv.columns)}")
        else:
            print(f"WARNING: CSV not found: {self.csv_path}")
            df_csv = pd.DataFrame()

        # Try to load existing Excel
        df_excel = pd.DataFrame()
        excel_sheets = {}
        if self.excel_path.exists():
            try:
                print(f"\nLoading from Excel: {self.excel_path.name}")
                xl = pd.ExcelFile(self.excel_path)
                print(f"  Excel Sheets: {xl.sheet_names}")

                for sheet_name in xl.sheet_names:
                    df_sheet = pd.read_excel(xl, sheet_name=sheet_name)
                    excel_sheets[sheet_name] = df_sheet
                    print(f"    {sheet_name}: {len(df_sheet)} rows, {len(df_sheet.columns)} columns")

                    # Use first sheet as primary if df_excel is empty
                    if df_excel.empty:
                        df_excel = df_sheet
            except Exception as e:
                print(f"  WARNING: Could not read Excel: {e}")

        # Use CSV as primary source (most complete)
        if not df_csv.empty:
            df_primary = df_csv.copy()
            print(f"\nUsing CSV as primary data source")
        elif not df_excel.empty:
            df_primary = df_excel.copy()
            print(f"\nUsing Excel as primary data source")
        else:
            raise ValueError("No data source available!")

        return df_primary, excel_sheets

    def add_all_calculations(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add all possible option chain calculations."""
        print("\n[STEP 2] Adding All Calculations")
        print("-" * 80)

        df = df.copy()
        original_cols = len(df.columns)

        # 1. Use existing calculated columns function
        print("  Adding basic calculated columns...")
        try:
            fetch_timestamp = df.get("timestamp_ist", pd.Series([None]))[0] if "timestamp_ist" in df.columns else None
            df = add_calculated_columns(df, fetch_timestamp)
        except Exception as e:
            print(f"  WARNING: Error in add_calculated_columns: {e}")

        # 2. Add moneyness classification
        print("  Adding moneyness classification...")
        if "strike" in df.columns and "spot_price" in df.columns:

            def classify_moneyness(row):
                try:
                    spot = float(row["spot_price"]) if pd.notna(row["spot_price"]) else None
                    strike = float(row["strike"]) if pd.notna(row["strike"]) else None
                    opt_type = str(row["option_type"]).upper() if "option_type" in row else None

                    if spot is None or strike is None:
                        return "UNKNOWN"

                    distance_pct = abs(strike - spot) / spot * 100

                    if distance_pct < 0.5:  # Within 0.5%
                        return "ATM"
                    elif opt_type == "CE":
                        if strike < spot:
                            return "ITM"
                        else:
                            return "OTM"
                    elif opt_type == "PE":
                        if strike > spot:
                            return "ITM"
                        else:
                            return "OTM"
                    else:
                        return "UNKNOWN"
                except:
                    return "UNKNOWN"

            df["moneyness"] = df.apply(classify_moneyness, axis=1)

        # 3. Calculate Greeks if missing (using Black-Scholes)
        print("  Calculating missing Greeks...")
        missing_greeks = []
        for col in ["delta", "gamma", "theta", "vega", "rho", "iv"]:
            if col not in df.columns or df[col].isna().sum() > len(df) * 0.5:
                missing_greeks.append(col)

        if missing_greeks:
            print(f"    Missing/Incomplete Greeks: {missing_greeks}")

            # Calculate Greeks for rows where they're missing
            for idx, row in df.iterrows():
                try:
                    spot = row.get("spot_price")
                    strike = row.get("strike")
                    time_to_expiry = row.get("time_to_expiry", 0.065)  # Default ~24 days
                    option_type = str(row.get("option_type", "CE")).upper()
                    market_price = row.get("ltp") or row.get("mid_price")

                    if all(pd.notna(x) for x in [spot, strike, time_to_expiry]) and market_price and market_price > 0:
                        # Try to get IV from row, or solve for it
                        iv = row.get("iv")
                        if pd.isna(iv) or iv <= 0:
                            # Solve for IV
                            iv = solve_implied_volatility(
                                spot, strike, time_to_expiry, self.risk_free_rate, market_price, option_type
                            )

                        if iv and iv > 0:
                            # Calculate all Greeks
                            greeks = calculate_greeks(
                                spot, strike, time_to_expiry, self.risk_free_rate, iv, option_type
                            )

                            # Fill missing values
                            for greek_name, greek_value in greeks.items():
                                if greek_name in missing_greeks or pd.isna(row.get(greek_name)):
                                    df.at[idx, greek_name] = greek_value

                            # Update IV if calculated
                            if "iv" in missing_greeks or pd.isna(row.get("iv")):
                                df.at[idx, "iv"] = iv
                except:
                    continue

        # 4. Add PCR (Put-Call Ratio) per strike
        print("  Adding PCR per strike...")
        if "strike" in df.columns and "option_type" in df.columns and "oi" in df.columns:

            def calculate_strike_pcr(strike_val):
                strike_df = df[df["strike"] == strike_val]
                ce_oi = (
                    strike_df[strike_df["option_type"] == "CE"]["oi"].sum()
                    if len(strike_df[strike_df["option_type"] == "CE"]) > 0
                    else 0
                )
                pe_oi = (
                    strike_df[strike_df["option_type"] == "PE"]["oi"].sum()
                    if len(strike_df[strike_df["option_type"] == "PE"]) > 0
                    else 0
                )
                return pe_oi / ce_oi if ce_oi > 0 else 0.0

            df["strike_pcr"] = df["strike"].apply(calculate_strike_pcr)

        # 5. Add expected move
        print("  Adding expected move...")
        if "spot_price" in df.columns and "iv" in df.columns and "time_to_expiry" in df.columns:

            def calc_expected_move(row):
                try:
                    spot = float(row["spot_price"]) if pd.notna(row["spot_price"]) else None
                    iv = float(row["iv"]) if pd.notna(row["iv"]) else 0.20  # Default 20%
                    tte = float(row["time_to_expiry"]) if pd.notna(row["time_to_expiry"]) else 0.065

                    if spot and iv > 0 and tte > 0:
                        return spot * iv * np.sqrt(tte)
                    return None
                except:
                    return None

            df["expected_move"] = df.apply(calc_expected_move, axis=1)

        # 6. Add gamma exposure (simplified)
        print("  Adding gamma exposure...")
        if "gamma" in df.columns and "oi" in df.columns and "spot_price" in df.columns:
            # Gamma exposure = Gamma * OI * Spot^2 * 0.01
            mask = df["gamma"].notna() & df["oi"].notna() & df["spot_price"].notna()
            df.loc[mask, "gamma_exposure"] = (
                df.loc[mask, "gamma"] * df.loc[mask, "oi"] * (df.loc[mask, "spot_price"] ** 2) * 0.01
            )

        # 7. Add theta exposure
        print("  Adding theta exposure...")
        if "theta" in df.columns and "oi" in df.columns:
            mask = df["theta"].notna() & df["oi"].notna()
            df.loc[mask, "theta_exposure"] = df.loc[mask, "theta"] * df.loc[mask, "oi"]

        # 8. Add vega exposure
        print("  Adding vega exposure...")
        if "vega" in df.columns and "oi" in df.columns:
            mask = df["vega"].notna() & df["oi"].notna()
            df.loc[mask, "vega_exposure"] = df.loc[mask, "vega"] * df.loc[mask, "oi"]

        # 9. Add liquidity score
        print("  Adding liquidity score...")
        if "volume" in df.columns and "oi" in df.columns and "bid_ask_spread_pct" in df.columns:

            def calc_liquidity_score(row):
                try:
                    score = 100.0

                    # Volume component (40 points)
                    volume = row.get("volume", 0) or 0
                    if volume > 10000:
                        vol_score = 40
                    elif volume > 5000:
                        vol_score = 30
                    elif volume > 1000:
                        vol_score = 20
                    else:
                        vol_score = 10

                    # OI component (30 points)
                    oi = row.get("oi", 0) or 0
                    if oi > 100000:
                        oi_score = 30
                    elif oi > 50000:
                        oi_score = 20
                    elif oi > 10000:
                        oi_score = 10
                    else:
                        oi_score = 5

                    # Spread component (30 points) - lower spread = better
                    spread_pct = row.get("bid_ask_spread_pct", 10) or 10
                    if spread_pct < 1:
                        spread_score = 30
                    elif spread_pct < 3:
                        spread_score = 20
                    elif spread_pct < 5:
                        spread_score = 10
                    else:
                        spread_score = 0

                    return vol_score + oi_score + spread_score
                except:
                    return 0.0

            df["liquidity_score"] = df.apply(calc_liquidity_score, axis=1)

        # 10. Add IV rank and percentile
        print("  Adding IV rank and percentile...")
        if "iv" in df.columns and "underlying" in df.columns:
            for underlying in df["underlying"].unique():
                underlying_df = df[df["underlying"] == underlying]
                iv_values = underlying_df["iv"].dropna()

                if len(iv_values) > 0:
                    iv_min = iv_values.min()
                    iv_max = iv_values.max()
                    iv_range = iv_max - iv_min

                    mask = df["underlying"] == underlying
                    if iv_range > 0:
                        df.loc[mask, "iv_rank"] = ((df.loc[mask, "iv"] - iv_min) / iv_range * 100).fillna(0)
                    else:
                        df.loc[mask, "iv_rank"] = 50.0  # Middle if no range

                    # IV Percentile (simplified - would need historical data for true percentile)
                    df.loc[mask, "iv_percentile"] = df.loc[mask, "iv_rank"]  # Use rank as proxy

        # 11. Add time value decay rate
        print("  Adding time value decay rate...")
        if "theta" in df.columns and "extrinsic_value" in df.columns and "days_to_expiry" in df.columns:
            mask = (
                df["theta"].notna()
                & df["extrinsic_value"].notna()
                & df["days_to_expiry"].notna()
                & (df["days_to_expiry"] > 0)
                & (df["extrinsic_value"] > 0)
            )
            df.loc[mask, "time_decay_rate"] = abs(df.loc[mask, "theta"]) / df.loc[mask, "extrinsic_value"] * 100

        # 12. Add break-even price
        print("  Adding break-even price...")
        if "strike" in df.columns and "option_type" in df.columns and "ltp" in df.columns:

            def calc_breakeven(row):
                try:
                    strike = float(row["strike"]) if pd.notna(row["strike"]) else None
                    premium = float(row["ltp"]) if pd.notna(row["ltp"]) else None
                    opt_type = str(row["option_type"]).upper() if "option_type" in row else None

                    if strike and premium and opt_type:
                        if opt_type == "CE":
                            return strike + premium
                        else:  # PE
                            return strike - premium
                    return None
                except:
                    return None

            df["breakeven_price"] = df.apply(calc_breakeven, axis=1)

        # 13. Add max profit and max loss (for long positions)
        print("  Adding max profit/loss...")
        if "option_type" in df.columns and "ltp" in df.columns:

            def calc_max_profit_loss(row):
                try:
                    opt_type = str(row["option_type"]).upper() if "option_type" in row else None
                    premium = float(row["ltp"]) if pd.notna(row["ltp"]) else None

                    if premium:
                        # For long option: max loss = premium paid, max profit = unlimited (set to None)
                        max_loss = -premium
                        max_profit = None  # Unlimited for long options
                        return max_profit, max_loss
                    return None, None
                except:
                    return None, None

            df[["max_profit", "max_loss"]] = df.apply(lambda row: pd.Series(calc_max_profit_loss(row)), axis=1)

        # 14. Add risk-reward ratio
        print("  Adding risk-reward ratio...")
        if "max_profit" in df.columns and "max_loss" in df.columns:

            def calc_risk_reward(row):
                try:
                    max_profit = row.get("max_profit")
                    max_loss = abs(row.get("max_loss", 0)) if pd.notna(row.get("max_loss")) else 0

                    if max_loss > 0:
                        if max_profit is None or pd.isna(max_profit):
                            return None  # Unlimited profit
                        else:
                            return max_profit / max_loss
                    return None
                except:
                    return None

            df["risk_reward_ratio"] = df.apply(calc_risk_reward, axis=1)

        # 15. Add theoretical price (Black-Scholes)
        print("  Adding theoretical price (Black-Scholes)...")
        if all(col in df.columns for col in ["spot_price", "strike", "time_to_expiry", "iv", "option_type"]):

            def calc_theoretical_price(row):
                try:
                    spot = float(row["spot_price"]) if pd.notna(row["spot_price"]) else None
                    strike = float(row["strike"]) if pd.notna(row["strike"]) else None
                    tte = float(row["time_to_expiry"]) if pd.notna(row["time_to_expiry"]) else None
                    iv = float(row["iv"]) if pd.notna(row["iv"]) else None
                    opt_type = str(row["option_type"]).upper() if "option_type" in row else None

                    if all(x is not None for x in [spot, strike, tte, iv]) and iv > 0 and tte > 0:
                        price = black_scholes_price(spot, strike, tte, self.risk_free_rate, iv, opt_type)
                        return price
                    return None
                except:
                    return None

            df["theoretical_price"] = df.apply(calc_theoretical_price, axis=1)

            # Add price difference
            if "ltp" in df.columns:
                mask = df["theoretical_price"].notna() & df["ltp"].notna()
                df.loc[mask, "price_vs_theoretical"] = df.loc[mask, "ltp"] - df.loc[mask, "theoretical_price"]
                df.loc[mask, "price_vs_theoretical_pct"] = (
                    df.loc[mask, "price_vs_theoretical"] / df.loc[mask, "theoretical_price"]
                ) * 100

        new_cols = len(df.columns) - original_cols
        print(f"\n  Added {new_cols} new calculated columns")
        print(f"  Total columns: {len(df.columns)}")

        return df

    def fill_missing_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Fill missing data using calculations and defaults."""
        print("\n[STEP 3] Filling Missing Data")
        print("-" * 80)

        df = df.copy()
        filled_count = 0

        # Fill missing spot_price (use median for same underlying)
        if "spot_price" in df.columns and "underlying" in df.columns:
            missing_spot = df["spot_price"].isna().sum()
            if missing_spot > 0:
                print(f"  Filling {missing_spot} missing spot_price values...")
                for underlying in df["underlying"].unique():
                    mask = (df["underlying"] == underlying) & df["spot_price"].isna()
                    median_spot = df[df["underlying"] == underlying]["spot_price"].median()
                    if pd.notna(median_spot):
                        df.loc[mask, "spot_price"] = median_spot
                        filled_count += mask.sum()

        # Fill missing mid_price (calculate from bid/ask or use LTP)
        if "mid_price" in df.columns:
            missing_mid = df["mid_price"].isna().sum()
            if missing_mid > 0:
                print(f"  Filling {missing_mid} missing mid_price values...")
                # Try to calculate from bid/ask
                if "bidPrice" in df.columns and "offerPrice" in df.columns:
                    mask = df["mid_price"].isna() & df["bidPrice"].notna() & df["offerPrice"].notna()
                    df.loc[mask, "mid_price"] = (df.loc[mask, "bidPrice"] + df.loc[mask, "offerPrice"]) / 2
                    filled_count += mask.sum()

                # Use LTP as fallback
                if "ltp" in df.columns:
                    mask = df["mid_price"].isna() & df["ltp"].notna()
                    df.loc[mask, "mid_price"] = df.loc[mask, "ltp"]
                    filled_count += mask.sum()

        # Fill missing bid/ask from mid_price
        if "bidPrice" in df.columns and "mid_price" in df.columns:
            missing_bid = df["bidPrice"].isna().sum()
            if missing_bid > 0:
                print(f"  Filling {missing_bid} missing bidPrice values...")
                mask = df["bidPrice"].isna() & df["mid_price"].notna()
                # Assume 0.5% spread
                df.loc[mask, "bidPrice"] = df.loc[mask, "mid_price"] * 0.995
                filled_count += mask.sum()

        if "offerPrice" in df.columns and "mid_price" in df.columns:
            missing_ask = df["offerPrice"].isna().sum()
            if missing_ask > 0:
                print(f"  Filling {missing_ask} missing offerPrice values...")
                mask = df["offerPrice"].isna() & df["mid_price"].notna()
                df.loc[mask, "offerPrice"] = df.loc[mask, "mid_price"] * 1.005
                filled_count += mask.sum()

        # Fill missing volume/OI with 0 (no trading activity)
        for col in ["volume", "oi"]:
            if col in df.columns:
                missing = df[col].isna().sum()
                if missing > 0:
                    print(f"  Filling {missing} missing {col} values with 0...")
                    df[col] = df[col].fillna(0)
                    filled_count += missing

        print(f"\n  Total values filled: {filled_count:,}")

        return df

    def create_excel_file(self, df: pd.DataFrame, excel_sheets: dict = None):
        """Create production-grade Excel file with formatting."""
        print("\n[STEP 4] Creating Excel File")
        print("-" * 80)

        output_path = self.excel_path

        # Create backup if exists
        if output_path.exists():
            backup_path = output_path.with_suffix(".xlsx.backup")
            if not backup_path.exists():
                import shutil

                shutil.copy2(output_path, backup_path)
                print(f"  Backup created: {backup_path.name}")

        # Sort data for better organization
        sort_cols = []
        for col in ["underlying", "expiry", "strike", "option_type"]:
            if col in df.columns:
                sort_cols.append(col)

        if sort_cols:
            df = df.sort_values(sort_cols).reset_index(drop=True)

        # Create Excel writer
        print(f"  Writing to Excel: {output_path.name}")
        print(f"  Rows: {len(df):,}")
        print(f"  Columns: {len(df.columns)}")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Main data sheet (CHAIN_RAW - preserve existing name if sheet exists)
            main_sheet_name = "CHAIN_RAW" if excel_sheets and "CHAIN_RAW" in excel_sheets else "OptionChain_Data"
            df.to_excel(writer, sheet_name=main_sheet_name, index=False)

            # Create CHAIN_CALC sheet with all calculated columns
            calc_cols = [
                col
                for col in df.columns
                if col
                not in [
                    "underlying",
                    "exchange",
                    "token",
                    "symbol",
                    "strike",
                    "option_type",
                    "expiry",
                    "ltp",
                    "oi",
                    "volume",
                    "bidPrice",
                    "offerPrice",
                    "spot_price",
                ]
            ]
            if calc_cols:
                calc_df = df[["underlying", "strike", "option_type"] + calc_cols].copy()
                calc_df.to_excel(writer, sheet_name="CHAIN_CALC", index=False)

            # Create summary sheet
            self._create_summary_sheet(writer, df)

            # Create OI_BUILDUP analysis sheet
            if "oi_buildup" in df.columns:
                oi_df = df[
                    ["underlying", "strike", "option_type", "oi", "dOI", "dVolume", "oi_buildup", "ltp", "dLTP"]
                ].copy()
                oi_df = oi_df[oi_df["oi_buildup"] != "Neutral"].sort_values("dOI", ascending=False)
                oi_df.to_excel(writer, sheet_name="OI_BUILDUP", index=False)

            # Create GREEKS sheet
            greeks_cols = ["underlying", "strike", "option_type", "delta", "gamma", "theta", "vega", "rho", "iv"]
            if all(col in df.columns for col in greeks_cols):
                greeks_df = df[greeks_cols].copy()
                greeks_df.to_excel(writer, sheet_name="GREEKS", index=False)

            # Create GAMMA_THETA_MAP sheet
            if all(col in df.columns for col in ["strike", "gamma_exposure", "theta_exposure", "underlying"]):
                exposure_df = df[
                    ["underlying", "strike", "option_type", "gamma_exposure", "theta_exposure", "oi", "iv"]
                ].copy()
                exposure_df = exposure_df.sort_values("gamma_exposure", ascending=False, na_position="last")
                exposure_df.to_excel(writer, sheet_name="GAMMA_THETA_MAP", index=False)

            # Create underlying-wise sheets
            if "underlying" in df.columns:
                for underlying in df["underlying"].unique():
                    underlying_df = df[df["underlying"] == underlying].copy()
                    sheet_name = f"{underlying}_Chain"[:31]  # Excel sheet name limit
                    underlying_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Preserve existing sheets if they exist and are not overwritten
            if excel_sheets:
                for sheet_name, sheet_df in excel_sheets.items():
                    # Skip if we already wrote to this sheet
                    if sheet_name not in writer.sheets:
                        try:
                            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            print(f"  Preserved existing sheet: {sheet_name}")
                        except:
                            print(f"  WARNING: Could not preserve sheet: {sheet_name}")

        # Apply formatting
        print("  Applying Excel formatting...")
        self._apply_excel_formatting(output_path, df)

        print(f"\n  Excel file created: {output_path}")
        print(f"  Size: {output_path.stat().st_size:,} bytes")

        return output_path

    def _create_summary_sheet(self, writer, df: pd.DataFrame):
        """Create summary sheet with key metrics."""
        summary_data = []

        if "underlying" in df.columns:
            for underlying in df["underlying"].unique():
                underlying_df = df[df["underlying"] == underlying]

                summary_row = {
                    "Underlying": underlying,
                    "Total_Contracts": len(underlying_df),
                    "CE_Contracts": len(underlying_df[underlying_df["option_type"] == "CE"]),
                    "PE_Contracts": len(underlying_df[underlying_df["option_type"] == "PE"]),
                }

                if "spot_price" in underlying_df.columns:
                    summary_row["Spot_Price"] = underlying_df["spot_price"].median()

                if "oi" in underlying_df.columns:
                    summary_row["Total_OI"] = underlying_df["oi"].sum()
                    summary_row["CE_OI"] = underlying_df[underlying_df["option_type"] == "CE"]["oi"].sum()
                    summary_row["PE_OI"] = underlying_df[underlying_df["option_type"] == "PE"]["oi"].sum()

                if "volume" in underlying_df.columns:
                    summary_row["Total_Volume"] = underlying_df["volume"].sum()

                if "pcr" in underlying_df.columns or ("oi" in underlying_df.columns):
                    ce_oi = underlying_df[underlying_df["option_type"] == "CE"]["oi"].sum()
                    pe_oi = underlying_df[underlying_df["option_type"] == "PE"]["oi"].sum()
                    summary_row["PCR"] = pe_oi / ce_oi if ce_oi > 0 else 0

                summary_data.append(summary_row)

        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

    def _apply_excel_formatting(self, file_path: Path, df: pd.DataFrame):
        """Apply professional formatting to Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path)

            # Format all data sheets
            data_sheets = ["OptionChain_Data", "CHAIN_RAW", "CHAIN_CALC", "OI_BUILDUP", "GREEKS", "GAMMA_THETA_MAP"]

            for sheet_name in data_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Header formatting
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)

                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # Auto-adjust column widths
                    try:
                        sheet_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
                        for idx, col in enumerate(sheet_df.columns, 1):
                            col_letter = get_column_letter(idx)
                            max_length = max(
                                sheet_df[col].astype(str).map(len).max() if len(sheet_df) > 0 else 0, len(str(col))
                            )
                            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                    except:
                        # Fallback: use df columns if available
                        for idx, col in enumerate(df.columns, 1):
                            col_letter = get_column_letter(idx)
                            ws.column_dimensions[col_letter].width = min(len(str(col)) + 2, 50)

                    # Freeze first row
                    ws.freeze_panes = "A2"

            wb.save(file_path)
            print("  Formatting applied successfully")
        except Exception as e:
            print(f"  WARNING: Could not apply formatting: {e}")

    def verify_excel_file(self, file_path: Path):
        """Verify the created Excel file."""
        print("\n[STEP 5] Verifying Excel File")
        print("-" * 80)

        if not file_path.exists():
            print("  ERROR: File not created!")
            return False

        try:
            xl = pd.ExcelFile(file_path)
            print(f"  File exists: OK")
            print(f"  Size: {file_path.stat().st_size:,} bytes")
            print(f"  Sheets: {len(xl.sheet_names)}")

            # Verify main sheet
            if "OptionChain_Data" in xl.sheet_names:
                df_check = pd.read_excel(xl, sheet_name="OptionChain_Data", nrows=10)
                print(f"  Main sheet columns: {len(df_check.columns)}")
                print(f"  Sample rows readable: OK")

            # Check for critical columns
            df_full = pd.read_excel(xl, sheet_name="OptionChain_Data")
            critical_cols = ["underlying", "strike", "option_type", "ltp", "spot_price"]
            missing_critical = [col for col in critical_cols if col not in df_full.columns]

            if missing_critical:
                print(f"  WARNING: Missing critical columns: {missing_critical}")
            else:
                print(f"  All critical columns present: OK")

            # Check data completeness
            total_cells = len(df_full) * len(df_full.columns)
            filled_cells = df_full.notna().sum().sum()
            completeness = (filled_cells / total_cells) * 100

            print(f"  Data completeness: {completeness:.1f}%")

            if completeness > 80:
                print(f"  Status: EXCELLENT")
            elif completeness > 60:
                print(f"  Status: GOOD")
            else:
                print(f"  Status: NEEDS IMPROVEMENT")

            return True

        except Exception as e:
            print(f"  ERROR: {str(e)}")
            import traceback

            traceback.print_exc()
            return False


def main():
    """Main execution."""
    print("\n" + "=" * 80)
    print("  OPTIONCHAIN MASTER - PRODUCTION BUILD")
    print("=" * 80)

    builder = OptionChainMasterBuilder()

    try:
        # Step 1: Load data
        df, excel_sheets = builder.load_existing_data()

        # Step 2: Add all calculations
        df = builder.add_all_calculations(df)

        # Step 3: Fill missing data
        df = builder.fill_missing_data(df)

        # Step 4: Create Excel file
        output_path = builder.create_excel_file(df, excel_sheets)

        # Step 5: Verify
        success = builder.verify_excel_file(output_path)

        if success:
            print("\n" + "=" * 80)
            print("  BUILD COMPLETE - PRODUCTION READY")
            print("=" * 80)
            print(f"\nOutput file: {output_path}")
            print(f"Total columns: {len(df.columns)}")
            print(f"Total rows: {len(df):,}")
            print("\nStatus: SUCCESS")
        else:
            print("\nStatus: COMPLETED WITH WARNINGS")

        return success

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
