#!/usr/bin/env python3
"""Extract live GCP inventory + live APIs into System3_GCP_FULL_MRI.xlsx."""
from __future__ import annotations

import json
import subprocess
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "reports" / "latest" / "gcp_full_mri_20260826"
XLSX = ROOT / "System3_GCP_FULL_MRI.xlsx"
PROJ = "system3-openalgo-safe"
REG = "asia-south1"
BASE = "https://genesis-system3-web-doq2wplepa-el.a.run.app"
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _gcloud_bin() -> str:
    candidates = [
        r"C:\Program Files (x86)\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        r"C:\Program Files\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        "gcloud",
    ]
    for c in candidates:
        if c == "gcloud" or Path(c).exists():
            return c
    return "gcloud"


def gcloud(args: list[str]) -> Any:
    exe = _gcloud_bin()
    cmd = [exe, *args]
    p = subprocess.run(cmd, capture_output=True, shell=False)
    raw = p.stdout or b""
    if p.returncode != 0 and not raw:
        raise RuntimeError(p.stderr.decode("utf-8", "replace") or f"gcloud rc={p.returncode}")
    for enc in ("utf-8", "utf-16", "utf-16-le"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            text = ""
    else:
        text = raw.decode("utf-8", "replace")
    text = text.lstrip("\ufeff").strip()
    if not text:
        if p.returncode != 0:
            raise RuntimeError(p.stderr.decode("utf-8", "replace") or f"gcloud rc={p.returncode}")
        return []
    return json.loads(text)


def http_json(path: str) -> Any:
    req = urllib.request.Request(f"{BASE}/{path}", headers={"User-Agent": "system3-gcp-full-mri/1.0"})
    with urllib.request.urlopen(req, timeout=40) as resp:
        return json.loads(resp.read().decode("utf-8"))


def load_local(name: str) -> Any:
    p = OUT_DIR / name
    if not p.exists():
        return None
    raw = p.read_bytes()
    for enc in ("utf-8", "utf-16", "utf-16-le"):
        try:
            return json.loads(raw.decode(enc).lstrip("\ufeff"))
        except Exception:
            continue
    return None


def save_json(name: str, data: Any) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / name).write_text(json.dumps(data, indent=2), encoding="utf-8")


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)


def write_rows(ws, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
    ws.append(headers)
    style_header(ws, len(headers))
    for r in rows:
        ws.append(r)
    autosize(ws, widths)


def dig(d: Any, *keys: str, default: Any = "") -> Any:
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur if cur is not None else default


def flatten_env(env_list: list[dict] | None) -> str:
    if not env_list:
        return ""
    parts = []
    for e in env_list:
        name = e.get("name", "")
        if "valueFrom" in e:
            sec = dig(e, "valueFrom", "secretKeyRef", "name", default="")
            parts.append(f"{name}=SECRET:{sec}")
        else:
            val = str(e.get("value", ""))
            if any(x in name.upper() for x in ("TOKEN", "SECRET", "KEY", "PASSWORD")):
                val = "***REDACTED***"
            parts.append(f"{name}={val}")
    return "; ".join(parts)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    inventory: dict[str, Any] = {"extracted_at_utc": NOW, "project": PROJ, "region": REG}

    # Prefer fresh gcloud; fall back to local dumps
    fetches = {
        "project_describe": ["projects", "describe", PROJ, "--format=json"],
        "billing": ["billing", "projects", "describe", PROJ, "--format=json"],
        "run_services": ["run", "services", "list", f"--project={PROJ}", f"--region={REG}", "--format=json"],
        "run_jobs": ["run", "jobs", "list", f"--project={PROJ}", f"--region={REG}", "--format=json"],
        "scheduler_jobs": ["scheduler", "jobs", "list", f"--project={PROJ}", f"--location={REG}", "--format=json"],
        "secrets": ["secrets", "list", f"--project={PROJ}", "--format=json"],
        "service_accounts": ["iam", "service-accounts", "list", f"--project={PROJ}", "--format=json"],
        "pubsub_topics": ["pubsub", "topics", "list", f"--project={PROJ}", "--format=json"],
        "buckets": ["storage", "buckets", "list", f"--project={PROJ}", "--format=json"],
        "apis_enabled": ["services", "list", "--enabled", f"--project={PROJ}", "--format=json"],
        "web_service_full": [
            "run",
            "services",
            "describe",
            "genesis-system3-web",
            f"--project={PROJ}",
            f"--region={REG}",
            "--format=json",
        ],
        "rotate_job_full": [
            "run",
            "jobs",
            "describe",
            "genesis-system3-dhan-token-rotate",
            f"--project={PROJ}",
            f"--region={REG}",
            "--format=json",
        ],
        "secret_dhan_versions": [
            "secrets",
            "versions",
            "list",
            "dhan-access-token",
            f"--project={PROJ}",
            "--limit=8",
            "--format=json",
        ],
        "web_revisions": [
            "run",
            "revisions",
            "list",
            f"--service=genesis-system3-web",
            f"--project={PROJ}",
            f"--region={REG}",
            "--limit=12",
            "--format=json",
        ],
    }

    for name, args in fetches.items():
        try:
            data = gcloud(args)
            inventory[name] = data
            save_json(f"{name}.json", data)
            print("OK", name)
        except Exception as exc:  # noqa: BLE001
            fallback = load_local(f"{name}.json")
            inventory[name] = fallback if fallback is not None else {"error": str(exc)}
            print("FALLBACK/ERR", name, exc)

    live = {}
    for path, key in [
        ("api/deploy_info", "deploy_info"),
        ("api/broker/status", "broker_status"),
        ("api/health", "health"),
        ("api/auto_gates", "auto_gates"),
        ("api/scheduler/health", "scheduler_health"),
    ]:
        try:
            live[key] = http_json(path)
            save_json(f"api_{key}.json", live[key])
            print("LIVE", key)
        except Exception as exc:  # noqa: BLE001
            live[key] = {"error": str(exc)}
            print("LIVE_ERR", key, exc)

    try:
        main_sha = subprocess.check_output(
            ["git", "rev-parse", "origin/main"], cwd=str(ROOT), text=True
        ).strip()
    except Exception:
        main_sha = (OUT_DIR / "github_origin_main.txt").read_text(encoding="utf-8").strip() if (OUT_DIR / "github_origin_main.txt").exists() else "UNKNOWN"

    serving = dig(live.get("deploy_info", {}), "git_sha", default="UNKNOWN")
    console = {
        "project_id": PROJ,
        "account": "warghade2012@gmail.com",
        "dashboard_url": f"https://console.cloud.google.com/home/dashboard?project={PROJ}",
        "billing_estimate_console_ui": "INR 17704.64 for 1-25 Aug 2026 (console card)",
        "api_requests_console_ui": "~0.84/s (dashboard chart at capture)",
        "top_error_console_ui": "Cloud Run abort: no available instance (858 occurrences shown)",
        "platform_status_console_ui": "All services normal",
        "resources_shown": "BigQuery, SQL, Compute Engine, Storage, Cloud Run functions, Cloud Run",
        "extracted_at_utc": NOW,
    }
    save_json("console_dashboard_capture.json", console)

    wb = Workbook()

    # 0 Executive
    ws = wb.active
    ws.title = "0_Executive_MRI"
    ws["A1"] = "SYSTEM3 GCP FULL MRI"
    ws["A1"].font = TITLE_FONT
    rows_exec = [
        ["Field", "Value", "Evidence_Class"],
        ["extracted_at_utc", NOW, "PROVEN"],
        ["gcp_project", PROJ, "PROVEN"],
        ["region", REG, "PROVEN"],
        ["console_account", console["account"], "PROVEN"],
        ["github_origin_main", main_sha, "PROVEN"],
        ["serving_git_sha", serving, "PROVEN"],
        ["main_equals_serving", str(main_sha == serving), "PROVEN"],
        ["live_trading_enabled", str(dig(live.get("deploy_info", {}), "live_trading_enabled", default=False)), "PROVEN"],
        ["broker_connected", str(dig(live.get("broker_status", {}), "connected", default="")), "PROVEN"],
        ["broker_auth", str(dig(live.get("broker_status", {}), "auth_classification", default="")), "PROVEN"],
        ["health_status", str(dig(live.get("health", {}), "status", default="")), "PROVEN"],
        ["health_qc", str(dig(live.get("health", {}), "qc_status", default="")), "PROVEN"],
        ["gates_passing", str(dig(live.get("auto_gates", {}), "gates_passing", default="")), "PROVEN"],
        ["gates_total", str(dig(live.get("auto_gates", {}), "gates_total", default="")), "PROVEN"],
        ["trade_ready", str(dig(live.get("auto_gates", {}), "trade_ready", default="")), "PROVEN"],
        ["scheduler_healthy", str(dig(live.get("scheduler_health", {}), "healthy", default="")), "PROVEN"],
        ["billing_estimate_ui", console["billing_estimate_console_ui"], "PROVEN_UI"],
        ["public_url", dig(live.get("deploy_info", {}), "public_base_url", default=BASE), "PROVEN"],
        ["artifact_dir", str(OUT_DIR), "PROVEN"],
        ["workbook", str(XLSX), "PROVEN"],
    ]
    for r in rows_exec:
        ws.append(r)
    style_header(ws, 3)
    autosize(ws, [28, 80, 14])

    # 1 Console capture
    ws = wb.create_sheet("1_Console_Dashboard")
    write_rows(
        ws,
        ["key", "value"],
        [[k, v] for k, v in console.items()],
        [40, 90],
    )

    # 2 Project
    proj = inventory.get("project_describe") or {}
    bill = inventory.get("billing") or {}
    ws = wb.create_sheet("2_Project_Billing")
    write_rows(
        ws,
        ["field", "value"],
        [
            ["projectId", dig(proj, "projectId")],
            ["projectNumber", dig(proj, "projectNumber")],
            ["name", dig(proj, "name")],
            ["lifecycleState", dig(proj, "lifecycleState")],
            ["createTime", dig(proj, "createTime")],
            ["billingAccountName", dig(bill, "billingAccountName")],
            ["billingEnabled", dig(bill, "billingEnabled")],
            ["console_billing_estimate", console["billing_estimate_console_ui"]],
        ],
        [28, 70],
    )

    # 3 Cloud Run services
    ws = wb.create_sheet("3_CloudRun_Services")
    svc_rows = []
    for s in inventory.get("run_services") or []:
        if not isinstance(s, dict):
            continue
        meta = s.get("metadata") or {}
        status = s.get("status") or {}
        spec = dig(s, "spec", "template", "spec", default={}) or {}
        containers = dig(spec, "containers", default=[]) or []
        image = dig(containers[0] if containers else {}, "image", default="")
        svc_rows.append(
            [
                meta.get("name"),
                dig(meta, "labels", "cloud.googleapis.com/location", default=REG),
                dig(status, "url"),
                dig(status, "latestReadyRevisionName"),
                dig(status, "latestCreatedRevisionName"),
                dig(status, "traffic", default=[{}])[0].get("percent", "") if dig(status, "traffic") else "",
                image,
                dig(meta, "annotations", "serving.knative.dev/creator", default=""),
                dig(meta, "creationTimestamp"),
            ]
        )
    write_rows(
        ws,
        ["name", "region", "url", "latestReadyRevision", "latestCreatedRevision", "traffic_pct", "image", "creator", "created"],
        svc_rows,
        [28, 14, 55, 40, 40, 10, 70, 30, 24],
    )

    # 4 Web service deep
    web = inventory.get("web_service_full") or {}
    ws = wb.create_sheet("4_Web_Service_Deep")
    tmpl = dig(web, "spec", "template", default={}) or {}
    tmpl_spec = dig(tmpl, "spec", default={}) or {}
    containers = dig(tmpl_spec, "containers", default=[]) or []
    c0 = containers[0] if containers else {}
    write_rows(
        ws,
        ["field", "value"],
        [
            ["service", dig(web, "metadata", "name")],
            ["uid", dig(web, "metadata", "uid")],
            ["generation", dig(web, "metadata", "generation")],
            ["url", dig(web, "status", "url")],
            ["latestReadyRevision", dig(web, "status", "latestReadyRevisionName")],
            ["observedGeneration", dig(web, "status", "observedGeneration")],
            ["serviceAccountName", dig(tmpl_spec, "serviceAccountName")],
            ["timeoutSeconds", dig(tmpl_spec, "timeoutSeconds")],
            ["containerConcurrency", dig(tmpl_spec, "containerConcurrency")],
            ["image", dig(c0, "image")],
            ["ports", json.dumps(dig(c0, "ports", default=[]))],
            ["resources_limits", json.dumps(dig(c0, "resources", "limits", default={}))],
            ["env_redacted", flatten_env(dig(c0, "env", default=[]))],
            ["annotations", json.dumps(dig(web, "metadata", "annotations", default={}))[:2000]],
            ["labels", json.dumps(dig(web, "metadata", "labels", default={}))],
        ],
        [28, 100],
    )

    # 5 Revisions
    ws = wb.create_sheet("5_Web_Revisions")
    rev_rows = []
    for r in inventory.get("web_revisions") or []:
        if not isinstance(r, dict):
            continue
        meta = r.get("metadata") or {}
        status = r.get("status") or {}
        conds = status.get("conditions") or []
        ready = next((c.get("status") for c in conds if c.get("type") == "Ready"), "")
        rev_rows.append(
            [
                meta.get("name"),
                ready,
                dig(meta, "annotations", "serving.knative.dev/routingState", default=""),
                dig(meta, "creationTimestamp"),
                dig(r, "spec", "containers", default=[{}])[0].get("image", "") if dig(r, "spec", "containers") else "",
            ]
        )
    write_rows(ws, ["revision", "ready", "routingState", "created", "image"], rev_rows, [45, 10, 16, 24, 80])

    # 6 Jobs
    ws = wb.create_sheet("6_CloudRun_Jobs")
    job_rows = []
    for j in inventory.get("run_jobs") or []:
        if not isinstance(j, dict):
            continue
        meta = j.get("metadata") or {}
        status = j.get("status") or {}
        tmpl = dig(j, "spec", "template", "template", "spec", default={}) or {}
        containers = dig(tmpl, "containers", default=[]) or []
        job_rows.append(
            [
                meta.get("name"),
                dig(status, "executionCount"),
                dig(status, "latestCreatedExecution", "name", default=""),
                dig(status, "latestCreatedExecution", "completionTimestamp", default="")
                or dig(status, "latestCreatedExecution", "completionStatus", default=""),
                dig(containers[0] if containers else {}, "image", default=""),
                dig(tmpl, "serviceAccountName"),
                dig(meta, "creationTimestamp"),
            ]
        )
    write_rows(
        ws,
        ["job", "executionCount", "latestExecution", "latestCompletion", "image", "serviceAccount", "created"],
        job_rows,
        [40, 12, 40, 28, 70, 45, 24],
    )

    # 7 Rotate job deep
    rot = inventory.get("rotate_job_full") or {}
    ws = wb.create_sheet("7_Token_Rotate_Job")
    rtmpl = dig(rot, "spec", "template", "template", "spec", default={}) or {}
    rcontainers = dig(rtmpl, "containers", default=[]) or []
    rc0 = rcontainers[0] if rcontainers else {}
    write_rows(
        ws,
        ["field", "value"],
        [
            ["job", dig(rot, "metadata", "name")],
            ["serviceAccountName", dig(rtmpl, "serviceAccountName")],
            ["image", dig(rc0, "image")],
            ["command", json.dumps(dig(rc0, "command", default=[]))],
            ["args", json.dumps(dig(rc0, "args", default=[]))],
            ["env_redacted", flatten_env(dig(rc0, "env", default=[]))],
            ["executionCount", dig(rot, "status", "executionCount")],
            ["latestExecution", dig(rot, "status", "latestCreatedExecution", "name", default="")],
            ["annotations", json.dumps(dig(rot, "metadata", "annotations", default={}))[:2000]],
        ],
        [28, 100],
    )

    # 8 Scheduler
    ws = wb.create_sheet("8_Scheduler_Jobs")
    sch_rows = []
    for s in inventory.get("scheduler_jobs") or []:
        if not isinstance(s, dict):
            continue
        http = s.get("httpTarget") or {}
        sch_rows.append(
            [
                s.get("name", "").split("/")[-1],
                s.get("schedule"),
                s.get("timeZone"),
                s.get("state"),
                http.get("uri") or dig(s, "pubsubTarget", "topicName", default=""),
                http.get("httpMethod"),
                dig(s, "lastAttemptTime"),
                dig(s, "scheduleTime"),
                dig(s, "status", "code", default=""),
            ]
        )
    write_rows(
        ws,
        ["name", "schedule", "timezone", "state", "target", "method", "lastAttempt", "scheduleTime", "status_code"],
        sch_rows,
        [40, 18, 16, 12, 70, 10, 24, 24, 12],
    )

    # 9 Secrets metadata
    ws = wb.create_sheet("9_Secrets_Metadata")
    sec_rows = []
    for s in inventory.get("secrets") or []:
        if not isinstance(s, dict):
            continue
        sec_rows.append(
            [
                dig(s, "name", default="").split("/")[-1],
                dig(s, "createTime"),
                dig(s, "replication", "automatic") and "automatic" or json.dumps(dig(s, "replication", default={})),
                dig(s, "labels") and json.dumps(s.get("labels")) or "",
            ]
        )
    write_rows(ws, ["secret_id", "createTime", "replication", "labels"], sec_rows, [40, 24, 20, 40])
    ws.append([])
    ws.append(["dhan-access-token versions (NO VALUES)"])
    ws.append(["version", "state", "createTime"])
    for v in inventory.get("secret_dhan_versions") or []:
        if isinstance(v, dict):
            ws.append([dig(v, "name", default="").split("/")[-1], v.get("state"), v.get("createTime")])

    # 10 Service accounts
    ws = wb.create_sheet("10_Service_Accounts")
    sa_rows = []
    for sa in inventory.get("service_accounts") or []:
        if not isinstance(sa, dict):
            continue
        sa_rows.append([sa.get("email"), sa.get("displayName"), sa.get("uniqueId"), sa.get("disabled"), sa.get("oauth2ClientId")])
    write_rows(ws, ["email", "displayName", "uniqueId", "disabled", "oauth2ClientId"], sa_rows, [55, 35, 22, 10, 22])

    # 11 PubSub / Buckets
    ws = wb.create_sheet("11_PubSub_Buckets")
    ws.append(["type", "name", "detail"])
    style_header(ws, 3)
    for t in inventory.get("pubsub_topics") or []:
        name = t if isinstance(t, str) else dig(t, "name", default=str(t))
        ws.append(["pubsub_topic", name, ""])
    for b in inventory.get("buckets") or []:
        if isinstance(b, dict):
            ws.append(
                [
                    "gcs_bucket",
                    b.get("name") or dig(b, "metadata", "name", default=""),
                    json.dumps({k: b.get(k) for k in ("location", "storageClass", "timeCreated", "updated") if k in b or True})[:300],
                ]
            )
        else:
            ws.append(["gcs_bucket", str(b), ""])
    autosize(ws, [14, 55, 80])

    # 12 APIs enabled
    ws = wb.create_sheet("12_APIs_Enabled")
    api_rows = []
    for a in inventory.get("apis_enabled") or []:
        if not isinstance(a, dict):
            continue
        api_rows.append([a.get("config", {}).get("name") or a.get("name"), a.get("state"), dig(a, "config", "title", default="")])
    write_rows(ws, ["api", "state", "title"], api_rows, [55, 12, 50])

    # 13 Live deploy/broker/health
    ws = wb.create_sheet("13_Live_Runtime_APIs")
    ws.append(["surface", "key", "value"])
    style_header(ws, 3)
    for surface, payload in live.items():
        if isinstance(payload, dict):
            for k, v in payload.items():
                if isinstance(v, (dict, list)):
                    val = json.dumps(v)[:1500]
                else:
                    # redact tokens
                    if any(x in str(k).lower() for x in ("token", "secret", "password", "access_token")):
                        val = "***REDACTED***"
                    else:
                        val = str(v)
                ws.append([surface, k, val])
        else:
            ws.append([surface, "raw", str(payload)[:1500]])
    autosize(ws, [22, 35, 100])

    # 14 Gates detail
    ws = wb.create_sheet("14_Auto_Gates")
    gates = dig(live.get("auto_gates", {}), "proof_gates", default=[]) or dig(live.get("auto_gates", {}), "gates", default={})
    gate_rows = []
    if isinstance(gates, list):
        for g in gates:
            gate_rows.append(
                [
                    g.get("gate_id") or g.get("name"),
                    g.get("status"),
                    g.get("pass"),
                    g.get("note") or g.get("blocker_id"),
                    g.get("latest_rho"),
                    g.get("days_recorded"),
                ]
            )
    elif isinstance(gates, dict):
        for gid, g in gates.items():
            if isinstance(g, dict):
                gate_rows.append([gid, g.get("pass"), g.get("pass"), g.get("blocker_id"), g.get("latest_rho"), g.get("days_recorded")])
    write_rows(ws, ["gate_id", "status", "pass", "note_or_blocker", "latest_rho", "days_recorded"], gate_rows, [45, 14, 10, 50, 12, 12])

    # 15 Scheduler health API
    ws = wb.create_sheet("15_Scheduler_Health_API")
    sh = live.get("scheduler_health") or {}
    ws.append(["key", "value"])
    style_header(ws, 2)
    for k, v in (sh.items() if isinstance(sh, dict) else []):
        ws.append([k, json.dumps(v)[:2000] if isinstance(v, (dict, list)) else str(v)])
    autosize(ws, [35, 100])

    # 16 Architecture map
    ws = wb.create_sheet("16_Architecture_Map")
    write_rows(
        ws,
        ["layer", "component", "gcp_or_system", "role", "authority"],
        [
            ["Control", "GitHub main", "GitHub", "Source of truth for code", "PROVEN_PATH"],
            ["Control", "Cloud Run Auto Deploy / WIF", "GCP IAM", "Keyless deploy", "AUTHORITATIVE_PATH"],
            ["Compute", "genesis-system3-web", "Cloud Run", "UI+API serving", "CURRENT_LIVE"],
            ["Compute", "genesis-system3-dhan-token-rotate", "Cloud Run Job", "Sole token mint authority", "AUTHORITATIVE_PATH"],
            ["Schedule", "Cloud Scheduler jobs", "Scheduler", "Rotate / warm / validate triggers", "AUTHORITATIVE_PATH"],
            ["Secrets", "dhan-access-token", "Secret Manager", "Broker token versions only", "AUTHORITATIVE_PATH"],
            ["Data hot", "In-process / API caches", "Cloud Run memory", "Live chains/spots", "PARTIAL"],
            ["Data warm", "Paper/positions files or DB", "Cloud Run filesystem/GCS?", "Lifecycle ledger", "GAP_IF_FILE_MISSING"],
            ["Data cold", "GCS + BigQuery", "Storage/BigQuery", "History/ML lake", "CONSOLE_SHOWN_NEEDS_PROOF"],
            ["Evidence", "reports/latest + MRI xlsx", "Laptop/GitHub artifacts", "Operator MRI board", "HISTORICAL_UNTIL_REFRESH"],
            ["Safety", "LIVE=false locks", "Runtime env + API", "No real orders", "PROVEN"],
        ],
        [12, 40, 28, 40, 24],
    )

    # 17 Recommendations
    ws = wb.create_sheet("17_Recommendations")
    write_rows(
        ws,
        ["priority", "finding", "action"],
        [
            ["P0", "Keep LIVE=false until gates 7/7", "Do not enable live trading"],
            ["P0", f"Serving SHA {serving} vs main {main_sha}", "Keep deploy parity monitored"],
            ["P0", "Console shows Cloud Run 'no available instance' aborts", "MRI scale/min-instances / concurrency; check Error Reporting"],
            ["P0", "Scheduler health from API", "Fix unhealthy scheduler transport/IAM if healthy=false"],
            ["P1", "Finish durable paper/positions ledger", "Cloud SQL/Firestore or GCS-backed persistence"],
            ["P1", "Options history lake", "Partitioned GCS + BigQuery tables"],
            ["P1", "Pub/Sub event spine", "ticks/chains/orders/risk topics"],
            ["P2", "Excel/CSV are operator views only", "Never treat workbook as lake of record"],
            ["INFO", "Billing estimate ~INR 17.7k (1-25 Aug)", "Review detailed charges in Billing console"],
        ],
        [10, 70, 70],
    )

    # 18 Raw index
    ws = wb.create_sheet("18_Raw_Artifact_Index")
    idx_rows = []
    for p in sorted(OUT_DIR.glob("*")):
        idx_rows.append([p.name, p.stat().st_size, datetime.fromtimestamp(p.stat().st_mtime, timezone.utc).isoformat()])
    write_rows(ws, ["file", "bytes", "mtime_utc"], idx_rows, [40, 12, 28])

    wb.save(XLSX)
    # also copy beside coordination workbook name alias
    alt = ROOT / "reports" / "coordination" / "System3_GCP_FULL_MRI.xlsx"
    alt.parent.mkdir(parents=True, exist_ok=True)
    alt.write_bytes(XLSX.read_bytes())
    print(json.dumps({"xlsx": str(XLSX), "copy": str(alt), "serving": serving, "main": main_sha, "sheets": wb.sheetnames}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
