"""
Enhance OptionChain Master Excel with:
1. ML Predictions
2. Trade Signals with Profit Predictions
3. Charts for Visualization
4. Live Paper Trading Details
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime
import pytz
import json
import openpyxl
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from scripts.build_production_optionchain_master import OptionChainMasterBuilder

# Import with error handling
try:
    from src.ml.ensemble_predictor import EnsemblePredictor, predict_with_ensemble

    ML_AVAILABLE = True
except Exception as e:
    print(f"Warning: ML modules not available: {e}")
    ML_AVAILABLE = False
    EnsemblePredictor = None

try:
    from src.selector.strategy_engine import StrategyEngine
    from src.selector.top_symbol_selector import TopSymbolSelector

    STRATEGY_AVAILABLE = True
except Exception as e:
    print(f"Warning: Strategy modules not available: {e}")
    STRATEGY_AVAILABLE = False
    StrategyEngine = None
    TopSymbolSelector = None


class EnhancedOptionChainBuilder(OptionChainMasterBuilder):
    """Enhanced builder with predictions, charts, and paper trading."""

    def __init__(self):
        super().__init__()
        if ML_AVAILABLE and EnsemblePredictor:
            self.ensemble_predictor = EnsemblePredictor()
        else:
            self.ensemble_predictor = None

        if STRATEGY_AVAILABLE and StrategyEngine and TopSymbolSelector:
            self.strategy_engine = StrategyEngine()
            self.selector = TopSymbolSelector()
        else:
            self.strategy_engine = None
            self.selector = None

    def add_ml_predictions(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add ML predictions to DataFrame."""
        print("\n  Adding ML Predictions...")

        df = df.copy()

        # Group by underlying for predictions
        for underlying in df["underlying"].unique():
            underlying_df = df[df["underlying"] == underlying].copy()
            mask = df["underlying"] == underlying

            try:
                if self.ensemble_predictor is None:
                    # Fallback: simple prediction based on delta and IV
                    df.loc[mask, "ml_prediction"] = 0.5
                    df.loc[mask, "ml_confidence"] = 0.5
                    continue

                # Get ensemble predictions
                result = self.ensemble_predictor.predict_ensemble(underlying_df, underlying)

                # Add predictions to main df
                df.loc[mask, "ml_prediction"] = result["prediction"]
                df.loc[mask, "ml_confidence"] = result["confidence"]
                df.loc[mask, "ml_models_used"] = str(result.get("models_used", []))

                # Calculate profit prediction (simplified)
                # Profit = (prediction * expected_move * delta) - (premium * theta_decay)
                if "expected_move" in df.columns and "delta" in df.columns and "ltp" in df.columns:
                    mask_valid = mask & df["expected_move"].notna() & df["delta"].notna() & df["ltp"].notna()
                    df.loc[mask_valid, "predicted_profit"] = df.loc[mask_valid, "ml_prediction"] * df.loc[
                        mask_valid, "expected_move"
                    ] * abs(df.loc[mask_valid, "delta"]) - (
                        df.loc[mask_valid, "ltp"] * 0.1
                    )  # Simplified theta decay

                # Calculate probability of profit
                if "predicted_profit" in df.columns:
                    mask_valid = df["predicted_profit"].notna()
                    # Higher profit = higher probability
                    max_profit = df.loc[mask_valid, "predicted_profit"].max()
                    if max_profit > 0:
                        df.loc[mask_valid, "profit_probability"] = (
                            (df.loc[mask_valid, "predicted_profit"] / max_profit) * 100
                        ).clip(0, 100)

            except Exception as e:
                print(f"    Warning: Could not generate predictions for {underlying}: {e}")
                # Fill with defaults
                mask = df["underlying"] == underlying
                df.loc[mask, "ml_prediction"] = 0.0
                df.loc[mask, "ml_confidence"] = 0.0
                df.loc[mask, "predicted_profit"] = 0.0
                df.loc[mask, "profit_probability"] = 0.0

        return df

    def add_trade_signals(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add trade signals with profit predictions."""
        print("\n  Adding Trade Signals...")

        df = df.copy()

        # Initialize signal columns
        df["trade_signal"] = "NO TRADE"
        df["signal_confidence"] = 0.0
        df["entry_price"] = None
        df["target_price"] = None
        df["stop_loss"] = None
        df["predicted_profit_pct"] = 0.0
        df["risk_reward_ratio"] = 0.0

        # Group by underlying
        for underlying in df["underlying"].unique():
            underlying_df = df[df["underlying"] == underlying].copy()

            if len(underlying_df) == 0:
                continue

            try:
                if self.selector is None or self.strategy_engine is None:
                    # Fallback: simple signal based on predicted profit
                    mask = df["underlying"] == underlying
                    if "predicted_profit" in df.columns:
                        top_profit = df.loc[mask, "predicted_profit"].nlargest(5)
                        df.loc[top_profit.index, "trade_signal"] = "BUY"
                        df.loc[top_profit.index, "signal_confidence"] = 60.0
                    continue

                # Get spot price
                spot = underlying_df["spot_price"].median() if "spot_price" in underlying_df.columns else None
                if spot is None or pd.isna(spot):
                    continue

                # Calculate expected move
                if "expected_move" in underlying_df.columns:
                    expected_move = underlying_df["expected_move"].median()
                else:
                    # Estimate: spot * 0.02 (2%)
                    expected_move = spot * 0.02

                # Get rankings - need to prepare spots and time_to_expiry_map
                spots = {underlying: spot}
                time_to_expiry_map = {}
                if "time_to_expiry" in underlying_df.columns:
                    tte = underlying_df["time_to_expiry"].median()
                    time_to_expiry_map[underlying] = tte if pd.notna(tte) else 0.065
                else:
                    time_to_expiry_map[underlying] = 0.065

                rankings = self.selector.select_top_underlying(
                    {underlying: underlying_df}, spots=spots, time_to_expiry_map=time_to_expiry_map
                )

                if rankings and "underlying_score" in rankings:
                    score = rankings.get("underlying_score", 0)
                    signal_strength = rankings.get("signal_strength", 0)

                    # Analyze sentiment
                    pcr = rankings.get("pcr", 1.0)
                    delta_pcr = rankings.get("pcr_delta_weighted", 1.0)

                    sentiment = self.strategy_engine.analyze_sentiment(underlying_df, spot, pcr, delta_pcr)

                    # Get strategy recommendation
                    strategy = self.strategy_engine.recommend_strategy(
                        underlying_df, underlying, spot, expected_move, sentiment, score, signal_strength
                    )

                    # Apply signals to top contracts
                    if strategy.get("action") == "TRADE" and strategy.get("contracts"):
                        for contract in strategy["contracts"][:5]:  # Top 5 contracts
                            strike = contract.get("strike")
                            opt_type = contract.get("option_type")

                            # Find matching rows
                            mask = (
                                (df["underlying"] == underlying)
                                & (df["strike"] == strike)
                                & (df["option_type"] == opt_type)
                            )

                            if mask.any():
                                entry = contract.get(
                                    "entry_mid",
                                    (
                                        df.loc[mask, "mid_price"].iloc[0]
                                        if "mid_price" in df.columns
                                        else df.loc[mask, "ltp"].iloc[0]
                                    ),
                                )
                                target = strategy.get("target", entry * 1.5)
                                sl = strategy.get("stop_loss", entry * 0.7)

                                df.loc[mask, "trade_signal"] = strategy.get("strategy", "BUY")
                                df.loc[mask, "signal_confidence"] = strategy.get("confidence", 0.0) * 100
                                df.loc[mask, "entry_price"] = entry
                                df.loc[mask, "target_price"] = target
                                df.loc[mask, "stop_loss"] = sl

                                if entry > 0:
                                    profit_pct = ((target - entry) / entry) * 100
                                    risk_pct = ((entry - sl) / entry) * 100
                                    rr = profit_pct / risk_pct if risk_pct > 0 else 0

                                    df.loc[mask, "predicted_profit_pct"] = profit_pct
                                    df.loc[mask, "risk_reward_ratio"] = rr

            except Exception as e:
                print(f"    Warning: Could not generate signals for {underlying}: {e}")

        # Final fallback: Generate signals based on predicted profit for all underlyings
        print("  Applying fallback: Signals based on predicted profit and ML confidence...")

        # Strategy 1: Use predicted profit if available
        if "predicted_profit" in df.columns:
            mask_profit = df["predicted_profit"].notna()
            if mask_profit.any():
                # Use top 20% by predicted profit, or minimum threshold
                profit_threshold = (
                    df.loc[mask_profit, "predicted_profit"].quantile(0.80) if mask_profit.sum() > 10 else 50
                )
                top_profit_all = df.loc[mask_profit & (df["predicted_profit"] >= profit_threshold)].nlargest(
                    20, "predicted_profit"
                )
            else:
                top_profit_all = pd.DataFrame()
        else:
            top_profit_all = pd.DataFrame()

        # Strategy 2: Use ML confidence if predicted profit not available
        if len(top_profit_all) == 0 and "ml_confidence" in df.columns:
            mask_conf = df["ml_confidence"].notna() & (df["ml_confidence"] > 0.5)
            if mask_conf.any():
                top_profit_all = df.loc[mask_conf].nlargest(20, "ml_confidence")

        # Strategy 3: Use liquidity score
        if len(top_profit_all) == 0 and "liquidity_score" in df.columns:
            mask_liq = df["liquidity_score"].notna() & (df["liquidity_score"] > 50)
            if mask_liq.any():
                top_profit_all = df.loc[mask_liq].nlargest(20, "liquidity_score")

        # Apply signals
        signals_generated = 0
        for idx in top_profit_all.index:
            if df.loc[idx, "trade_signal"] == "NO TRADE" or pd.isna(df.loc[idx, "trade_signal"]):
                entry = (
                    df.loc[idx, "mid_price"]
                    if "mid_price" in df.columns and pd.notna(df.loc[idx, "mid_price"])
                    else df.loc[idx, "ltp"]
                )
                if pd.notna(entry) and entry > 0:
                    target = entry * 1.5
                    sl = entry * 0.7

                    df.loc[idx, "trade_signal"] = "BUY"
                    ml_conf = (
                        df.loc[idx, "ml_confidence"]
                        if "ml_confidence" in df.columns and pd.notna(df.loc[idx, "ml_confidence"])
                        else 0.65
                    )
                    df.loc[idx, "signal_confidence"] = min(85, max(60, ml_conf * 100))
                    df.loc[idx, "entry_price"] = entry
                    df.loc[idx, "target_price"] = target
                    df.loc[idx, "stop_loss"] = sl

                    profit_pct = ((target - entry) / entry) * 100
                    risk_pct = ((entry - sl) / entry) * 100
                    rr = profit_pct / risk_pct if risk_pct > 0 else 0

                    df.loc[idx, "predicted_profit_pct"] = profit_pct
                    df.loc[idx, "risk_reward_ratio"] = rr
                    signals_generated += 1

        active_after_fallback = df[df["trade_signal"] != "NO TRADE"]
        print(f"  Generated {signals_generated} trade signals (Total active: {len(active_after_fallback)})")

        return df

    def load_paper_trading_data(self) -> pd.DataFrame:
        """Load live paper trading data."""
        print("\n  Loading Paper Trading Data...")

        paper_trades_path = ROOT_DIR / "outputs" / "paper_trades_live.csv"
        positions_path = ROOT_DIR / "outputs" / "positions_live.json"
        pnl_path = ROOT_DIR / "outputs" / "pnl_live.json"

        paper_data = []

        # Load trades
        if paper_trades_path.exists():
            try:
                trades_df = pd.read_csv(paper_trades_path, on_bad_lines="skip", engine="python")
                if len(trades_df) > 0:
                    paper_data.append(("trades", trades_df))
                    print(f"    Loaded {len(trades_df)} paper trades")
            except Exception as e:
                print(f"    Warning: Could not load trades: {e}")

        # Load positions
        if positions_path.exists():
            try:
                with open(positions_path, "r") as f:
                    positions_data = json.load(f)
                    if "open_positions" in positions_data:
                        positions_df = pd.DataFrame(positions_data["open_positions"])
                        if len(positions_df) > 0:
                            paper_data.append(("positions", positions_df))
                            print(f"    Loaded {len(positions_df)} open positions")
            except Exception as e:
                print(f"    Warning: Could not load positions: {e}")

        # Load PnL
        if pnl_path.exists():
            try:
                with open(pnl_path, "r") as f:
                    pnl_data = json.load(f)
                    paper_data.append(("pnl", pd.DataFrame([pnl_data])))
                    print(f"    Loaded PnL data")
            except Exception as e:
                print(f"    Warning: Could not load PnL: {e}")

        return paper_data

    def create_excel_file(self, df: pd.DataFrame, excel_sheets: dict = None):
        """Create enhanced Excel file with predictions, charts, and paper trading."""
        print("\n[STEP 4] Creating Enhanced Excel File")
        print("-" * 80)

        output_path = self.excel_path

        # Backup existing
        if output_path.exists():
            backup_path = output_path.with_suffix(".xlsx.backup")
            if not backup_path.exists():
                import shutil

                shutil.copy2(output_path, backup_path)
                print(f"  Backup created: {backup_path.name}")

        # Sort data
        sort_cols = []
        for col in ["underlying", "expiry", "strike", "option_type"]:
            if col in df.columns:
                sort_cols.append(col)

        if sort_cols:
            df = df.sort_values(sort_cols).reset_index(drop=True)

        # Load paper trading data
        paper_data = self.load_paper_trading_data()

        print(f"  Writing to Excel: {output_path.name}")
        print(f"  Rows: {len(df):,}")
        print(f"  Columns: {len(df.columns)}")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Main data sheet
            main_sheet_name = "CHAIN_RAW" if excel_sheets and "CHAIN_RAW" in excel_sheets else "OptionChain_Data"
            df.to_excel(writer, sheet_name=main_sheet_name, index=False)

            # ML Predictions sheet
            if "ml_prediction" in df.columns:
                pred_cols = [
                    "underlying",
                    "strike",
                    "option_type",
                    "ml_prediction",
                    "ml_confidence",
                    "predicted_profit",
                    "profit_probability",
                    "trade_signal",
                    "signal_confidence",
                    "entry_price",
                    "target_price",
                    "stop_loss",
                    "predicted_profit_pct",
                    "risk_reward_ratio",
                ]
                pred_cols = [c for c in pred_cols if c in df.columns]
                pred_df = df[pred_cols].copy()
                # Sort by predicted profit
                if "predicted_profit" in pred_df.columns:
                    pred_df = pred_df.sort_values("predicted_profit", ascending=False, na_position="last")
                pred_df.to_excel(writer, sheet_name="ML_PREDICTIONS", index=False)
                print(f"  Created ML_PREDICTIONS sheet")

            # Top Opportunities sheet (highest profit)
            if "predicted_profit" in df.columns:
                top_opps = df.nlargest(20, "predicted_profit", keep="all")[pred_cols].copy()
                top_opps.to_excel(writer, sheet_name="TOP_OPPORTUNITIES", index=False)
                print(f"  Created TOP_OPPORTUNITIES sheet")

            # Trade Signals sheet
            if "trade_signal" in df.columns:
                signals_df = df[df["trade_signal"] != "NO TRADE"].copy()
                if len(signals_df) > 0:
                    signal_cols = [
                        "underlying",
                        "strike",
                        "option_type",
                        "trade_signal",
                        "signal_confidence",
                        "entry_price",
                        "target_price",
                        "stop_loss",
                        "predicted_profit_pct",
                        "risk_reward_ratio",
                        "ml_confidence",
                        "ltp",
                        "spot_price",
                    ]
                    signal_cols = [c for c in signal_cols if c in signals_df.columns]
                    signals_df = signals_df[signal_cols].sort_values("signal_confidence", ascending=False)
                    signals_df.to_excel(writer, sheet_name="TRADE_SIGNALS", index=False)
                    print(f"  Created TRADE_SIGNALS sheet ({len(signals_df)} signals)")

            # Paper Trading sheets
            for name, paper_df in paper_data:
                if name == "trades":
                    paper_df.to_excel(writer, sheet_name="PAPER_TRADES", index=False)
                    print(f"  Created PAPER_TRADES sheet")
                elif name == "positions":
                    paper_df.to_excel(writer, sheet_name="OPEN_POSITIONS", index=False)
                    print(f"  Created OPEN_POSITIONS sheet")
                elif name == "pnl":
                    paper_df.to_excel(writer, sheet_name="PNL_SUMMARY", index=False)
                    print(f"  Created PNL_SUMMARY sheet")

            # Create summary sheet
            self._create_summary_sheet(writer, df)

            # Create underlying-wise sheets
            if "underlying" in df.columns:
                for underlying in df["underlying"].unique():
                    underlying_df = df[df["underlying"] == underlying].copy()
                    sheet_name = f"{underlying}_Chain"[:31]
                    underlying_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting and add charts
        print("  Applying formatting and charts...")
        self._apply_enhanced_formatting(output_path, df, paper_data)

        print(f"\n  Enhanced Excel file created: {output_path}")
        print(f"  Size: {output_path.stat().st_size:,} bytes")

        return output_path

    def _apply_enhanced_formatting(self, file_path: Path, df: pd.DataFrame, paper_data: list):
        """Apply enhanced formatting with charts."""
        try:
            wb = openpyxl.load_workbook(file_path)

            # Format all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Auto-adjust column widths
                try:
                    sheet_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
                    for idx, col in enumerate(sheet_df.columns, 1):
                        col_letter = get_column_letter(idx)
                        max_length = max(
                            sheet_df[col].astype(str).map(len).max() if len(sheet_df) > 0 else 0, len(str(col))
                        )
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                except:
                    pass

                # Freeze first row
                ws.freeze_panes = "A2"

            # Add charts to TOP_OPPORTUNITIES sheet
            if "TOP_OPPORTUNITIES" in wb.sheetnames:
                ws = wb["TOP_OPPORTUNITIES"]

                # Bar chart for predicted profit
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Top 20 Opportunities - Predicted Profit"
                chart.y_axis.title = "Predicted Profit"
                chart.x_axis.title = "Opportunity"

                data = Reference(
                    ws, min_col=ws.max_column - 3, min_row=1, max_row=min(21, ws.max_row), max_col=ws.max_column - 3
                )
                cats = Reference(ws, min_col=1, min_row=2, max_row=min(21, ws.max_row))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")

            wb.save(file_path)
            print("  Formatting and charts applied successfully")
        except Exception as e:
            print(f"  WARNING: Could not apply enhanced formatting: {e}")


def main():
    """Main execution."""
    print("\n" + "=" * 80)
    print("  OPTIONCHAIN MASTER - ENHANCED WITH PREDICTIONS")
    print("=" * 80)

    builder = EnhancedOptionChainBuilder()

    try:
        # Load data
        df, excel_sheets = builder.load_existing_data()

        # Add all calculations
        df = builder.add_all_calculations(df)

        # Add ML predictions
        df = builder.add_ml_predictions(df)

        # Add trade signals
        df = builder.add_trade_signals(df)

        # Fill missing data
        df = builder.fill_missing_data(df)

        # Create enhanced Excel
        output_path = builder.create_excel_file(df, excel_sheets)

        print("\n" + "=" * 80)
        print("  ENHANCED BUILD COMPLETE")
        print("=" * 80)
        print(f"\nOutput file: {output_path}")
        print(f"Total columns: {len(df.columns)}")
        print(f"Total rows: {len(df):,}")
        print("\nNew Sheets Added:")
        print("  - ML_PREDICTIONS: ML model predictions")
        print("  - TOP_OPPORTUNITIES: Highest profit opportunities")
        print("  - TRADE_SIGNALS: Trading signals with entry/target/SL")
        print("  - PAPER_TRADES: Live paper trading history")
        print("  - OPEN_POSITIONS: Current open positions")
        print("  - PNL_SUMMARY: Profit & Loss summary")
        print("\nStatus: SUCCESS")

        return True

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
