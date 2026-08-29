#!/usr/bin/env python3
"""Build overwrite-only AGENT_OPERATING_OPTIONS.xlsx + companion CSVs.

Sheets:
  1_User_Actions      - what user must do (minimal)
  2_Options_Priority  - alternative solutions ranked agent-first
  3_Pending_Live      - from TRACKING_CHECKLIST.json if present
  4_UI_Tab_Impact     - tab-by-tab impact
  5_GCP_GitHub_Levers - control levers
  6_Progress_Chart    - confidence/progress series for Excel charts
  7_Failure_Playbook  - unpredictable failure responses
  8_MD_Upgrades       - docs that train agents

Always overwrites reports/coordination/AGENT_OPERATING_OPTIONS.xlsx
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "reports" / "coordination"
TRACK_JSON = OUT / "TRACKING_CHECKLIST.json"


def style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col[:40]:
            width = max(width, min(max_width, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    autosize(ws)


def load_pending() -> list[list]:
    if not TRACK_JSON.exists():
        return [["NO_TRACKING_JSON", "OPEN", "P0", "Run tracker first", "", "", ""]]
    data = json.loads(TRACK_JSON.read_text(encoding="utf-8"))
    rows = []
    for r in data.get("rows") or []:
        rows.append(
            [
                r.get("id"),
                r.get("pri"),
                r.get("status"),
                r.get("title"),
                r.get("live_proof"),
                r.get("need_user"),
                r.get("rec"),
            ]
        )
    return rows


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat()
    live = {}
    if TRACK_JSON.exists():
        live = json.loads(TRACK_JSON.read_text(encoding="utf-8")).get("live") or {}

    wb = Workbook()
    # remove default
    wb.remove(wb.active)

    # --- Sheet 1: User actions (minimal involvement) ---
    write_sheet(
        wb,
        "1_User_Actions",
        [
            "priority_rank",
            "action",
            "why",
            "blocks_agent_if_skipped",
            "time_needed",
            "one_time_or_repeat",
            "command_or_path",
        ],
        [
            [1, "Open Cursor on primary clone only", "Broken Git at C:\\System3\\...", "YES", "1 min", "every session", r"C:\Users\ADMIN\Genesis_System3\Genesis_System3"],
            [2, "Reply: commit + PR (or approve PR)", "Local OptionChain/API aliases not on Cloud Run", "YES for UI change", "2 min", "per batch", "Chat: commit + PR"],
            [3, "Keep LIVE OFF", "Gates 2/7; safety", "NO (agent refuses LIVE)", "0", "always", "UI Paper mode"],
            [4, "Keep Dhan web logged in when asked", "Parity visual proof", "SOFT", "0 (already open)", "on request", "https://web.dhan.co/"],
            [5, "Decide API key policy PEND-025", "public_readonly vs enforce", "SOFT", "5 min", "once", "Enforce or document"],
            [6, "Refresh Claude project memory", "Stops Aug16 myth loops", "SOFT", "10 min", "once", "Claude project memory"],
            [7, "Optional: want RUHI board on UI?", "PEND-026 product choice", "SOFT", "1 min", "once", "Yes/No in chat"],
        ],
    )

    # --- Sheet 2: Options priority (agent-first) ---
    # Priority score: lower = better. Prefer agent-only / little user.
    write_sheet(
        wb,
        "2_Options_Priority",
        [
            "option_id",
            "priority_rank",
            "user_involvement",
            "who_triggers",
            "auto_trigger",
            "timeframe",
            "goal_slice",
            "alternative_solution",
            "proof_after",
            "ui_tabs_impacted",
            "depends_on",
            "confidence_now_to_after",
            "recommended",
        ],
        [
            ["OPT-A1", 1, "LOW (approve PR once)", "Agent", "post-edit command_center", "2–6 h", "Dhan chain parity Batch-1", "PR OptionChain+holdings/funds aliases → Auto Deploy → re-snap", "serving SHA change + chain headers + TRACKING DONE", "Trade,Chain,Broker APIs", "User: commit+PR", "35%→70%", "YES — FIRST"],
            ["OPT-A2", 2, "NONE", "Agent", "hourly tracker + post-edit", "continuous", "Truth board", "Overwrite TRACKING_CHECKLIST from live APIs", "OPEN/DONE counts refresh", "All (status)", "None", "50%→60%", "YES"],
            ["OPT-A3", 3, "NONE", "Agent", "command_center", "4–12 h", "Deploy lag PEND-001", "MRI Auto Deploy / Actions; fix gate; catch serving to main", "deploy_info==main", "All tabs get new build", "gh+gcloud", "40%→65%", "YES"],
            ["OPT-A4", 4, "NONE", "Agent", "command_center", "1–2 d", "Scheduler health PEND-002", "MRI IAM/Scheduler invoke; named gate green", "scheduler/health healthy", "Paper/ML jobs", "gcloud", "30%→55%", "YES"],
            ["OPT-A5", 5, "NONE", "Agent", "post-edit", "2–5 d", "Paper lifecycle 014–017", "Fix cloud positions file path + lifecycle proof job", "paper open or honest + gate", "Paper,Positions", "deploy", "25%→60%", "YES"],
            ["OPT-A6", 6, "NONE", "Agent", "post-edit", "3–7 d", "Signals 021", "Persist scanner→signal; rate-limit 429", "signals not NO_TRADE file-missing", "Signals,Paper", "deploy", "20%→50%", "YES"],
            ["OPT-A7", 7, "NONE", "Agent", "market-hours jobs", "2–4 wk", "ML gates 018–020", "daily_gain_validate + retrain; never weaken thresholds", "Spearman/expectancy pass", "Performance,ML,Gates", "paper data", "15%→45%", "YES (slow)"],
            ["OPT-B1", 8, "MEDIUM", "User+Agent", "manual", "1 d", "API key", "Mount Secret Manager key + REQUIRE_API_KEY", "auth/status required=true", "All APIs", "User decision", "N/A", "OPTIONAL"],
            ["OPT-B2", 9, "MEDIUM", "User", "manual", "30 min", "Claude memory", "Delete Aug16 SHAs; point to TRACKING_CHECKLIST", "Agents stop myth loops", "Ops", "User", "N/A", "OPTIONAL"],
            ["OPT-C1", 10, "HIGH — avoid", "User", "n/a", "risky", "Force LIVE", "Break-glass LIVE before 7/7", "NOT ALLOWED", "Trade", "Human break-glass", "0%", "NO"],
            ["OPT-C2", 11, "HIGH — avoid", "Agent", "n/a", "tech debt", "Weaken gates", "Lower Spearman/expectancy", "False PASS", "Gates", "Forbidden", "0%", "NO"],
            ["OPT-C3", 12, "HIGH — avoid", "Laptop mint", "n/a", "unsafe", "Mint token on laptop", "Local .env mint", "Authority violation", "Broker", "Forbidden", "0%", "NO"],
            ["OPT-A8", 5, "NONE", "Agent", "post-edit", "1–3 d", "Charts/predictions 404", "Implement routes OR label MISSING on UI", "no false-green", "Charts,Performance", "deploy", "30%→50%", "YES"],
            ["OPT-A9", 6, "LOW", "Agent", "post-edit", "3–7 d", "Equity security_id", "Scrip master map → equity chain", "equity CE load vs Dhan", "Trade equity panel", "Dhan master", "25%→55%", "YES"],
            ["OPT-A10", 4, "NONE", "Agent+Sched", "hourly+post-edit", "same day", "Single info source", "command_center_refresh replaces all probes", "one MD+JSON+XLSX", "Ops", "None", "55%→80%", "YES — CORE"],
        ],
    )

    # --- Sheet 3: Pending live ---
    write_sheet(
        wb,
        "3_Pending_Live",
        ["issue_id", "priority", "status", "title", "live_proof", "need_from_user", "recommendation"],
        load_pending(),
    )

    # --- Sheet 4: UI tab impact ---
    write_sheet(
        wb,
        "4_UI_Tab_Impact",
        ["tab", "url", "present_state", "after_OPT_A1_deploy", "later_batches", "proof_method"],
        [
            ["Trade", "/ui/?tab=trade", "Top CE OK; chain columns incomplete", "ATM±10 + LTP% Buildup Greeks", "Equity security_id", "snap + headers"],
            ["Option Chain", "/ui/?tab=chain", "Stale snapshot risk; deep OTM default", "STALE honesty + ATM default + columns", "Tick health gate", "snap vs Dhan"],
            ["Paper", "/ui/?tab=paper", "0 open; Positions file not found", "unchanged until A5", "Lifecycle proof", "API+snap"],
            ["Positions", "/ui/?tab=positions", "Paper 0; Dhan 1 shown on broker", "unchanged until A5", "Align files", "snap"],
            ["Broker", "/ui/?tab=broker", "AUTH_OK v319; holdings UI ok", "holdings/funds API 200", "Watch rotate", "API"],
            ["Multibagger", "/ui/?tab=multibagger", "0 candidates Delayed", "honest Delayed", "Research pipeline", "snap"],
            ["Signals", "/ui/?tab=signals", "NO_TRADE / file or 429", "unchanged until A6", "Persist signals", "API"],
            ["Performance/ML", "/ui/?tab=performance / ml", "0 preds; rho low", "honest BLOCKED", "A7 validate", "auto_gates"],
            ["Overview/Gates", "/ui/?tab=overview / gates", "2/7 LIVE OFF", "same until gates rise", "A4–A7", "auto_gates"],
            ["Charts", "/ui/?tab=charts", "404/empty", "MISSING label or route", "A8", "HTTP"],
        ],
    )

    # --- Sheet 5: GCP / GitHub levers ---
    write_sheet(
        wb,
        "5_GCP_GitHub_Levers",
        ["lever", "system", "best_practice", "why_fast", "agent_can_run", "user_needed", "schedule"],
        [
            ["origin/main SHA", "GitHub", "Only merge via PR+CI", "Prevents broken serve", "YES", "approve PR", "per PR"],
            ["Auto Deploy", "GitHub Actions→Cloud Build", "Image tag = full 40-char SHA", "Serving matches code", "YES MRI", "if secrets missing", "on merge"],
            ["deploy_info", "Cloud Run", "Public truth of serving", "No laptop PASS", "YES", "None", "on demand"],
            ["Secret Manager dhan-access-token", "GCP", "Rotate via job not laptop", "Auth without exposure", "YES workflow", "None", "scheduled"],
            ["Cloud Run service account", "GCP IAM", "Least privilege + soft-fail binds", "Deploy stays green", "YES", "None", "on deploy"],
            ["Cloud Scheduler jobs", "GCP", "Named health gate", "Jobs drive paper/ML", "YES MRI", "None", "IST schedule"],
            ["Firestore gate evidence", "GCP", "auto_gates runtime_driven", "Gates not laptop DB", "YES read", "None", "continuous"],
            ["Artifact Registry", "GCP", "asia-south1 containers", "Region latency", "YES deploy", "None", "on build"],
            ["TRACKING_CHECKLIST", "Laptop+repo", "Overwrite-only", "One source", "YES", "None", "hourly+post-edit"],
            ["Issue #188", "GitHub", "RUHI bus", "Multi-agent sync", "YES comment", "None", "on state change"],
            ["GCS for paper positions", "GCP Storage", "Prefer durable object over ephemeral container FS", "Fixes Positions file not found", "YES implement", "None", "with paper fix"],
            ["Cloud Logging", "GCP", "MRI blast radius", "Faster root cause", "YES", "None", "on incident"],
        ],
    )

    # --- Sheet 6: Progress chart data ---
    ws = wb.create_sheet("6_Progress_Chart")
    headers = ["metric", "present_pct", "after_batch1_pct", "after_paper_pct", "after_gates_pct", "target_pct"]
    ws.append(headers)
    metrics = [
        ["UI_Dhan_chain_parity", 35, 75, 80, 90, 95],
        ["Paper_truth_visible", 20, 25, 70, 85, 90],
        ["Deploy_truth_sync", 40, 80, 85, 95, 100],
        ["Scheduler_health", 25, 40, 70, 90, 95],
        ["ML_gate_readiness", 15, 20, 40, 80, 100],
        ["Agent_automation_leverage", 55, 75, 85, 90, 95],
        ["User_involvement_burden_inv", 40, 70, 80, 90, 95],
        ["Overall_goal_confidence", 28, 55, 68, 82, 92],
    ]
    for m in metrics:
        ws.append(m)
    style_header(ws, len(headers))
    autosize(ws)

    chart = LineChart()
    chart.title = "Confidence / Progress vs Batches"
    chart.style = 10
    chart.y_axis.title = "Confidence %"
    chart.x_axis.title = "Metric"
    data = Reference(ws, min_col=2, min_row=1, max_col=6, max_row=1 + len(metrics))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(metrics))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "H2")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Present vs Target (Overall Goal)"
    bar.y_axis.title = "%"
    # overall row is last
    last = 1 + len(metrics)
    data2 = Reference(ws, min_col=2, min_row=last, max_col=6, max_row=last)
    bar.add_data(data2, from_rows=True, titles_from_data=False)
    # simpler: chart present vs target for all metrics
    bar2 = BarChart()
    bar2.type = "col"
    bar2.grouping = "clustered"
    bar2.title = "Present vs Target by Metric"
    bar2.y_axis.title = "%"
    data3 = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=last)
    data4 = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=last)
    bar2.add_data(data3, titles_from_data=True)
    bar2.add_data(data4, titles_from_data=True)
    bar2.set_categories(cats)
    bar2.width = 18
    bar2.height = 10
    ws.add_chart(bar2, "H20")

    # --- Sheet 7: Failure playbook ---
    write_sheet(
        wb,
        "7_Failure_Playbook",
        ["failure", "detection", "auto_response", "user_needed", "fallback_truth", "prevent_repeat"],
        [
            ["CI/Auto Deploy red", "gh run / checklist PEND-001", "MRI logs; fix PR; soft-fail IAM", "If secrets/WIF", "Keep OPEN with proof", "Named gates"],
            ["Serving SHA stuck", "deploy_info vs main", "Rerun deploy; diagnose health canary", "Rare approve", "Label pre-deploy snaps", "Deploy lag row"],
            ["Tracker task dead", "checklist stale timestamp", "post-edit command_center; re-register task", "None", "Manual refresh PS1", "Hourly+post-edit"],
            ["Snap still wrong", "browser re-snap", "§10 investigate loop; new PR", "None", "Never DONE", "ATM/stale tests"],
            ["Broker DH-906", "broker/status", "Cloud rotate only", "None", "AUTH class on UI", "PR#303 pattern"],
            ["Wrong folder open", "git toplevel", "Refuse edit; switch primary", "Open correct folder", "Hard ban rule", "cursor rules"],
            ["Disk full", "probe", "Prefer E: worktree", "Free space", "Partial reports", "Worktree policy"],
            ["API 429 scanner", "signals/scanner", "Backoff + cache", "None", "Honest NO_TRADE", "Rate limit"],
        ],
    )

    # --- Sheet 8: MD upgrades ---
    write_sheet(
        wb,
        "8_MD_Upgrades",
        ["doc", "change_needed", "why_trains_agents_better", "status"],
        [
            ["SESSION_ISSUES_MASTER.md", "Point status to TRACKING_CHECKLIST only", "Stops stale status tables", "DONE"],
            ["SYSTEM3_MASTER_AUTOMATION_RUNBOOK.md §11", "Overwrite-only + post-edit trigger", "Stops duplicate commands", "DONE/extend"],
            ["session-issues-master.mdc", "Force read TRACKING + command_center", "Every session same path", "DONE/extend"],
            ["RUHI_RULE_V2.md", "Memory not authority; tracker is live", "Multi-agent sync", "DONE §17"],
            ["LIVE_PRIORITY_URLS.md", "Keep open tabs list", "UI proof targets", "exists"],
            ["AGENT_COMMAND_CENTER.md", "NEW one-page: run one script get all truth", "Replace manual probe spam", "CREATE"],
            ["AGENTS.md", "Link command_center + Excel path", "Repo entrypoint clarity", "UPDATE"],
        ],
    )

    # Meta sheet
    write_sheet(
        wb,
        "0_Meta",
        ["field", "value"],
        [
            ["generated_utc", now],
            ["serving_sha", live.get("serving_sha") or ""],
            ["gates", f"{live.get('gates_passing')}/{live.get('gates_total')}"],
            ["broker", f"{live.get('broker_auth')} v{live.get('broker_secret_version')}"],
            ["excel_path", str(OUT / "AGENT_OPERATING_OPTIONS.xlsx")],
            ["rule", "Overwrite only — do not create dated option workbooks"],
            ["first_priority", "OPT-A1 then OPT-A10 command_center"],
            ["thinking_before", "Inventory+local fix+tracker"],
            ["thinking_after_user_ask", "User-minimal options+Excel+auto one-source+post-edit trigger"],
        ],
    )

    # Move meta first
    wb.move_sheet("0_Meta", offset=-len(wb.sheetnames) + 1)

    xlsx = OUT / "AGENT_OPERATING_OPTIONS.xlsx"
    wb.save(xlsx)
    print(f"WROTE {xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
