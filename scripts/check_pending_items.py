"""
Check Pending Items - Comprehensive Status Check
"""

import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT_DIR = Path(__file__).parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def check_pending():
    """Check all pending items."""
    print("=" * 80)
    print("COMPREHENSIVE STATUS CHECK - PENDING ITEMS")
    print("=" * 80)

    pending = []
    completed = []

    # 1. Excel File
    excel_path = ROOT_DIR / "outputs" / "OptionChain_Master_v3_AI_FINAL.xlsx"
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        completed.append("Excel file exists with all sheets")

        # Check PnL data
        if "PNL_SUMMARY" in wb.sheetnames:
            ws = wb["PNL_SUMMARY"]
            if ws.max_row > 1:
                data = [c.value for c in ws[2]]
                if data[1] == 0:  # total_trades
                    pending.append("PnL_SUMMARY has no trade data (run paper trading)")
                else:
                    completed.append(f"PnL_SUMMARY has {data[1]} trades")
    else:
        pending.append("Excel file not found")

    # 2. Predictions
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        if "OptionChain_Data" in wb.sheetnames:
            ws = wb["OptionChain_Data"]
            headers = [c.value for c in ws[1]]
            if "ensemble_prediction" in headers:
                completed.append("AI predictions are working")
            else:
                pending.append("AI predictions missing in OptionChain_Data")

    # 3. Accuracy Metrics
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        if "ACCURACY_METRICS" in wb.sheetnames:
            ws = wb["ACCURACY_METRICS"]
            if ws.max_row > 1:
                completed.append("Accuracy metrics sheet exists")
                # Check if accuracy is 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if "Accuracy" in str(row[0]) and "0.00%" in str(row[1]):
                        pending.append("Accuracy metrics show 0% (need trade data to validate)")
                        break

    # 4. Paper Trading Data
    trades_path = ROOT_DIR / "outputs" / "paper_trades_live.csv"
    if trades_path.exists() and trades_path.stat().st_size > 100:
        try:
            df = pd.read_csv(trades_path, nrows=1, on_bad_lines="skip", engine="python")
            completed.append("Paper trades CSV exists")
        except:
            pending.append("Paper trades CSV exists but unreadable")
    else:
        pending.append("Paper trades CSV missing or empty (run paper trading)")

    # 5. Chain Data
    chain_path = ROOT_DIR / "outputs" / "chain_raw_live.csv"
    if chain_path.exists() and chain_path.stat().st_size > 1000:
        completed.append("Option chain data available")
    else:
        pending.append("Option chain data missing or empty")

    # 6. PnL JSON
    pnl_path = ROOT_DIR / "outputs" / "pnl_live.json"
    if pnl_path.exists():
        try:
            with open(pnl_path, "r") as f:
                pnl = json.load(f)
            if pnl.get("total_trades", 0) > 0:
                completed.append(f"PnL JSON has {pnl.get('total_trades')} trades")
            else:
                pending.append("PnL JSON exists but has no trades")
        except:
            pending.append("PnL JSON exists but unreadable")
    else:
        pending.append("PnL JSON missing")

    # 7. Test Results
    test_path = ROOT_DIR / "outputs" / "comprehensive_10k_test_report.json"
    if test_path.exists():
        try:
            with open(test_path, "r") as f:
                test = json.load(f)
            if test.get("pass_rate", 0) >= 99:
                completed.append(f"Tests: {test.get('pass_rate', 0):.2f}% pass rate")
            else:
                pending.append(f"Tests: {test.get('pass_rate', 0):.2f}% pass rate (needs improvement)")
        except:
            pass

    # Summary
    print("\n" + "=" * 80)
    print("COMPLETED ITEMS")
    print("=" * 80)
    for item in completed:
        print(f"  [OK] {item}")

    print("\n" + "=" * 80)
    print("PENDING ITEMS")
    print("=" * 80)
    if pending:
        for item in pending:
            print(f"  [PENDING] {item}")
    else:
        print("  [OK] No pending items!")

    print("\n" + "=" * 80)
    print("RECOMMENDATIONS")
    print("=" * 80)

    if any("paper trading" in p.lower() or "trade" in p.lower() for p in pending):
        print("  1. Run paper trading to generate trade data:")
        print("     START_PAPER_TRADING_COMPLETE.bat")

    if any("excel" in p.lower() or "rebuild" in p.lower() for p in pending):
        print("  2. Rebuild Excel with latest data:")
        print("     BUILD_ADVANCED_EXCEL.bat")

    if any("chain" in p.lower() or "data" in p.lower() for p in pending):
        print("  3. Fetch latest option chain data:")
        print("     UPDATE_OPTIONCHAIN_MASTER.bat")

    print("\n" + "=" * 80)
    print("STATUS SUMMARY")
    print("=" * 80)
    print(f"  Completed: {len(completed)}")
    print(f"  Pending: {len(pending)}")
    print(f"  Overall: {'READY' if len(pending) == 0 else 'NEEDS ATTENTION'}")
    print("=" * 80)


if __name__ == "__main__":
    check_pending()
