"""
Verify Excel Output - Show Proof from Actual File
No MD files - only show actual output
"""

import sys
from pathlib import Path
import openpyxl

ROOT_DIR = Path(__file__).parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def verify_excel():
    """Verify Excel file and show proof."""
    excel_path = ROOT_DIR / "outputs" / "OptionChain_Master_v3_AI_FINAL.xlsx"

    if not excel_path.exists():
        print("ERROR: Excel file not found!")
        return

    print("=" * 80)
    print("EXCEL FILE VERIFICATION - PROOF FROM ACTUAL FILE")
    print("=" * 80)

    wb = openpyxl.load_workbook(excel_path)

    print(f"\nFile: {excel_path.name}")
    print(f"Size: {excel_path.stat().st_size / (1024*1024):.2f} MB")
    print(f"\nTotal Sheets: {len(wb.sheetnames)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")

    # Verify PNL_SUMMARY
    print("\n" + "=" * 80)
    print("PNL_SUMMARY SHEET")
    print("=" * 80)
    if "PNL_SUMMARY" in wb.sheetnames:
        ws = wb["PNL_SUMMARY"]
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        print("\nHeaders:")
        headers = [c.value for c in ws[1]]
        print(f"  {headers}")
        if ws.max_row > 1:
            print("\nData (Row 2):")
            data = [c.value for c in ws[2]]
            print(f"  {data}")
    else:
        print("MISSING!")

    # Verify ACCURACY_METRICS
    print("\n" + "=" * 80)
    print("ACCURACY_METRICS SHEET")
    print("=" * 80)
    if "ACCURACY_METRICS" in wb.sheetnames:
        ws = wb["ACCURACY_METRICS"]
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        print("\nAll Data:")
        for row in ws.iter_rows(values_only=True):
            print(f"  {row}")
    else:
        print("MISSING!")

    # Verify AI_PREDICTIONS
    print("\n" + "=" * 80)
    print("AI_PREDICTIONS SHEET")
    print("=" * 80)
    if "AI_PREDICTIONS" in wb.sheetnames:
        ws = wb["AI_PREDICTIONS"]
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        if ws.max_row > 1:
            print("\nHeaders (first 8):")
            headers = [str(c.value)[:20] for c in ws[1]][:8]
            print(f"  {headers}")
            print("\nSample Data (Row 2, first 8 columns):")
            data = [str(c.value)[:20] if c.value else "" for c in ws[2]][:8]
            print(f"  {data}")
            print("\nSample Data (Row 3, first 8 columns):")
            data = [str(c.value)[:20] if c.value else "" for c in ws[3]][:8]
            print(f"  {data}")
    else:
        print("MISSING!")

    # Verify TOP_OPPORTUNITIES
    print("\n" + "=" * 80)
    print("TOP_OPPORTUNITIES SHEET")
    print("=" * 80)
    if "TOP_OPPORTUNITIES" in wb.sheetnames:
        ws = wb["TOP_OPPORTUNITIES"]
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        if ws.max_row > 1:
            print("\nHeaders:")
            headers = [c.value for c in ws[1]]
            print(f"  {headers}")
            print("\nTop 3 Opportunities:")
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
                print(f"  {i}: {row[:6]}")
    else:
        print("MISSING!")

    # Verify OptionChain_Data
    print("\n" + "=" * 80)
    print("OPTIONCHAIN_DATA SHEET")
    print("=" * 80)
    if "OptionChain_Data" in wb.sheetnames:
        ws = wb["OptionChain_Data"]
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        if ws.max_row > 1:
            # Check for prediction columns
            headers = [c.value for c in ws[1]]
            has_predictions = "ensemble_prediction" in headers
            has_confidence = "ensemble_confidence" in headers
            has_profit = "predicted_profit" in headers
            print(f"\nHas ensemble_prediction: {has_predictions}")
            print(f"Has ensemble_confidence: {has_confidence}")
            print(f"Has predicted_profit: {has_profit}")
            if has_predictions and ws.max_row > 1:
                # Get sample prediction values
                pred_col = headers.index("ensemble_prediction") + 1
                conf_col = headers.index("ensemble_confidence") + 1
                print(f"\nSample Predictions (first 5 rows):")
                for row_idx in range(2, min(7, ws.max_row + 1)):
                    pred_val = ws.cell(row=row_idx, column=pred_col).value
                    conf_val = ws.cell(row=row_idx, column=conf_col).value
                    print(f"  Row {row_idx}: prediction={pred_val:.3f}, confidence={conf_val:.3f}")
    else:
        print("MISSING!")

    print("\n" + "=" * 80)
    print("VERIFICATION COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    verify_excel()
