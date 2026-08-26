#!/usr/bin/env python3
"""SYSTEM3 Live Proof Center — single scheduled SSOT for all agents.

Purpose:
  Agents without laptop/gcloud/Dhan access still get a continuously refreshed,
  sanitized forensic pack on GitHub: APIs, dashboard tabs, GCP inventory (when
  WIF/gcloud available), and an Excel workbook.

Outputs (always overwritten):
  reports/latest/live_proof_center/LATEST/
    INDEX.md                         — agent landing page
    MANIFEST.json                    — machine SSOT
    System3_LIVE_PROOF_CENTER.xlsx   — 12 forensic sheets
    api/*.json                       — sanitized endpoint dumps
    dashboard_tabs.json
    gcp_inventory.json               — present when gcloud auth works
    CROSS_VERIFY.json                — main tip vs serving (best-effort)

Never exports secret values / tokens / private keys.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl required: pip install openpyxl ({exc})") from exc

ROOT = Path(__file__).resolve().parents[1]
BASE = os.environ.get(
    "SYSTEM3_PUBLIC_BASE_URL",
    "https://genesis-system3-web-doq2wplepa-el.a.run.app",
).rstrip("/")
PROJ = os.environ.get("GOOGLE_CLOUD_PROJECT", "system3-openalgo-safe")
REG = os.environ.get("GCP_REGION", "asia-south1")
SERVICE = os.environ.get("GCP_CLOUD_RUN_SERVICE", "genesis-system3-web")
OUT = ROOT / "reports" / "latest" / "live_proof_center" / "LATEST"
API_DIR = OUT / "api"
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
STAMP = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
UA = "system3-live-proof-center/1.0"

HF = PatternFill("solid", fgColor="1F4E79")
HFont = Font(color="FFFFFF", bold=True)
OKF = PatternFill("solid", fgColor="C6EFCE")
FAILF = PatternFill("solid", fgColor="FFC7CE")
WARNF = PatternFill("solid", fgColor="FFEB9C")

# 22 production dashboard tabs (frontend Sidebar.tsx)
DASHBOARD_TABS = [
    ("decision-intel", "Decision Intel", "main"),
    ("truth", "Truth Control", "main"),
    ("genesis", "Genesis Brain", "main"),
    ("e2e-proof", "E2E Proof", "main"),
    ("overview", "Overview", "main"),
    ("sim-live", "Sim Live", "main"),
    ("options-intel", "Options Intel", "market"),
    ("chain", "Option Chain", "market"),
    ("signals", "Signals", "market"),
    ("trade", "Trade", "trading"),
    ("paper", "Paper Trades", "trading"),
    ("positions", "Positions", "trading"),
    ("risk-scenarios", "Risk & Scenarios", "analysis"),
    ("multibagger", "Multibagger", "analysis"),
    ("prediction-audit", "Prediction Audit", "analysis"),
    ("performance", "Performance", "analysis"),
    ("ml", "ML Model", "analysis"),
    ("data-integrity", "Data Integrity", "system"),
    ("broker", "Broker", "system"),
    ("alerts", "Alerts", "system"),
    ("system", "System", "system"),
    ("gates", "Live Gate", "system"),
]

# Public / read-only APIs probed every run (mutation endpoints excluded)
API_ENDPOINTS = [
    ("deploy_info", "/api/deploy_info"),
    ("deploy_info_alias", "/api/deploy/info"),
    ("health", "/api/health"),
    ("state", "/api/state"),
    ("broker_status", "/api/broker/status"),
    ("broker_truth", "/api/broker/truth"),
    ("auth_status", "/api/auth/status"),
    ("system_health", "/api/system_health"),
    ("scheduler_health", "/api/scheduler/health"),
    ("auto_gates", "/api/auto_gates"),
    ("batch_chains", "/api/batch/chains"),
    ("chain_nifty", "/api/chain/NIFTY"),
    ("holdings", "/api/holdings"),
    ("funds", "/api/funds"),
    ("positions", "/api/positions"),
    ("signals_top", "/api/signal/top"),
    ("proof_ledger", "/api/proof_ledger"),
    ("continuous_closure", "/api/continuous_closure"),
    ("instruments_health", "/api/instruments/health"),
    ("live_trading_gate", "/api/live-trading/gate"),
    ("kill_switch", "/api/kill-switch/status"),
    ("status", "/api/status"),
    ("ui_root", "/ui/"),
]

# Frontend tab → primary backend API(s) for forensic mapping
TAB_API_MAP = {
    "decision-intel": ["/api/state", "/api/auto_gates"],
    "truth": ["/api/broker/status", "/api/deploy_info", "/api/health"],
    "genesis": ["/api/state", "/api/signal/top"],
    "e2e-proof": ["/api/proof_ledger", "/api/continuous_closure"],
    "overview": ["/api/state", "/api/health"],
    "sim-live": ["/api/state"],
    "options-intel": ["/api/batch/chains", "/api/chain/NIFTY"],
    "chain": ["/api/batch/chains", "/api/chain/NIFTY"],
    "signals": ["/api/signal/top"],
    "trade": ["/api/state"],
    "paper": ["/api/positions"],
    "positions": ["/api/positions", "/api/broker/positions/live"],
    "risk-scenarios": ["/api/state"],
    "multibagger": ["/api/research/multibagger"],
    "prediction-audit": ["/api/accuracy_trend", "/api/auto_gates"],
    "performance": ["/api/perf", "/api/pnl"],
    "ml": ["/api/auto_gates"],
    "data-integrity": ["/api/instruments/health", "/api/scheduler/health"],
    "broker": ["/api/broker/status", "/api/broker/truth", "/api/holdings", "/api/funds"],
    "alerts": ["/api/state"],
    "system": ["/api/system_health", "/api/deploy_info", "/api/memory"],
    "gates": ["/api/auto_gates", "/api/live-trading/gate"],
}

SENSITIVE_KEY = re.compile(
    r"(token|password|secret|pin|totp|authorization|api_key|access_key|private_key|credential)",
    re.I,
)
KEEP_META = re.compile(
    r"(secret_id|secret_version|version|expires|hours_remaining|present|classification|source|status)",
    re.I,
)


def gcloud_bin() -> str:
    for c in (
        r"C:\Program Files (x86)\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        r"C:\Program Files\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd",
        "gcloud",
    ):
        if c == "gcloud" or Path(c).exists():
            return c
    return "gcloud"


def run_gcloud(args: list[str], timeout: int = 90) -> tuple[Any, str | None]:
    exe = gcloud_bin()
    shell = exe.lower().endswith(".cmd")
    try:
        p = subprocess.run(
            [exe, *args],
            capture_output=True,
            shell=shell,
            timeout=timeout,
            check=False,
        )
    except Exception as exc:
        return None, str(exc)
    raw = p.stdout or b""
    err = (p.stderr or b"").decode("utf-8", "replace")
    if p.returncode != 0 and not raw:
        return None, err or f"gcloud rc={p.returncode}"
    for enc in ("utf-8", "utf-16", "utf-16-le"):
        try:
            text = raw.decode(enc).lstrip("\ufeff").strip()
            break
        except UnicodeDecodeError:
            text = ""
    else:
        text = raw.decode("utf-8", "replace").lstrip("\ufeff").strip()
    if not text:
        return None if p.returncode else {}, err or None
    try:
        return json.loads(text), None
    except json.JSONDecodeError:
        return {"raw": text[:4000]}, None


def redact(obj: Any, path: str = "") -> Any:
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            key = str(k)
            child = f"{path}.{key}" if path else key
            if SENSITIVE_KEY.search(key) and not KEEP_META.search(key):
                if isinstance(v, (str, bytes)) and v:
                    out[key] = f"<REDACTED len={len(v)}>"
                elif v is None or v is False or v is True or isinstance(v, (int, float)):
                    out[key] = v
                else:
                    out[key] = "<REDACTED>"
            else:
                out[key] = redact(v, child)
        return out
    if isinstance(obj, list):
        # Cap huge chain payloads
        if len(obj) > 40:
            return [redact(x, path) for x in obj[:20]] + [
                {"_truncated": True, "omitted": len(obj) - 20}
            ]
        return [redact(x, path) for x in obj]
    if isinstance(obj, str) and len(obj) > 500 and obj.startswith("eyJ"):
        return f"<REDACTED_JWT len={len(obj)}>"
    return obj


def http_get(path: str, timeout: int = 45) -> dict[str, Any]:
    url = path if path.startswith("http") else f"{BASE}{path}"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    started = datetime.now(timezone.utc)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read()
            elapsed_ms = int((datetime.now(timezone.utc) - started).total_seconds() * 1000)
            ctype = (resp.headers.get("Content-Type") or "").split(";")[0].strip()
            if "json" in ctype or path.startswith("/api/"):
                try:
                    data = json.loads(body.decode("utf-8", "replace"))
                except json.JSONDecodeError:
                    data = {"_non_json": True, "preview": body[:300].decode("utf-8", "replace")}
            else:
                text = body.decode("utf-8", "replace")
                data = {
                    "_html": True,
                    "bytes": len(body),
                    "has_root": "id=" in text or "root" in text.lower(),
                    "title_hint": _title_hint(text),
                    "preview": text[:240].replace("\n", " "),
                }
            return {
                "ok": True,
                "http_status": getattr(resp, "status", 200),
                "elapsed_ms": elapsed_ms,
                "content_type": ctype,
                "data": redact(data),
            }
    except urllib.error.HTTPError as exc:
        elapsed_ms = int((datetime.now(timezone.utc) - started).total_seconds() * 1000)
        body = exc.read()[:500].decode("utf-8", "replace")
        return {
            "ok": False,
            "http_status": exc.code,
            "elapsed_ms": elapsed_ms,
            "error": f"HTTPError {exc.code}",
            "body_preview": body,
        }
    except Exception as exc:
        elapsed_ms = int((datetime.now(timezone.utc) - started).total_seconds() * 1000)
        return {
            "ok": False,
            "http_status": None,
            "elapsed_ms": elapsed_ms,
            "error": f"{type(exc).__name__}: {exc}",
        }


def _title_hint(html: str) -> str:
    m = re.search(r"<title>(.*?)</title>", html, re.I | re.S)
    return (m.group(1).strip()[:120] if m else "")


def slim_chains(payload: Any) -> Any:
    if not isinstance(payload, dict):
        return payload
    chains = payload.get("chains") or {}
    slim_chains_map = {}
    if isinstance(chains, dict):
        for sym, item in chains.items():
            if not isinstance(item, dict):
                slim_chains_map[sym] = {"type": type(item).__name__}
                continue
            n = None
            if isinstance(item.get("contracts"), list):
                n = len(item["contracts"])
            slim_chains_map[sym] = {
                "status": item.get("status"),
                "stale": item.get("stale"),
                "fetched_at_utc": item.get("fetched_at_utc"),
                "n_contracts": n or item.get("contract_count"),
                "source": item.get("source") or item.get("source_class"),
            }
    return {
        "generated_at": payload.get("generated_at"),
        "required_symbols": payload.get("required_symbols"),
        "required_symbols_ready": payload.get("required_symbols_ready"),
        "symbols": payload.get("symbols"),
        "chains": slim_chains_map,
        "live_trading_enabled": payload.get("live_trading_enabled"),
    }


def collect_apis() -> dict[str, Any]:
    results: dict[str, Any] = {}
    for name, path in API_ENDPOINTS:
        print(f"API {path}", flush=True)
        result = http_get(path)
        if name == "batch_chains" and result.get("ok") and isinstance(result.get("data"), dict):
            result = dict(result)
            result["data"] = slim_chains(result["data"])
        if name.startswith("chain_") and result.get("ok") and isinstance(result.get("data"), dict):
            d = result["data"]
            contracts = d.get("contracts")
            result = dict(result)
            result["data"] = {
                "underlying": d.get("underlying") or d.get("symbol"),
                "status": d.get("status"),
                "stale": d.get("stale"),
                "fetched_at_utc": d.get("fetched_at_utc"),
                "n_contracts": len(contracts) if isinstance(contracts, list) else d.get("contract_count"),
                "source": d.get("source"),
            }
        results[name] = {"path": path, **result}
        # Persist per-endpoint
        API_DIR.mkdir(parents=True, exist_ok=True)
        (API_DIR / f"{name}.json").write_text(
            json.dumps(results[name], indent=2, default=str), encoding="utf-8"
        )
    return results


def collect_tabs() -> list[dict[str, Any]]:
    rows = []
    for tab_id, label, group in DASHBOARD_TABS:
        path = f"/ui/?tab={tab_id}"
        print(f"TAB {path}", flush=True)
        result = http_get(path, timeout=40)
        mapped = TAB_API_MAP.get(tab_id, [])
        rows.append(
            {
                "tab_id": tab_id,
                "label": label,
                "group": group,
                "url": f"{BASE}{path}",
                "http_ok": bool(result.get("ok")),
                "http_status": result.get("http_status"),
                "elapsed_ms": result.get("elapsed_ms"),
                "bytes": (result.get("data") or {}).get("bytes") if isinstance(result.get("data"), dict) else None,
                "has_root": (result.get("data") or {}).get("has_root") if isinstance(result.get("data"), dict) else None,
                "title_hint": (result.get("data") or {}).get("title_hint") if isinstance(result.get("data"), dict) else None,
                "primary_apis": mapped,
                "error": result.get("error"),
                "semantic_proof": "NOT_PROVEN",  # HTTP mount ≠ semantic values
            }
        )
    return rows


def collect_gcp() -> dict[str, Any]:
    out: dict[str, Any] = {"available": False, "errors": []}
    svc, err = run_gcloud(
        [
            "run",
            "services",
            "describe",
            SERVICE,
            f"--project={PROJ}",
            f"--region={REG}",
            "--format=json",
        ]
    )
    if err:
        out["errors"].append(f"service:{err[:300]}")
    else:
        out["available"] = True
        status = (svc or {}).get("status") or {}
        traffic = status.get("traffic") or []
        out["cloud_run"] = {
            "latestReadyRevisionName": status.get("latestReadyRevisionName"),
            "latestCreatedRevisionName": status.get("latestCreatedRevisionName"),
            "traffic": [
                {"revision": t.get("revisionName"), "percent": t.get("percent")}
                for t in traffic
                if isinstance(t, dict)
            ],
            "url": status.get("url"),
        }

    jobs, err = run_gcloud(
        ["run", "jobs", "list", f"--project={PROJ}", f"--region={REG}", "--format=json"]
    )
    if err:
        out["errors"].append(f"jobs:{err[:300]}")
    else:
        out["jobs"] = [
            {
                "name": (j.get("metadata") or {}).get("name"),
                "creationTimestamp": (j.get("metadata") or {}).get("creationTimestamp"),
            }
            for j in (jobs or [])
            if isinstance(j, dict)
        ][:50]

    sched, err = run_gcloud(
        ["scheduler", "jobs", "list", f"--project={PROJ}", f"--location={REG}", "--format=json"]
    )
    if err:
        # try global location fallback
        sched2, err2 = run_gcloud(
            ["scheduler", "jobs", "list", f"--project={PROJ}", "--location=asia-south1", "--format=json"]
        )
        if err2:
            out["errors"].append(f"scheduler:{err[:200]}|{err2[:200]}")
        else:
            sched = sched2
    if isinstance(sched, list):
        out["scheduler_jobs"] = [
            {
                "name": s.get("name", "").split("/")[-1],
                "schedule": s.get("schedule"),
                "timeZone": s.get("timeZone"),
                "state": s.get("state"),
                "httpTarget": bool(s.get("httpTarget")),
            }
            for s in sched
            if isinstance(s, dict)
        ]

    secrets, err = run_gcloud(
        ["secrets", "list", f"--project={PROJ}", "--format=json"]
    )
    if err:
        out["errors"].append(f"secrets:{err[:300]}")
    else:
        out["secret_names"] = [
            (s.get("name") or "").split("/")[-1]
            for s in (secrets or [])
            if isinstance(s, dict)
        ]

    return out


def git_main_sha() -> str | None:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "origin/main"],
            cwd=str(ROOT),
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip()
    except Exception:
        env_sha = os.environ.get("GITHUB_SHA")
        return env_sha


def sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HF
        cell.font = HFont
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        ws.append(row)
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(42, max(12, len(headers[i - 1]) + 4))


def build_workbook(
    apis: dict[str, Any],
    tabs: list[dict[str, Any]],
    gcp: dict[str, Any],
    cross: dict[str, Any],
    gaps: list[list[Any]],
) -> Workbook:
    wb = Workbook()
    # 00 README
    ws = wb.active
    ws.title = "00_Agent_README"
    lines = [
        ["SYSTEM3 LIVE PROOF CENTER"],
        ["generated_utc", NOW],
        ["public_base", BASE],
        ["project", PROJ],
        ["region", REG],
        ["service", SERVICE],
        ["purpose", "Sanitized live forensic pack for agents lacking laptop/gcloud access"],
        ["authority", "GitHub main + GCP serving — laptop is NON-AUTH"],
        ["secrets", "NEVER exported — names/versions/metadata only"],
        ["semantic_note", "HTTP 200 / HTML mount ≠ semantic API↔UI acceptance"],
        ["update_path", "reports/latest/live_proof_center/LATEST/"],
        ["workflow", ".github/workflows/live-proof-center.yml"],
        ["sheets", "00 README + 12 forensic data sheets"],
    ]
    for row in lines:
        ws.append(row)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90

    # 01 Executive
    dep = (apis.get("deploy_info") or {}).get("data") or {}
    br = (apis.get("broker_status") or {}).get("data") or {}
    health = (apis.get("health") or {}).get("data") or {}
    state = (apis.get("state") or {}).get("data") or {}
    sched = (apis.get("scheduler_health") or {}).get("data") or {}
    gates = (apis.get("auto_gates") or {}).get("data") or {}
    chains = (apis.get("batch_chains") or {}).get("data") or {}
    obs = (sched.get("observability") or {}) if isinstance(sched, dict) else {}
    sheet(
        wb,
        "01_Executive",
        ["field", "value"],
        [
            ["generated_utc", NOW],
            ["github_origin_main", cross.get("github_origin_main")],
            ["serving_git_sha", dep.get("git_sha")],
            ["sha_match_main_serving", cross.get("sha_match")],
            ["main_class", cross.get("main_class")],
            ["revision", ((gcp.get("cloud_run") or {}).get("latestReadyRevisionName"))],
            ["traffic", json.dumps((gcp.get("cloud_run") or {}).get("traffic"))],
            ["broker_connected", br.get("connected")],
            ["auth_classification", br.get("auth_classification")],
            ["live_trading_enabled", br.get("live_trading_enabled") or dep.get("live_trading_enabled")],
            ["orders_allowed", br.get("order_placement_allowed")],
            ["health_qc", health.get("qc_status")],
            ["state_qc", (state.get("qc") or {}).get("status") if isinstance(state.get("qc"), dict) else None],
            ["scheduler_healthy", sched.get("healthy")],
            ["alert_severity", obs.get("alert_severity")],
            ["unhealthy_reasons", json.dumps(sched.get("unhealthy_reasons"))],
            ["gates_passing", f"{gates.get('gates_passing')}/{gates.get('gates_total')}"],
            ["required_symbols_ready", chains.get("required_symbols_ready")],
            ["tabs_http_ok", sum(1 for t in tabs if t.get("http_ok"))],
            ["tabs_total", len(tabs)],
            ["gcp_inventory_available", gcp.get("available")],
            ["rhui_status_hint", "NOT_ACCEPTED until semantic+scheduler+#367 gates green"],
        ],
    )

    # 02 Deploy
    sheet(
        wb,
        "02_Deploy_Identity",
        ["key", "value"],
        [[k, json.dumps(v) if isinstance(v, (dict, list)) else v] for k, v in sorted(dep.items())]
        if isinstance(dep, dict)
        else [["error", "no deploy_info"]],
    )

    # 03 Broker
    token = br.get("token_reload") if isinstance(br, dict) else {}
    broker_rows = [
        ["connected", br.get("connected")],
        ["auth_classification", br.get("auth_classification")],
        ["mode", br.get("mode")],
        ["live_trading_enabled", br.get("live_trading_enabled")],
        ["order_placement_allowed", br.get("order_placement_allowed")],
        ["access_token_present", br.get("access_token_present")],
        ["client_id_present", br.get("client_id_present")],
        ["secret_version", (token or {}).get("secret_version") if isinstance(token, dict) else None],
        ["hours_remaining", (token or {}).get("hours_remaining") if isinstance(token, dict) else None],
        ["expires_at_utc", (token or {}).get("expires_at_utc") if isinstance(token, dict) else None],
    ]
    sheet(wb, "03_Broker_TokenMeta", ["field", "value"], broker_rows)

    # 04 QC
    sheet(
        wb,
        "04_Health_State_QC",
        ["source", "field", "value"],
        [
            ["health", "status", health.get("status")],
            ["health", "qc_status", health.get("qc_status")],
            ["health", "qc_failures", json.dumps(health.get("qc_failures"))],
            ["health", "broker_status", health.get("broker_status")],
            ["state", "qc_status", (state.get("qc") or {}).get("status") if isinstance(state.get("qc"), dict) else None],
            ["state", "qc_reasons", json.dumps((state.get("qc") or {}).get("reasons")) if isinstance(state.get("qc"), dict) else None],
            ["state", "market_is_open", (state.get("market") or {}).get("is_open") if isinstance(state.get("market"), dict) else None],
            ["convergence", "health_state_both_not_ready", str(health.get("qc_status")) == "NOT_READY" and ((state.get("qc") or {}).get("status") == "NOT_READY")],
        ],
    )

    # 05 Chains
    chain_rows = []
    for sym, item in (chains.get("chains") or {}).items() if isinstance(chains, dict) else []:
        if isinstance(item, dict):
            chain_rows.append(
                [
                    sym,
                    item.get("status"),
                    item.get("stale"),
                    item.get("n_contracts"),
                    item.get("fetched_at_utc"),
                    item.get("source"),
                ]
            )
    sheet(
        wb,
        "05_Option_Chains",
        ["symbol", "status", "stale", "n_contracts", "fetched_at_utc", "source"],
        chain_rows
        or [["required_symbols_ready", chains.get("required_symbols_ready"), "", "", "", ""]],
    )

    # 06 Scheduler
    sheet(
        wb,
        "06_Scheduler_Obs",
        ["field", "value"],
        [
            ["healthy", sched.get("healthy")],
            ["status", sched.get("status")],
            ["business_readiness", sched.get("business_readiness")],
            ["alert_severity", obs.get("alert_severity")],
            ["alert_signals", json.dumps(obs.get("alert_signals"))],
            ["contract_matched", obs.get("contract_matched")],
            ["collector_control_ok", obs.get("collector_control_ok")],
            ["unhealthy_reasons", json.dumps(sched.get("unhealthy_reasons"))],
            ["deploy_git_sha", sched.get("deploy_git_sha")],
        ],
    )

    # 07 Gates
    gate_map = gates.get("gates") if isinstance(gates, dict) else {}
    gate_rows = []
    if isinstance(gate_map, dict):
        for gid, g in gate_map.items():
            if isinstance(g, dict):
                gate_rows.append([gid, g.get("pass"), g.get("blocker_id"), json.dumps(g)[:200]])
    sheet(
        wb,
        "07_Auto_Gates",
        ["gate_id", "pass", "blocker_id", "snippet"],
        gate_rows
        or [
            [
                "summary",
                f"{gates.get('gates_passing')}/{gates.get('gates_total')}",
                json.dumps(gates.get("open_blockers")),
                "",
            ]
        ],
    )

    # 08 API matrix
    api_rows = []
    for name, meta in apis.items():
        api_rows.append(
            [
                name,
                meta.get("path"),
                meta.get("ok"),
                meta.get("http_status"),
                meta.get("elapsed_ms"),
                meta.get("error") or "",
            ]
        )
    sheet(
        wb,
        "08_API_Probe_Matrix",
        ["name", "path", "ok", "http_status", "elapsed_ms", "error"],
        api_rows,
    )

    # 09 Tabs
    sheet(
        wb,
        "09_Dashboard_Tabs_22",
        [
            "tab_id",
            "label",
            "group",
            "http_ok",
            "http_status",
            "elapsed_ms",
            "bytes",
            "has_root",
            "primary_apis",
            "semantic_proof",
        ],
        [
            [
                t["tab_id"],
                t["label"],
                t["group"],
                t["http_ok"],
                t["http_status"],
                t["elapsed_ms"],
                t.get("bytes"),
                t.get("has_root"),
                ",".join(t.get("primary_apis") or []),
                t.get("semantic_proof"),
            ]
            for t in tabs
        ],
    )

    # 10 FE/BE map
    map_rows = []
    for tab_id, label, group in DASHBOARD_TABS:
        map_rows.append([tab_id, label, group, ",".join(TAB_API_MAP.get(tab_id, []))])
    sheet(
        wb,
        "10_FE_BE_Map",
        ["tab_id", "label", "group", "primary_apis"],
        map_rows,
    )

    # 11 GCP inventory
    gcp_rows = [["available", gcp.get("available")], ["errors", json.dumps(gcp.get("errors"))]]
    cr = gcp.get("cloud_run") or {}
    for k, v in cr.items():
        gcp_rows.append([f"cloud_run.{k}", json.dumps(v) if isinstance(v, (dict, list)) else v])
    for j in gcp.get("jobs") or []:
        gcp_rows.append(["job", j.get("name")])
    for s in gcp.get("scheduler_jobs") or []:
        gcp_rows.append(
            [
                "scheduler",
                f"{s.get('name')} | {s.get('schedule')} | {s.get('state')} | {s.get('timeZone')}",
            ]
        )
    for name in gcp.get("secret_names") or []:
        gcp_rows.append(["secret_name", name])
    sheet(wb, "11_GCP_Inventory", ["field", "value"], gcp_rows)

    # 12 Gaps
    sheet(wb, "12_Gaps_Blockers", ["severity", "gap", "evidence", "next"], gaps)

    return wb


def build_gaps(
    apis: dict[str, Any],
    tabs: list[dict[str, Any]],
    gcp: dict[str, Any],
    cross: dict[str, Any],
) -> list[list[Any]]:
    gaps: list[list[Any]] = []
    if not cross.get("sha_match"):
        gaps.append(
            [
                "INFO",
                "main tip != serving SHA",
                f"main={cross.get('github_origin_main')} serving={cross.get('serving_sha')}",
                "Classify docs-only vs runtime; do not blind redeploy",
            ]
        )
    failed_apis = [n for n, m in apis.items() if not m.get("ok")]
    if failed_apis:
        gaps.append(["HIGH", "API probe failures", ",".join(failed_apis), "Inspect api/*.json"])
    failed_tabs = [t["tab_id"] for t in tabs if not t.get("http_ok")]
    if failed_tabs:
        gaps.append(["HIGH", "Dashboard tab HTTP failures", ",".join(failed_tabs), "Check /ui routing"])
    sched = (apis.get("scheduler_health") or {}).get("data") or {}
    if isinstance(sched, dict) and sched.get("healthy") is False:
        gaps.append(
            [
                "CRITICAL",
                "Scheduler unhealthy",
                json.dumps(sched.get("unhealthy_reasons")),
                "Attribute workload; no token mint",
            ]
        )
    chains = (apis.get("batch_chains") or {}).get("data") or {}
    if isinstance(chains, dict) and chains.get("required_symbols_ready") is not True:
        gaps.append(
            [
                "HIGH",
                "4/4 chains not ready",
                str(chains.get("required_symbols_ready")),
                "Market-open recheck",
            ]
        )
    gaps.append(
        [
            "MODERATE",
            "22-tab semantic API↔UI",
            "HTTP mounts only",
            "Run semantic continuous session",
        ]
    )
    if not gcp.get("available"):
        gaps.append(
            [
                "INFO",
                "GCP inventory unavailable this run",
                json.dumps(gcp.get("errors")),
                "Workflow WIF should fill sheet 11",
            ]
        )
    return gaps


def write_index(cross: dict[str, Any], apis: dict[str, Any], tabs: list[dict[str, Any]], gcp: dict[str, Any]) -> str:
    dep = (apis.get("deploy_info") or {}).get("data") or {}
    br = (apis.get("broker_status") or {}).get("data") or {}
    sched = (apis.get("scheduler_health") or {}).get("data") or {}
    gates = (apis.get("auto_gates") or {}).get("data") or {}
    chains = (apis.get("batch_chains") or {}).get("data") or {}
    obs = (sched.get("observability") or {}) if isinstance(sched, dict) else {}
    ok_tabs = sum(1 for t in tabs if t.get("http_ok"))
    ok_apis = sum(1 for m in apis.values() if m.get("ok"))
    md = f"""# SYSTEM3 Live Proof Center

**Generated (UTC):** `{NOW}`  
**Public base:** {BASE}  
**Audience:** ChatGPT / Claude / Cursor / Codex — **no laptop gcloud required**

## Authority

1. GitHub `origin/main` tip (code history)
2. GCP Cloud Run `/api/deploy_info` (what is serving now)
3. This folder — continuously refreshed sanitized forensic pack

Laptop checkouts are **NON-AUTHORITATIVE**.

## Cross-verify snapshot

| Field | Value |
|---|---|
| GitHub main | `{cross.get("github_origin_main")}` |
| Serving SHA | `{dep.get("git_sha")}` |
| Match | `{cross.get("sha_match")}` |
| Class | `{cross.get("main_class")}` |
| Revision | `{(gcp.get("cloud_run") or {}).get("latestReadyRevisionName")}` |
| Broker | connected=`{br.get("connected")}` auth=`{br.get("auth_classification")}` LIVE=`{br.get("live_trading_enabled")}` |
| Scheduler healthy | `{sched.get("healthy")}` severity=`{obs.get("alert_severity")}` |
| Gates | `{gates.get("gates_passing")}/{gates.get("gates_total")}` |
| 4/4 chains ready | `{chains.get("required_symbols_ready")}` |
| API probes OK | {ok_apis}/{len(apis)} |
| UI tabs HTTP OK | {ok_tabs}/{len(tabs)} (mount only — semantic NOT_PROVEN) |

## Files in this pack

| File | Purpose |
|---|---|
| `INDEX.md` | This landing page |
| `MANIFEST.json` | Machine SSOT |
| `System3_LIVE_PROOF_CENTER.xlsx` | **12 forensic sheets** (Excel MRI) |
| `CROSS_VERIFY.json` | main vs serving classification |
| `dashboard_tabs.json` | All 22 tab HTTP forensics |
| `gcp_inventory.json` | Cloud Run / jobs / scheduler / secret **names** |
| `api/*.json` | Per-endpoint sanitized dumps |

## Excel sheets (12)

1. `00_Agent_README`
2. `01_Executive`
3. `02_Deploy_Identity`
4. `03_Broker_TokenMeta`
5. `04_Health_State_QC`
6. `05_Option_Chains`
7. `06_Scheduler_Obs`
8. `07_Auto_Gates`
9. `08_API_Probe_Matrix`
10. `09_Dashboard_Tabs_22`
11. `10_FE_BE_Map`
12. `11_GCP_Inventory` + `12_Gaps_Blockers`

## Safety

- No secret **values**, no LIVE enablement, no order calls
- Workflow: `.github/workflows/live-proof-center.yml` (schedule + manual)
- Coordination bus: Issue #188

## How agents must use this

1. Read `MANIFEST.json` + `INDEX.md` first every session  
2. Do not invent PASS from stale laptop MRI  
3. HTTP tab OK ≠ semantic acceptance  
4. If GCP sheet empty, trust public API sheets; ask Cursor only for WIF failures  
"""
    return md


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    API_DIR.mkdir(parents=True, exist_ok=True)

    print("SYSTEM3 Live Proof Center — collecting public APIs...", flush=True)
    apis = collect_apis()
    print("Collecting dashboard tabs...", flush=True)
    tabs = collect_tabs()
    print("Collecting GCP inventory (optional)...", flush=True)
    gcp = collect_gcp()

    serving = ((apis.get("deploy_info") or {}).get("data") or {}).get("git_sha")
    main_sha = git_main_sha()
    sha_match = bool(main_sha and serving and str(main_sha)[:40] == str(serving)[:40])
    cross = {
        "generated_utc": NOW,
        "github_origin_main": main_sha,
        "serving_sha": serving,
        "sha_match": sha_match,
        "main_class": "MATCH" if sha_match else "DIVERGED_CHECK_PATH_FILTER",
        "public_base": BASE,
        "github_run_id": os.environ.get("GITHUB_RUN_ID"),
        "github_run_url": (
            f"https://github.com/{os.environ.get('GITHUB_REPOSITORY')}/actions/runs/{os.environ.get('GITHUB_RUN_ID')}"
            if os.environ.get("GITHUB_RUN_ID") and os.environ.get("GITHUB_REPOSITORY")
            else None
        ),
    }
    (OUT / "CROSS_VERIFY.json").write_text(json.dumps(cross, indent=2), encoding="utf-8")
    (OUT / "dashboard_tabs.json").write_text(json.dumps(tabs, indent=2), encoding="utf-8")
    (OUT / "gcp_inventory.json").write_text(json.dumps(gcp, indent=2), encoding="utf-8")

    gaps = build_gaps(apis, tabs, gcp, cross)
    wb = build_workbook(apis, tabs, gcp, cross, gaps)
    xlsx = OUT / "System3_LIVE_PROOF_CENTER.xlsx"
    wb.save(xlsx)
    # coordination copy
    coord = ROOT / "reports" / "coordination"
    coord.mkdir(parents=True, exist_ok=True)
    try:
        (coord / "System3_LIVE_PROOF_CENTER.xlsx").write_bytes(xlsx.read_bytes())
    except Exception as exc:
        print(f"WARN coord copy: {exc}", flush=True)

    index = write_index(cross, apis, tabs, gcp)
    (OUT / "INDEX.md").write_text(index, encoding="utf-8")
    # Stable pointer at pack root
    pack_root = ROOT / "reports" / "latest" / "live_proof_center"
    pack_root.mkdir(parents=True, exist_ok=True)
    (pack_root / "README.md").write_text(
        f"# Live Proof Center\n\nOpen **[LATEST/INDEX.md](LATEST/INDEX.md)** (updated `{NOW}`).\n\n"
        "Scheduled by `.github/workflows/live-proof-center.yml`.\n",
        encoding="utf-8",
    )
    ptr = ROOT / "reports" / "coordination" / "LIVE_PROOF_CENTER_POINTER.md"
    ptr.write_text(
        "\n".join(
            [
                "# Live Proof Center pointer",
                "",
                f"**Updated UTC:** `{NOW}`",
                "",
                "- Pack: [`../latest/live_proof_center/LATEST/INDEX.md`](../latest/live_proof_center/LATEST/INDEX.md)",
                "- Excel: [`../latest/live_proof_center/LATEST/System3_LIVE_PROOF_CENTER.xlsx`](../latest/live_proof_center/LATEST/System3_LIVE_PROOF_CENTER.xlsx)",
                "- Manifest: [`../latest/live_proof_center/LATEST/MANIFEST.json`](../latest/live_proof_center/LATEST/MANIFEST.json)",
                "- Coordination copy: [`System3_LIVE_PROOF_CENTER.xlsx`](System3_LIVE_PROOF_CENTER.xlsx)",
                "",
                "Agents blocked by laptop/gcloud access: read this pointer first, then INDEX.md.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    manifest = {
        "schema": "system3_live_proof_center_v1",
        "generated_utc": NOW,
        "stamp": STAMP,
        "public_base": BASE,
        "cross_verify": cross,
        "api_ok": sum(1 for m in apis.values() if m.get("ok")),
        "api_total": len(apis),
        "tabs_http_ok": sum(1 for t in tabs if t.get("http_ok")),
        "tabs_total": len(tabs),
        "gcp_available": bool(gcp.get("available")),
        "files": [
            "INDEX.md",
            "MANIFEST.json",
            "System3_LIVE_PROOF_CENTER.xlsx",
            "CROSS_VERIFY.json",
            "dashboard_tabs.json",
            "gcp_inventory.json",
            "api/",
        ],
        "gaps": [{"severity": g[0], "gap": g[1], "evidence": g[2], "next": g[3]} for g in gaps],
    }
    (OUT / "MANIFEST.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(OUT), **{k: manifest[k] for k in ("api_ok", "api_total", "tabs_http_ok", "tabs_total", "gcp_available")}}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
