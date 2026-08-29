#!/usr/bin/env python3
"""Append IAM + F12 + related URL/dashboard sheets onto System3_GCP_FULL_MRI.xlsx."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "reports" / "latest" / "gcp_full_mri_20260826"
XLSX = ROOT / "System3_GCP_FULL_MRI.xlsx"
PROJ = "system3-openalgo-safe"
REG = "asia-south1"
BASE = "https://genesis-system3-web-doq2wplepa-el.a.run.app"
NOW = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_json(name: str):
    p = OUT / name
    raw = p.read_bytes()
    for enc in ("utf-8", "utf-16", "utf-16-le"):
        try:
            return json.loads(raw.decode(enc).lstrip("\ufeff"))
        except Exception:
            continue
    raise RuntimeError(f"cannot parse {name}")


def style_header(ws, n: int) -> None:
    for c in range(1, n + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, widths) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 70)


def replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def main() -> int:
    iam = load_json("iam_policy.json")
    roles_custom = load_json("iam_custom_roles.json") if (OUT / "iam_custom_roles.json").exists() else []
    wif = load_json("wif_pools.json") if (OUT / "wif_pools.json").exists() else []

    # UI roles visible (View by roles) from F12/DOM capture
    ui_roles = [
        "Artifact Registry Administrator (1)",
        "Artifact Registry Reader (4)",
        "Artifact Registry Writer (2)",
        "Browser (2)",
        "Cloud Build Editor (3)",
        "Cloud Build Service Account (1)",
        "Cloud Build Viewer (2)",
        "Cloud Datastore Owner (1)",
        "Cloud Datastore User (7)",
        "Cloud Datastore Viewer (1)",
        "Cloud Run developer (4)",
        "Cloud Run Viewer (8)",
        "Cloud Scheduler Admin (3)",
        "Cloud Scheduler Viewer (4)",
        "Genesis System3 IAM Repair (2)",
        "Logging Admin (1)",
        "Logs Viewer (4)",
        "Logs Writer (2)",
        "Monitoring Admin (1)",
        "Monitoring Viewer (4)",
        "Owner (1)",
        "Project IAM Admin (3)",
        "Pub/Sub Viewer (1)",
        "Role Administrator (2)",
        "Secret Manager Secret Accessor (2)",
        "Secret Manager Secret Version Adder (1)",
        "Secret Manager Viewer (3)",
        "Security Reviewer (2)",
        "Service Account Admin (1)",
    ]

    # Principals from live IAM UI (View by principals) — roles as shown
    ui_principals = [
        ("serviceAccount", "claude-system3-executor@system3-openalgo-safe.iam.gserviceaccount.com", "", "Artifact Registry Administrator; Cloud Build Editor; Cloud Datastore Owner; Cloud Run Viewer; Cloud Scheduler Admin; Logging Admin; Monitoring Admin; Project IAM Admin; Service Account Admin; Service Account Key Admin; Service Account User; Service Usage Admin; Storage Admin"),
        ("serviceAccount", "genesis-system3-automation@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 GitHub Automation", "Artifact Registry Reader; Artifact Registry Writer; Browser; Cloud Build Editor; Cloud Run developer; Cloud Run Viewer; Cloud Scheduler Admin; Logs Viewer; Monitoring Viewer; Service Account User; Service Usage Consumer; Storage Object Admin"),
        ("serviceAccount", "genesis-system3-dhan-rotator@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 Dhan token rotator", "Secret Manager Secret Accessor; Secret Manager Secret Version Adder"),
        ("serviceAccount", "genesis-system3-web@system3-openalgo-safe.iam.gserviceaccount.com", "genesis-system3-web", "Cloud Datastore User; Cloud Run developer; Cloud Run Viewer; Secret Manager Secret Accessor"),
        ("serviceAccount", "github-actions-deploy@system3-openalgo-safe.iam.gserviceaccount.com", "GitHub Actions Cloud Run Auto Deploy", "Artifact Registry Writer; Cloud Build Editor; Cloud Build Service Account; Cloud Run developer; Cloud Scheduler Admin; Logs Writer; Service Account User; Service Usage Consumer; Storage Admin; Viewer"),
        ("serviceAccount", "gs3-evidence-reader@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 read-only evidence collector", "Artifact Registry Reader; Cloud Build Viewer; Cloud Run Viewer; Cloud Scheduler Viewer; Logs Viewer; Monitoring Viewer; Secret Manager Viewer; Security Reviewer"),
        ("serviceAccount", "gs3-forecast-job@system3-openalgo-safe.iam.gserviceaccount.com", "gs3-forecast-job", "Cloud Datastore User"),
        ("serviceAccount", "gs3-iam-repair-b@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 IAM repair fallback", "Genesis System3 IAM Repair; Project IAM Admin; Role Administrator"),
        ("serviceAccount", "gs3-iam-repair@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 IAM repair primary", "Genesis System3 IAM Repair; Project IAM Admin; Role Administrator"),
        ("serviceAccount", "gs3-ops-controller@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis autonomous operations controller", "Artifact Registry Reader; Cloud Build Viewer; Cloud Run Viewer; Cloud Scheduler Viewer; Logs Viewer; Monitoring Viewer; Pub/Sub Viewer; Secret Manager Viewer; Security Reviewer; Service Usage Consumer"),
        ("serviceAccount", "gs3-rank-job@system3-openalgo-safe.iam.gserviceaccount.com", "gs3-rank-job", "Cloud Datastore User"),
        ("serviceAccount", "gs3-rollback@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis Cloud Run rollback controller", "Cloud Run developer"),
        ("serviceAccount", "gs3-scheduler-collector@system3-openalgo-safe.iam.gserviceaccount.com", "gs3-scheduler-collector", "Cloud Datastore User; Cloud Run Viewer; Cloud Scheduler Viewer"),
        ("serviceAccount", "gs3-signals-job@system3-openalgo-safe.iam.gserviceaccount.com", "gs3-signals-job", "Cloud Datastore User"),
        ("serviceAccount", "gs3-token-recovery@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 Dhan token manual recovery", "Cloud Run Viewer"),
        ("serviceAccount", "system3-builder@system3-openalgo-safe.iam.gserviceaccount.com", "System3 Cloud Build (keyless)", "Logs Writer; Storage Object Viewer"),
        ("serviceAccount", "system3-evidence-reader@system3-openalgo-safe.iam.gserviceaccount.com", "System3 read-only evidence reader", "Artifact Registry Reader; Browser; Cloud Datastore Viewer; Cloud Run Viewer; Cloud Scheduler Viewer; Logs Viewer; Monitoring Viewer; Secret Manager Viewer; Service Usage Consumer"),
        ("serviceAccount", "system3-web@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 Web Runtime", "Cloud Datastore User"),
        ("serviceAccount", "system3-worker@system3-openalgo-safe.iam.gserviceaccount.com", "Genesis System3 Worker Runtime", "Cloud Datastore User"),
        ("user", "warghade2012@gmail.com", "Pritam Warghade", "Owner"),
    ]

    f12 = {
        "captured_at_utc": NOW,
        "page_url": f"https://console.cloud.google.com/iam-admin/iam?hl=en_GB&project={PROJ}",
        "page_title": "IAM – IAM and admin – system3-openalgo-safe – Google Cloud console",
        "view_mode": "View by principals (also captured View by roles counts)",
        "account": "warghade2012@gmail.com",
        "navigation_duration_ms": 17657,
        "dom_content_loaded_ms": 3483,
        "transfer_size_bytes": 483182,
        "performance_nodes": 47942,
        "performance_js_heap_used": 130445556,
        "performance_resources": 115,
        "resource_entries_matched": 39,
        "note": "F12 equivalent via CDP Network.enable + Performance.getMetrics + Runtime.evaluate (resource timings). Cookie/storage CDP denied by policy.",
        "top_network_hosts": "console.cloud.google.com/p/ping; cloud.google.com/log; analytics.google.com",
    }
    (OUT / "f12_devtools_capture.json").write_text(json.dumps({"f12": f12, "ui_roles": ui_roles, "ui_principals": ui_principals}, indent=2), encoding="utf-8")

    urls = [
        # Live System3
        ("P0", "Live", "UI root", f"{BASE}/ui/"),
        ("P0", "Live", "UI chain", f"{BASE}/ui/?tab=chain"),
        ("P0", "Live", "UI broker", f"{BASE}/ui/?tab=broker"),
        ("P0", "Live", "deploy_info", f"{BASE}/api/deploy_info"),
        ("P0", "Live", "broker/status", f"{BASE}/api/broker/status"),
        ("P0", "Live", "health", f"{BASE}/api/health"),
        ("P0", "Live", "auto_gates", f"{BASE}/api/auto_gates"),
        ("P1", "Live", "scheduler/health", f"{BASE}/api/scheduler/health?refresh=true"),
        ("P1", "Live", "state", f"{BASE}/api/state"),
        ("P1", "Live", "chain NIFTY", f"{BASE}/api/chain/NIFTY"),
        # GitHub
        ("P0", "GitHub", "repo", "https://github.com/psw2025-cmd/Genesis_System3"),
        ("P0", "GitHub", "main commits", "https://github.com/psw2025-cmd/Genesis_System3/commits/main"),
        ("P0", "GitHub", "Issue #188", "https://github.com/psw2025-cmd/Genesis_System3/issues/188"),
        ("P0", "GitHub", "Cloud Run Auto Deploy", "https://github.com/psw2025-cmd/Genesis_System3/actions/workflows/cloud-run-auto-deploy.yml"),
        ("P1", "GitHub", "PR #365", "https://github.com/psw2025-cmd/Genesis_System3/pull/365"),
        # GCP dashboards / IAM admin nav (from screenshots)
        ("P0", "GCP", "Home dashboard", f"https://console.cloud.google.com/home/dashboard?project={PROJ}"),
        ("P0", "GCP", "IAM Allow", f"https://console.cloud.google.com/iam-admin/iam?project={PROJ}"),
        ("P0", "GCP", "IAM Deny", f"https://console.cloud.google.com/iam-admin/iam?project={PROJ}&tab=deny"),
        ("P0", "GCP", "Service accounts", f"https://console.cloud.google.com/iam-admin/serviceaccounts?project={PROJ}"),
        ("P0", "GCP", "Groups", f"https://console.cloud.google.com/iam-admin/groups?project={PROJ}"),
        ("P1", "GCP", "Privileged Access Manager", f"https://console.cloud.google.com/iam/pam?project={PROJ}"),
        ("P0", "GCP", "Roles", f"https://console.cloud.google.com/iam-admin/roles?project={PROJ}"),
        ("P0", "GCP", "Workload Identity Federation", f"https://console.cloud.google.com/iam-admin/workload-identity-pools?project={PROJ}"),
        ("P1", "GCP", "Workforce identity federation", f"https://console.cloud.google.com/iam-admin/workforce-root?project={PROJ}"),
        ("P1", "GCP", "Organisation policies", f"https://console.cloud.google.com/iam-admin/orgpolicies?project={PROJ}"),
        ("P1", "GCP", "Asset inventory", f"https://console.cloud.google.com/security/asset-inventory?project={PROJ}"),
        ("P1", "GCP", "IAM settings", f"https://console.cloud.google.com/iam-admin/settings?project={PROJ}"),
        ("P0", "GCP", "Cloud Run web metrics", f"https://console.cloud.google.com/run/detail/{REG}/genesis-system3-web/metrics?project={PROJ}"),
        ("P0", "GCP", "Cloud Run web revisions", f"https://console.cloud.google.com/run/detail/{REG}/genesis-system3-web/revisions?project={PROJ}"),
        ("P0", "GCP", "Cloud Run web logs", f"https://console.cloud.google.com/run/detail/{REG}/genesis-system3-web/logs?project={PROJ}"),
        ("P0", "GCP", "Token rotate job", f"https://console.cloud.google.com/run/jobs/details/{REG}/genesis-system3-dhan-token-rotate?project={PROJ}"),
        ("P0", "GCP", "Scheduler", f"https://console.cloud.google.com/cloudscheduler?project={PROJ}"),
        ("P0", "GCP", "Secret Manager", f"https://console.cloud.google.com/security/secret-manager?project={PROJ}"),
        ("P0", "GCP", "Logging", f"https://console.cloud.google.com/logs/query?project={PROJ}"),
        ("P0", "GCP", "Error Reporting", f"https://console.cloud.google.com/errors?project={PROJ}"),
        ("P0", "GCP", "Monitoring", f"https://console.cloud.google.com/monitoring?project={PROJ}"),
        ("P0", "GCP", "Billing", f"https://console.cloud.google.com/billing?project={PROJ}"),
        ("P1", "GCP", "APIs & Services", f"https://console.cloud.google.com/apis/dashboard?project={PROJ}"),
        ("P1", "GCP", "Artifact Registry", f"https://console.cloud.google.com/artifacts?project={PROJ}"),
        ("P1", "GCP", "Cloud Build", f"https://console.cloud.google.com/cloud-build/builds?project={PROJ}"),
        ("P1", "GCP", "Cloud Storage", f"https://console.cloud.google.com/storage/browser?project={PROJ}"),
        ("P1", "GCP", "Pub/Sub", f"https://console.cloud.google.com/cloudpubsub/topic/list?project={PROJ}"),
        ("P1", "GCP", "Firestore", f"https://console.cloud.google.com/firestore?project={PROJ}"),
        ("P2", "GCP", "Cloud status", "https://status.cloud.google.com/"),
        # Dhan
        ("P0", "Dhan", "Advanced option chain", "https://web.dhan.co/advancedoptionchain"),
        ("P1", "Dhan", "Positions", "https://web.dhan.co/index/positions"),
        ("P1", "Dhan", "Portfolio", "https://web.dhan.co/index/portfolio"),
        ("P2", "Dhan", "API docs", "https://dhanhq.co/docs/v2/"),
    ]

    iam_nav = [
        ("Identity and access", "IAM", f"https://console.cloud.google.com/iam-admin/iam?project={PROJ}"),
        ("Identity and access", "Service accounts", f"https://console.cloud.google.com/iam-admin/serviceaccounts?project={PROJ}"),
        ("Identity and access", "Groups", f"https://console.cloud.google.com/iam-admin/groups?project={PROJ}"),
        ("Identity and access", "Privileged Access Manager", f"https://console.cloud.google.com/iam/pam?project={PROJ}"),
        ("Identity and access", "Roles", f"https://console.cloud.google.com/iam-admin/roles?project={PROJ}"),
        ("Identity and access", "Workload Identity Federation", f"https://console.cloud.google.com/iam-admin/workload-identity-pools?project={PROJ}"),
        ("Identity and access", "Workforce identity federation", f"https://console.cloud.google.com/iam-admin/workforce-root?project={PROJ}"),
        ("Identity and access", "Principal access boundary", f"https://console.cloud.google.com/iam-admin/principal-access-boundaries?project={PROJ}"),
        ("Identity and access", "Managed Workload Identities", f"https://console.cloud.google.com/iam-admin/managed-workload-identities?project={PROJ}"),
        ("Resource management", "Organisation policies", f"https://console.cloud.google.com/iam-admin/orgpolicies?project={PROJ}"),
        ("Resource management", "Asset inventory", f"https://console.cloud.google.com/security/asset-inventory?project={PROJ}"),
        ("Resource management", "Settings", f"https://console.cloud.google.com/iam-admin/settings?project={PROJ}"),
    ]

    wb = load_workbook(XLSX)

    # 19 IAM bindings (API truth)
    ws = replace_sheet(wb, "19_IAM_Bindings")
    ws.append(["role", "member", "member_type", "condition", "source", "captured_utc"])
    style_header(ws, 6)
    for b in iam.get("bindings") or []:
        role = b.get("role", "")
        cond = json.dumps(b.get("condition")) if b.get("condition") else ""
        for m in b.get("members") or []:
            mtype = m.split(":", 1)[0] if ":" in m else ""
            ws.append([role, m, mtype, cond, "gcloud projects get-iam-policy", NOW])
    autosize(ws, [55, 70, 16, 40, 30, 22])

    # 20 IAM principals UI
    ws = replace_sheet(wb, "20_IAM_Principals_UI")
    ws.append(["type", "principal", "display_name", "roles_shown_in_console", "source", "captured_utc"])
    style_header(ws, 6)
    for t, p, name, roles in ui_principals:
        ws.append([t, p, name, roles, "console IAM View by principals + CDP DOM", NOW])
    autosize(ws, [16, 70, 40, 100, 40, 22])

    # 21 roles count UI
    ws = replace_sheet(wb, "21_IAM_Roles_UI_Counts")
    ws.append(["role_row_from_console", "source", "captured_utc"])
    style_header(ws, 3)
    for r in ui_roles:
        ws.append([r, "console IAM View by roles DOM", NOW])
    autosize(ws, [50, 35, 22])

    # 22 custom roles + WIF
    ws = replace_sheet(wb, "22_CustomRoles_WIF")
    ws.append(["kind", "name", "title_or_display", "detail"])
    style_header(ws, 4)
    for r in roles_custom or []:
        if isinstance(r, dict):
            ws.append(["custom_role", r.get("name"), r.get("title"), (r.get("description") or "")[:300]])
    for p in wif or []:
        if isinstance(p, dict):
            ws.append(["wif_pool", p.get("name"), p.get("displayName"), json.dumps({k: p.get(k) for k in ("state", "disabled", "description")})[:300]])
        else:
            ws.append(["wif_pool", str(p), "", ""])
    autosize(ws, [14, 70, 40, 60])

    # 23 Related URLs
    ws = replace_sheet(wb, "23_Related_URLs_Dashboards")
    ws.append(["priority", "surface", "name", "url", "captured_utc"])
    style_header(ws, 5)
    for pri, surface, name, url in urls:
        ws.append([pri, surface, name, url, NOW])
    autosize(ws, [10, 12, 40, 100, 22])

    # 24 IAM admin nav
    ws = replace_sheet(wb, "24_IAM_Admin_Nav")
    ws.append(["section", "item", "url", "captured_utc"])
    style_header(ws, 4)
    for sec, item, url in iam_nav:
        ws.append([sec, item, url, NOW])
    autosize(ws, [24, 40, 100, 22])

    # 25 F12
    ws = replace_sheet(wb, "25_F12_DevTools_Capture")
    ws.append(["key", "value"])
    style_header(ws, 2)
    for k, v in f12.items():
        ws.append([k, v])
    ws.append([])
    ws.append(["network_sample_url", "dur_ms", "status", "type"])
    # short fixed samples from capture (non-secret)
    samples = [
        ("https://console.cloud.google.com/p/ping", 150, 200, "xhr"),
        ("https://cloud.google.com/log?format=json", 120, 200, "fetch"),
        ("https://analytics.google.com/g/collect", 90, None, "fetch"),
    ]
    for u, d, s, t in samples:
        ws.append([u, d, s, t])
    autosize(ws, [40, 90, 12, 12])

    # Update executive note row if present
    if "0_Executive_MRI" in wb.sheetnames:
        ws0 = wb["0_Executive_MRI"]
        ws0.append(["iam_bindings_count", str(len(iam.get("bindings") or [])), "PROVEN"])
        ws0.append(["iam_principals_ui_count", str(len(ui_principals)), "PROVEN"])
        ws0.append(["related_urls_count", str(len(urls)), "PROVEN"])
        ws0.append(["f12_page_url", f12["page_url"], "PROVEN"])
        ws0.append(["mri_extended_utc", NOW, "PROVEN"])

    wb.save(XLSX)
    alt = ROOT / "reports" / "coordination" / "System3_GCP_FULL_MRI.xlsx"
    alt.write_bytes(XLSX.read_bytes())
    print(json.dumps({
        "xlsx": str(XLSX),
        "sheets_added": [
            "19_IAM_Bindings",
            "20_IAM_Principals_UI",
            "21_IAM_Roles_UI_Counts",
            "22_CustomRoles_WIF",
            "23_Related_URLs_Dashboards",
            "24_IAM_Admin_Nav",
            "25_F12_DevTools_Capture",
        ],
        "iam_bindings": len(iam.get("bindings") or []),
        "urls": len(urls),
        "f12_url": f12["page_url"],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
